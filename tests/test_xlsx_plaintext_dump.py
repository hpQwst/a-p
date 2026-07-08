from __future__ import annotations

from io import BytesIO
import unittest

import openpyxl

from ppt_automator.xlsx_plaintext_dump import dump_xlsx_workbook


class XlsxPlaintextDumpTests(unittest.TestCase):
    def test_dump_preserves_decimal_formula_number_format_and_merges(self) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Planilha1"
        ws["A1"] = "Titulo"
        ws.merge_cells("A1:B1")
        ws["D15"] = 45.8419
        ws["D15"].number_format = "0.0000"
        ws["E15"] = "=D15*2"
        stream = BytesIO()
        wb.save(stream)

        dump = dump_xlsx_workbook(stream.getvalue(), file_name="slide_006/teste.xlsx")
        sheet = dump.sheets[0]
        by_cell = {cell.cell: cell for cell in sheet.cells}

        self.assertIn("A1:B1", sheet.merged_cells)
        self.assertEqual(by_cell["A1"].merged_range, "A1:B1")
        self.assertEqual(by_cell["D15"].raw, "45.8419")
        self.assertEqual(by_cell["D15"].number_format, "0.0000")
        self.assertEqual(by_cell["E15"].formula, "=D15*2")

    def test_prompt_text_is_compact_by_default(self) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Planilha1"
        ws["A1"] = "Titulo"
        ws.merge_cells("A1:B1")
        ws["B2"] = "Total"
        ws["C2"] = 12.345
        ws["C2"].number_format = "0.0"
        ws["D2"] = "=C2*2"
        stream = BytesIO()
        wb.save(stream)

        dump = dump_xlsx_workbook(stream.getvalue(), file_name="slide_001/teste.xlsx")
        prompt = dump.as_prompt_text()

        self.assertIn("WORKBOOK slide_001/teste.xlsx", prompt)
        self.assertIn("ROW 1:", prompt)
        self.assertIn('"c":"A1"', prompt)
        self.assertIn('"m":"A1:B1"', prompt)
        self.assertIn('"fmt":"0.0"', prompt)
        self.assertIn('"f":"=C2*2"', prompt)
        self.assertNotIn('"style"', prompt)
        self.assertNotIn('"number_format"', prompt)


if __name__ == "__main__":
    unittest.main()
