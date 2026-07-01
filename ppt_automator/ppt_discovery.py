from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from typing import Any, BinaryIO, Iterable
from zipfile import ZipFile
import os
import posixpath
import re
import unicodedata
import xml.etree.ElementTree as ET

import openpyxl
from openpyxl.utils.cell import range_boundaries


InputFile = str | Path | bytes | bytearray | BinaryIO

PML_NS = "http://schemas.openxmlformats.org/presentationml/2006/main"
DML_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
CHART_NS = "http://schemas.openxmlformats.org/drawingml/2006/chart"
REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
R_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
MC_NS = "http://schemas.openxmlformats.org/markup-compatibility/2006"
C14_NS = "http://schemas.microsoft.com/office/drawing/2007/8/2/chart"
C15_NS = "http://schemas.microsoft.com/office/drawing/2012/chart"
C16_NS = "http://schemas.microsoft.com/office/drawing/2014/chart"
C16R2_NS = "http://schemas.microsoft.com/office/drawing/2015/06/chart"
C16R3_NS = "http://schemas.microsoft.com/office/drawing/2017/03/chart"

NS = {
    "p": PML_NS,
    "a": DML_NS,
    "c": CHART_NS,
    "rel": REL_NS,
    "r": R_NS,
}

for prefix, uri in {
    "p": PML_NS,
    "a": DML_NS,
    "c": CHART_NS,
    "r": R_NS,
    "mc": MC_NS,
    "c14": C14_NS,
    "c15": C15_NS,
    "c16": C16_NS,
    "c16r2": C16R2_NS,
    "c16r3": C16R3_NS,
}.items():
    ET.register_namespace(prefix, uri)


@dataclass(frozen=True)
class PptTarget:
    slide_index: int
    slide_number: int
    slide_path: str
    shape_name: str
    shape_id: str
    object_type: str
    left_in: float
    top_in: float
    width_in: float
    height_in: float
    nearby_text: str = ""
    slide_text: str = ""
    chart_xml: str = ""
    workbook_embedded: str = ""
    sheet_name: str = ""
    expected_orientation: str = ""
    expected_categories: list[str] = field(default_factory=list)
    expected_series: list[str] = field(default_factory=list)
    expected_values: list[list[Any]] = field(default_factory=list)
    table_cells: list[list[str]] = field(default_factory=list)
    text: str = ""
    target_key: str = ""

    @property
    def target_id(self) -> str:
        return self.target_key or self.shape_name


@dataclass(frozen=True)
class _SlideTextBox:
    text: str
    left_in: float
    top_in: float
    width_in: float
    height_in: float


def read_bytes(file: InputFile) -> bytes:
    if isinstance(file, (bytes, bytearray)):
        return bytes(file)
    if isinstance(file, (str, Path)):
        return Path(file).read_bytes()
    if hasattr(file, "seek"):
        file.seek(0)
    data = file.read()
    return data if isinstance(data, bytes) else bytes(data)


def discover_ppt_targets(
    pptx_file: InputFile,
    numeric_only: bool = True,
    include_text_shapes: bool = True,
) -> list[PptTarget]:
    ppt_bytes = read_bytes(pptx_file)
    targets: list[PptTarget] = []
    with ZipFile(BytesIO(ppt_bytes)) as zf:
        slide_paths = _sorted_slide_paths(zf)
        for slide_index, slide_path in enumerate(slide_paths):
            slide_number = slide_index + 1
            slide_rels = _relationship_map(zf, _rels_path(slide_path))
            slide_root = ET.fromstring(zf.read(slide_path))
            text_boxes = _slide_text_boxes(slide_root)
            slide_text = _join_unique([box.text for box in text_boxes])

            for frame in slide_root.findall(".//p:graphicFrame", NS):
                cnv = frame.find("./p:nvGraphicFramePr/p:cNvPr", NS)
                if cnv is None:
                    continue
                shape_name = _text(cnv.attrib.get("name"))
                if numeric_only and not _is_update_name(shape_name):
                    continue
                shape_id = _text(cnv.attrib.get("id"))
                left, top, width, height = _graphic_frame_dimensions(frame)
                nearby_text = _nearby_text(left, top, width, height, text_boxes)
                chart_el = frame.find(".//c:chart", NS)
                table_el = frame.find(".//a:tbl", NS)
                if chart_el is not None:
                    target = _chart_target(
                        zf,
                        slide_rels,
                        slide_index,
                        slide_number,
                        slide_path,
                        shape_name,
                        shape_id,
                        left,
                        top,
                        width,
                        height,
                        nearby_text,
                        slide_text,
                        chart_el,
                    )
                    if target:
                        targets.append(target)
                elif table_el is not None:
                    targets.append(
                        PptTarget(
                            slide_index=slide_index,
                            slide_number=slide_number,
                            slide_path=slide_path,
                            shape_name=shape_name,
                            shape_id=shape_id,
                            object_type="table",
                            left_in=left,
                            top_in=top,
                            width_in=width,
                            height_in=height,
                            nearby_text=nearby_text,
                            slide_text=slide_text,
                            table_cells=_table_cells(table_el),
                        )
                    )

            if not include_text_shapes:
                continue

            for shape in slide_root.findall(".//p:sp", NS):
                cnv = shape.find("./p:nvSpPr/p:cNvPr", NS)
                if cnv is None:
                    continue
                shape_name = _text(cnv.attrib.get("name"))
                if numeric_only and not _is_update_name(shape_name):
                    continue
                left, top, width, height = _shape_dimensions(shape)
                texts = [_text(node.text) for node in shape.findall(".//a:t", NS)]
                text = _join_unique([part for part in texts if part])
                targets.append(
                    PptTarget(
                        slide_index=slide_index,
                        slide_number=slide_number,
                        slide_path=slide_path,
                        shape_name=shape_name,
                        shape_id=_text(cnv.attrib.get("id")),
                        object_type="text" if text else "shape",
                        left_in=left,
                        top_in=top,
                        width_in=width,
                        height_in=height,
                        nearby_text=text,
                        slide_text=slide_text,
                        text=text,
                    )
                )
    from .target_labeler import assign_slide_target_ids

    return assign_slide_target_ids(sorted(targets, key=lambda item: (item.slide_number, item.top_in, item.left_in, item.shape_name)))


def _chart_target(
    zf: ZipFile,
    slide_rels: dict[str, dict[str, str]],
    slide_index: int,
    slide_number: int,
    slide_path: str,
    shape_name: str,
    shape_id: str,
    left: float,
    top: float,
    width: float,
    height: float,
    nearby_text: str,
    slide_text: str,
    chart_el: ET.Element,
) -> PptTarget | None:
    rel_id = chart_el.attrib.get(f"{{{R_NS}}}id")
    if not rel_id or rel_id not in slide_rels:
        return None
    chart_path = _norm_join(slide_path, slide_rels[rel_id]["target"])
    chart_rels = _relationship_map(zf, _rels_path(chart_path))
    workbook_path = ""
    for rel in chart_rels.values():
        if rel["type"].endswith("/package"):
            workbook_path = _norm_join(chart_path, rel["target"])
            break

    sheet_name = ""
    expected_categories: list[str] = []
    expected_series: list[str] = []
    expected_values: list[list[Any]] = []
    orientation = ""
    if chart_path in zf.namelist():
        structure = _read_chart_structure(zf, chart_path, workbook_path)
        sheet_name = structure["sheet_name"]
        expected_categories = structure["categories"]
        expected_series = structure["series"]
        expected_values = structure["values"]
        orientation = structure["orientation"]

    return PptTarget(
        slide_index=slide_index,
        slide_number=slide_number,
        slide_path=slide_path,
        shape_name=shape_name,
        shape_id=shape_id,
        object_type="chart",
        left_in=left,
        top_in=top,
        width_in=width,
        height_in=height,
        nearby_text=nearby_text,
        slide_text=slide_text,
        chart_xml=chart_path,
        workbook_embedded=workbook_path,
        sheet_name=sheet_name,
        expected_orientation=orientation,
        expected_categories=expected_categories,
        expected_series=expected_series,
        expected_values=expected_values,
    )


def _read_chart_structure(zf: ZipFile, chart_path: str, workbook_path: str) -> dict[str, Any]:
    chart_root = ET.fromstring(zf.read(chart_path))
    series_elements = chart_root.findall(".//c:ser", NS)
    series_labels: list[str] = []
    categories: list[str] = []
    values_by_series: list[list[Any]] = []
    first_cat_formula = ""
    first_val_formula = ""
    workbook = None
    sheet_name = ""
    load_workbook = _chart_workbook_lookup_enabled()

    for ser in series_elements:
        tx_formula = _node_text(ser.find("./c:tx//c:f", NS))
        tx_label = _node_text(ser.find("./c:tx//c:v", NS))
        if not sheet_name:
            sheet_name = _formula_sheet_name(tx_formula)
        if load_workbook and not tx_label and tx_formula and workbook is None:
            workbook = _load_chart_workbook(zf, workbook_path)
        if not tx_label and tx_formula and workbook:
            tx_label = _formula_single_value(workbook, tx_formula)
        series_labels.append(_text(tx_label))

        cat_formula = _node_text(ser.find("./c:cat//c:f", NS))
        val_formula = _node_text(ser.find("./c:val//c:f", NS))
        if not sheet_name:
            sheet_name = _formula_sheet_name(cat_formula) or _formula_sheet_name(val_formula)
        if not first_cat_formula:
            first_cat_formula = cat_formula
        if not first_val_formula:
            first_val_formula = val_formula
        if not categories:
            categories = _cache_values(ser, "./c:cat//c:pt/c:v")
            if load_workbook and not categories and workbook is None:
                workbook = _load_chart_workbook(zf, workbook_path)
            if not categories and cat_formula and workbook:
                categories = [_text(value) for value in _formula_range_values(workbook, cat_formula)]
        vals = _cache_values(ser, "./c:val//c:pt/c:v", numeric=True)
        if load_workbook and not vals and workbook is None:
            workbook = _load_chart_workbook(zf, workbook_path)
        if not vals and val_formula and workbook:
            vals = _formula_range_values(workbook, val_formula)
        values_by_series.append(vals)

    orientation = _chart_orientation(first_cat_formula, first_val_formula)
    if orientation == "categories_rows_series_columns":
        values = _transpose_series_values(values_by_series, len(categories))
    else:
        values = values_by_series
    return {
        "sheet_name": sheet_name or (workbook.worksheets[0].title if workbook else ""),
        "categories": categories,
        "series": series_labels,
        "values": values,
        "orientation": orientation,
    }


def _load_chart_workbook(zf: ZipFile, workbook_path: str):
    workbook_bytes = zf.read(workbook_path) if workbook_path and workbook_path in zf.namelist() else b""
    return openpyxl.load_workbook(BytesIO(workbook_bytes), data_only=True) if workbook_bytes else None


def _chart_workbook_lookup_enabled() -> bool:
    return os.getenv("AUTO_PPT_LOAD_CHART_WORKBOOKS", "").strip().lower() in {"1", "true", "yes", "on"}


def _formula_sheet_name(formula: str) -> str:
    text = _text(formula).replace("$", "")
    if "!" not in text:
        return ""
    sheet, _ref = text.split("!", 1)
    return sheet.strip("'")


def _chart_orientation(cat_formula: str, val_formula: str) -> str:
    try:
        _cat_sheet, cat_range = _split_formula_ref(cat_formula)
        _val_sheet, val_range = _split_formula_ref(val_formula)
        cat_bounds = range_boundaries(cat_range)
        val_bounds = range_boundaries(val_range)
    except Exception:
        return "categories_rows_series_columns"

    cat_min_col, cat_min_row, cat_max_col, cat_max_row = cat_bounds
    val_min_col, val_min_row, val_max_col, val_max_row = val_bounds
    cat_vertical = cat_min_col == cat_max_col and cat_min_row != cat_max_row
    cat_horizontal = cat_min_row == cat_max_row and cat_min_col != cat_max_col
    val_vertical = val_min_col == val_max_col and val_min_row != val_max_row
    val_horizontal = val_min_row == val_max_row and val_min_col != val_max_col
    if cat_vertical and val_vertical:
        return "categories_rows_series_columns"
    if cat_horizontal and val_horizontal:
        return "series_rows_categories_columns"
    if cat_horizontal:
        return "series_rows_categories_columns"
    return "categories_rows_series_columns"


def _formula_single_value(workbook: Any, formula: str) -> str:
    values = _formula_range_values(workbook, formula)
    return _text(values[0]) if values else ""


def _formula_range_values(workbook: Any, formula: str) -> list[Any]:
    sheet_name, ref = _split_formula_ref(formula)
    if not sheet_name:
        sheet_name = workbook.worksheets[0].title
    if sheet_name not in workbook.sheetnames:
        return []
    ws = workbook[sheet_name]
    min_col, min_row, max_col, max_row = range_boundaries(ref)
    values: list[Any] = []
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            values.append(ws.cell(row=row, column=col).value)
    return values


def _split_formula_ref(formula: str) -> tuple[str, str]:
    text = _text(formula).replace("$", "")
    if "!" not in text:
        return "", text
    sheet, ref = text.split("!", 1)
    return sheet.strip("'"), ref


def _cache_values(ser: ET.Element, path: str, numeric: bool = False) -> list[Any]:
    values: list[Any] = []
    for node in ser.findall(path, NS):
        if numeric:
            try:
                values.append(float(node.text))
            except (TypeError, ValueError):
                values.append(node.text)
        else:
            values.append(_text(node.text))
    return values


def _transpose_series_values(values_by_series: list[list[Any]], category_count: int) -> list[list[Any]]:
    output: list[list[Any]] = []
    for row_index in range(category_count):
        output.append([values[row_index] if row_index < len(values) else None for values in values_by_series])
    return output


def _table_cells(table_el: ET.Element) -> list[list[str]]:
    rows: list[list[str]] = []
    for tr in table_el.findall("./a:tr", NS):
        cells: list[str] = []
        for tc in tr.findall("./a:tc", NS):
            texts = [_text(t.text) for t in tc.findall(".//a:t", NS)]
            cells.append("".join(texts))
        rows.append(cells)
    return rows


def _is_update_name(value: str) -> bool:
    return bool(re.fullmatch(r"\d{4,}", _text(value)))


def _sorted_slide_paths(zf: ZipFile) -> list[str]:
    return sorted(
        [n for n in zf.namelist() if re.fullmatch(r"ppt/slides/slide\d+\.xml", n)],
        key=lambda name: int(re.search(r"slide(\d+)\.xml", name).group(1)),
    )


def _rels_path(part_path: str) -> str:
    directory, filename = posixpath.split(part_path)
    return posixpath.join(directory, "_rels", f"{filename}.rels")


def _relationship_map(zf: ZipFile, path: str) -> dict[str, dict[str, str]]:
    if path not in zf.namelist():
        return {}
    root = ET.fromstring(zf.read(path))
    output: dict[str, dict[str, str]] = {}
    for rel in root.findall(f"{{{REL_NS}}}Relationship"):
        output[rel.attrib["Id"]] = {
            "type": rel.attrib.get("Type", ""),
            "target": rel.attrib.get("Target", ""),
        }
    return output


def _norm_join(base_part: str, target: str) -> str:
    return posixpath.normpath(posixpath.join(posixpath.dirname(base_part), target))


def _graphic_frame_dimensions(frame: ET.Element) -> tuple[float, float, float, float]:
    return _xfrm_dimensions(frame.find("./p:xfrm", NS))


def _shape_dimensions(shape: ET.Element) -> tuple[float, float, float, float]:
    return _xfrm_dimensions(shape.find("./p:spPr/a:xfrm", NS))


def _xfrm_dimensions(xfrm: ET.Element | None) -> tuple[float, float, float, float]:
    if xfrm is None:
        return 0.0, 0.0, 0.0, 0.0
    off = xfrm.find("./a:off", NS)
    ext = xfrm.find("./a:ext", NS)
    if off is None or ext is None:
        return 0.0, 0.0, 0.0, 0.0
    emu_per_inch = 914400
    return (
        round(int(off.attrib.get("x", "0")) / emu_per_inch, 3),
        round(int(off.attrib.get("y", "0")) / emu_per_inch, 3),
        round(int(ext.attrib.get("cx", "0")) / emu_per_inch, 3),
        round(int(ext.attrib.get("cy", "0")) / emu_per_inch, 3),
    )


def _slide_text_boxes(slide_root: ET.Element) -> list[_SlideTextBox]:
    boxes: list[_SlideTextBox] = []
    for shape in slide_root.findall(".//p:sp", NS):
        texts = [_text(node.text) for node in shape.findall(".//a:t", NS)]
        text = _join_unique([part for part in texts if part])
        if not text:
            continue
        left, top, width, height = _shape_dimensions(shape)
        boxes.append(_SlideTextBox(text, left, top, width, height))
    return boxes


def _nearby_text(left: float, top: float, width: float, height: float, boxes: list[_SlideTextBox]) -> str:
    if not boxes:
        return ""
    chart_center_x = left + width / 2
    chart_center_y = top + height / 2
    ranked: list[tuple[float, str]] = []
    for box in boxes:
        box_center_x = box.left_in + box.width_in / 2
        box_center_y = box.top_in + box.height_in / 2
        distance = abs(chart_center_x - box_center_x) + abs(chart_center_y - box_center_y)
        horizontal_overlap = _overlap_ratio(left, width, box.left_in, box.width_in)
        above_or_inside = box.top_in <= top + height and box.top_in + box.height_in <= top + height + 0.5
        if horizontal_overlap >= 0.2 and above_or_inside:
            distance *= 0.45
        ranked.append((distance, box.text))
    ranked.sort(key=lambda item: item[0])
    return _join_unique([text for _distance, text in ranked[:8]])


def _overlap_ratio(left_a: float, width_a: float, left_b: float, width_b: float) -> float:
    right_a = left_a + width_a
    right_b = left_b + width_b
    overlap = max(0.0, min(right_a, right_b) - max(left_a, left_b))
    smaller = min(width_a, width_b)
    return overlap / smaller if smaller else 0.0


def _join_unique(values: Iterable[Any], separator: str = " | ") -> str:
    parts: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = _text(value)
        key = _norm(text)
        if not text or key in seen:
            continue
        parts.append(text)
        seen.add(key)
    return separator.join(parts)


def _node_text(node: ET.Element | None) -> str:
    return _text(node.text) if node is not None else ""


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _norm(value: Any) -> str:
    text = _text(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.upper()
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()
