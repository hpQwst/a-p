from __future__ import annotations

from io import BytesIO
from pathlib import Path
from zipfile import ZipFile
import os
import unittest
import xml.etree.ElementTree as ET

import openpyxl

from ppt_automator import analyze_update_package, generate_updated_pptx


MB_DIR = Path(os.getenv("AUTO_PPT_MB_TEST_DIR", r"C:\Users\HugoRocha\Documents\automatizador-ppt-arquivos\mb"))
PPT = MB_DIR / "MBTESTE_formula.pptx"
DATASOURCES = MB_DIR / "datasources.zip"

NS = {
    "p": "http://schemas.openxmlformats.org/presentationml/2006/main",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "c": "http://schemas.openxmlformats.org/drawingml/2006/chart",
}


@unittest.skipUnless(PPT.exists() and DATASOURCES.exists(), "Arquivos MB de regressao nao encontrados.")
class MbUpdateTargetTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.targets, cls.sources, cls.plans = analyze_update_package(PPT, DATASOURCES)

    def test_discovers_chart_and_table_targets_on_slide_1(self) -> None:
        by_name = {target.shape_name: target for target in self.targets}
        self.assertEqual(by_name["3334180514"].slide_number, 1)
        self.assertEqual(by_name["3334180514"].object_type, "chart")
        self.assertEqual(by_name["1424058794"].slide_number, 1)
        self.assertEqual(by_name["1424058794"].object_type, "table")

    def test_xlsx_333_is_series_rows_months_columns(self) -> None:
        source = self._source("3334180514")
        self.assertEqual(source.orientation, "series_rows_categories_columns")
        self.assertEqual(source.categories, ["Nov/25", "Dez/25", "Jan/26", "Fev/26", "Mar/26"])
        self.assertEqual(source.series, ["Detrator", "Neutro", "Promotor", "NPS"])

    def test_chart_333_plan_transposes_and_preserves_scales(self) -> None:
        plan = self._plan("3334180514")
        self.assertEqual(plan.action, "transpose")
        self.assertEqual(plan.orientation_ppt, "categories_rows_series_columns")
        self.assertEqual(plan.categories, ["Nov/25", "Dez/25", "Jan/26", "Fev/26", "Mar/26"])
        self.assertEqual(plan.series, ["Detrator", "Neutro", "Promotor", "NPS"])
        self.assertLess(plan.values[0][0], 1)
        self.assertLess(plan.values[0][1], 1)
        self.assertLess(plan.values[0][2], 1)
        self.assertGreater(plan.values[0][3], 1)

    def test_xlsx_142_is_single_series_months_columns(self) -> None:
        source = self._source("1424058794")
        self.assertEqual(source.orientation, "single_series_row_categories_columns")
        self.assertEqual(source.categories, ["Nov/25", "Dez/25", "Jan/26", "Fev/26", "Mar/26"])
        self.assertEqual(source.values[0], [13126.0, 12626.0, 8483.0, 9401.0, 11929.0])

    def test_generated_ppt_updates_chart_and_powerpoint_table(self) -> None:
        output = generate_updated_pptx(PPT, self.plans)
        with ZipFile(BytesIO(output)) as zf:
            wb = openpyxl.load_workbook(BytesIO(zf.read("ppt/embeddings/Microsoft_Excel_Worksheet.xlsx")), data_only=True)
            ws = wb.worksheets[0]
            self.assertEqual([ws.cell(1, col).value for col in range(2, 6)], ["Detrator", "Neutro", "Promotor", "NPS"])
            self.assertEqual([ws.cell(row, 1).value for row in range(2, 7)], ["Nov/25", "Dez/25", "Jan/26", "Fev/26", "Mar/26"])
            self.assertAlmostEqual(ws.cell(2, 2).value, 0.2008989791254)
            self.assertAlmostEqual(ws.cell(2, 5).value, 44.67469145207984)

            table_plan = self._plan("1424058794")
            table_values = self._table_values(zf.read("ppt/slides/slide1.xml"), table_plan.target.target_id)
            self.assertEqual(table_values, [["13.126", "12.626", "8.483", "9.401", "11.929"]])

            chart_root = ET.fromstring(zf.read("ppt/charts/chart1.xml"))
            series_names = [
                ser.find(".//c:tx//c:v", NS).text
                for ser in chart_root.findall(".//c:ser", NS)
            ]
            self.assertEqual(series_names[:4], ["Detrator", "Neutro", "Promotor", "NPS"])

    def _source(self, graph_id: str):
        return next(source for source in self.sources if source.source_id == graph_id)

    def _plan(self, target_id: str):
        return next(plan for plan in self.plans if plan.target.shape_name == target_id)

    def _table_values(self, slide_xml: bytes, shape_name: str) -> list[list[str]]:
        root = ET.fromstring(slide_xml)
        for frame in root.findall(".//p:graphicFrame", NS):
            cnv = frame.find("./p:nvGraphicFramePr/p:cNvPr", NS)
            if cnv is None or cnv.attrib.get("name") != shape_name:
                continue
            table = frame.find(".//a:tbl", NS)
            return [
                ["".join(t.text or "" for t in cell.findall(".//a:t", NS)) for cell in row.findall("./a:tc", NS)]
                for row in table.findall("./a:tr", NS)
            ]
        return []


if __name__ == "__main__":
    unittest.main()
