from __future__ import annotations

from io import BytesIO
from typing import Any
from zipfile import ZipFile
import re
import xml.etree.ElementTree as ET

import openpyxl

from .ppt_discovery import CHART_NS, NS, PptTarget
from .table_normalizer import TransformPlan


PERCENT_FORMAT = "0.0%"


def chart_replacements(zf: ZipFile, target: PptTarget, plan: TransformPlan) -> dict[str, bytes]:
    if not target.chart_xml or not target.workbook_embedded:
        return {}
    return {
        target.workbook_embedded: _updated_workbook_bytes(zf, target, plan),
        target.chart_xml: _updated_chart_xml_bytes(zf, target, plan),
    }


def _updated_workbook_bytes(zf: ZipFile, target: PptTarget, plan: TransformPlan) -> bytes:
    wb = openpyxl.load_workbook(BytesIO(zf.read(target.workbook_embedded)))
    ws = wb[target.sheet_name] if target.sheet_name in wb.sheetnames else wb.worksheets[0]
    number_format = _plan_number_format(plan)

    if plan.orientation_ppt == "series_rows_categories_columns":
        max_row = max(ws.max_row, len(plan.series) + 1)
        max_col = max(ws.max_column, len(plan.categories) + 1)
        _clear_values(ws, max_row, max_col)
        ws.cell(row=1, column=1).value = None
        for col_idx, category in enumerate(plan.categories, 2):
            ws.cell(row=1, column=col_idx).value = category
        for row_idx, series_name in enumerate(plan.series, 2):
            ws.cell(row=row_idx, column=1).value = series_name
            row_values = plan.values[row_idx - 2] if row_idx - 2 < len(plan.values) else []
            for col_idx, value in enumerate(row_values, 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = _workbook_value(value, number_format)
                if number_format:
                    cell.number_format = number_format
    else:
        max_row = max(ws.max_row, len(plan.categories) + 1)
        max_col = max(ws.max_column, len(plan.series) + 1)
        _clear_values(ws, max_row, max_col)
        ws.cell(row=1, column=1).value = None
        for col_idx, series_name in enumerate(plan.series, 2):
            ws.cell(row=1, column=col_idx).value = series_name
        for row_idx, category in enumerate(plan.categories, 2):
            ws.cell(row=row_idx, column=1).value = category
            row_values = plan.values[row_idx - 2] if row_idx - 2 < len(plan.values) else []
            for col_idx, value in enumerate(row_values, 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = _workbook_value(value, number_format)
                if number_format:
                    cell.number_format = number_format

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _updated_chart_xml_bytes(zf: ZipFile, target: PptTarget, plan: TransformPlan) -> bytes:
    root = ET.fromstring(zf.read(target.chart_xml))
    series_elements = root.findall(".//c:ser", NS)
    sheet = _sheet_ref(target.sheet_name or "Sheet1")
    number_format = _plan_number_format(plan)

    if plan.orientation_ppt == "series_rows_categories_columns":
        end_col = _excel_col(len(plan.categories) + 1)
        for index, ser in enumerate(series_elements[: len(plan.series)]):
            excel_row = index + 2
            values = plan.values[index] if index < len(plan.values) else []
            _update_series_text(ser, f"{sheet}!$A${excel_row}", plan.series[index])
            _update_series_categories(ser, f"{sheet}!$B$1:${end_col}$1", plan.categories)
            _update_series_values(ser, f"{sheet}!$B${excel_row}:${end_col}${excel_row}", values, number_format)
    else:
        end_row = len(plan.categories) + 1
        for index, ser in enumerate(series_elements[: len(plan.series)]):
            excel_col = _excel_col(index + 2)
            values = [row[index] if index < len(row) else None for row in plan.values]
            _update_series_text(ser, f"{sheet}!${excel_col}$1", plan.series[index])
            _update_series_categories(ser, f"{sheet}!$A$2:$A${end_row}", plan.categories)
            _update_series_values(ser, f"{sheet}!${excel_col}$2:${excel_col}${end_row}", values, number_format)

    if number_format:
        _update_data_label_number_format(root, number_format)

    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _clear_values(ws: Any, max_row: int, max_col: int) -> None:
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.value = None


def _update_series_text(ser: ET.Element, formula: str, label: str) -> None:
    tx = ser.find("./c:tx/c:strRef", NS)
    if tx is None:
        tx_parent = ser.find("./c:tx", NS)
        if tx_parent is None:
            tx_parent = ET.SubElement(ser, f"{{{CHART_NS}}}tx")
        tx = ET.SubElement(tx_parent, f"{{{CHART_NS}}}strRef")
    _set_formula(tx, formula)
    cache = tx.find("./c:strCache", NS)
    if cache is None:
        cache = ET.SubElement(tx, f"{{{CHART_NS}}}strCache")
    _set_cache_values(cache, [label], numeric=False)


def _update_series_categories(ser: ET.Element, formula: str, labels: list[str]) -> None:
    cat = ser.find("./c:cat/c:strRef", NS)
    if cat is None:
        cat_parent = ser.find("./c:cat", NS)
        if cat_parent is None:
            cat_parent = ET.SubElement(ser, f"{{{CHART_NS}}}cat")
        cat = ET.SubElement(cat_parent, f"{{{CHART_NS}}}strRef")
    _set_formula(cat, formula)
    cache = cat.find("./c:strCache", NS)
    if cache is None:
        cache = ET.SubElement(cat, f"{{{CHART_NS}}}strCache")
    _set_cache_values(cache, labels, numeric=False)


def _update_series_values(ser: ET.Element, formula: str, values: list[Any], number_format: str = "") -> None:
    val = ser.find("./c:val/c:numRef", NS)
    if val is None:
        val_parent = ser.find("./c:val", NS)
        if val_parent is None:
            val_parent = ET.SubElement(ser, f"{{{CHART_NS}}}val")
        val = ET.SubElement(val_parent, f"{{{CHART_NS}}}numRef")
    _set_formula(val, formula)
    cache = val.find("./c:numCache", NS)
    if cache is None:
        cache = ET.SubElement(val, f"{{{CHART_NS}}}numCache")
    _set_cache_values(cache, values, numeric=True, number_format=number_format)


def _set_formula(parent: ET.Element, formula: str) -> None:
    formula_el = parent.find("./c:f", NS)
    if formula_el is None:
        formula_el = ET.SubElement(parent, f"{{{CHART_NS}}}f")
    formula_el.text = formula


def _set_cache_values(cache: ET.Element, values: list[Any], numeric: bool, number_format: str = "") -> None:
    if numeric and number_format:
        _set_format_code(cache, number_format)

    for child in list(cache):
        if child.tag in {f"{{{CHART_NS}}}ptCount", f"{{{CHART_NS}}}pt"}:
            cache.remove(child)

    insert_at = 0
    for i, child in enumerate(list(cache)):
        if child.tag == f"{{{CHART_NS}}}formatCode":
            insert_at = i + 1

    cache.insert(insert_at, ET.Element(f"{{{CHART_NS}}}ptCount", {"val": str(len(values))}))
    for offset, value in enumerate(values):
        pt = ET.Element(f"{{{CHART_NS}}}pt", {"idx": str(offset)})
        v = ET.SubElement(pt, f"{{{CHART_NS}}}v")
        v.text = _chart_value_text(value, numeric=numeric, number_format=number_format)
        cache.insert(insert_at + offset + 1, pt)


def _set_format_code(cache: ET.Element, number_format: str) -> None:
    format_el = cache.find("./c:formatCode", NS)
    if format_el is None:
        format_el = ET.Element(f"{{{CHART_NS}}}formatCode")
        cache.insert(0, format_el)
    format_el.text = number_format


def _update_data_label_number_format(root: ET.Element, number_format: str) -> None:
    for d_lbls in root.findall(".//c:dLbls", NS):
        num_fmt = d_lbls.find("./c:numFmt", NS)
        if num_fmt is None:
            num_fmt = ET.Element(f"{{{CHART_NS}}}numFmt")
            d_lbls.insert(0, num_fmt)
        num_fmt.attrib["formatCode"] = number_format
        num_fmt.attrib["sourceLinked"] = "0"


def _chart_value_text(value: Any, numeric: bool, number_format: str = "") -> str:
    if value is None:
        return "0" if numeric else ""
    if numeric:
        parsed = _numeric_value(value, percentage=bool(number_format))
        if parsed is not None:
            return f"{parsed:.12g}"
    if isinstance(value, float):
        return f"{value:.12g}"
    return str(value)


def _workbook_value(value: Any, number_format: str = "") -> Any:
    if not number_format:
        parsed = _numeric_value(value, percentage=False)
        return parsed if parsed is not None else value
    parsed = _numeric_value(value, percentage=True)
    return parsed if parsed is not None else value


def _plan_number_format(plan: TransformPlan) -> str:
    if plan.preserve_percentage_decimal:
        return PERCENT_FORMAT
    values = [value for row in plan.values for value in row]
    if any(_is_percent_text(value) for value in values):
        return PERCENT_FORMAT
    numeric = [_numeric_value(value, percentage=False) for value in values]
    numeric = [value for value in numeric if value is not None]
    if not numeric:
        return ""
    decimal_like = sum(1 for value in numeric if 0 <= abs(value) <= 1 and value not in {0, 1})
    return PERCENT_FORMAT if decimal_like >= max(2, len(numeric) * 0.5) else ""


def _numeric_value(value: Any, percentage: bool = False) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    is_percent = "%" in text
    text = text.replace("%", "").strip()
    text = re.sub(r"^,", "0,", text)
    text = re.sub(r"^- ,", "-0,", text)
    text = text.replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", ".")
    try:
        number = float(text)
    except ValueError:
        return None
    if percentage or is_percent:
        if is_percent or abs(number) > 1:
            return number / 100
    return number


def _is_percent_text(value: Any) -> bool:
    return isinstance(value, str) and "%" in value


def _sheet_ref(sheet_name: str) -> str:
    if any(ch in sheet_name for ch in " -'"):
        return "'" + sheet_name.replace("'", "''") + "'"
    return sheet_name


def _excel_col(index: int) -> str:
    letters = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters
