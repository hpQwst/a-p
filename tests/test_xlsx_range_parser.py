from __future__ import annotations

from datetime import datetime
from io import BytesIO
import unittest

import openpyxl
from openpyxl.worksheet.table import Table

from ppt_automator.xlsx_parser import parse_xlsx_table


class XlsxRangeParserTests(unittest.TestCase):
    def test_parse_xlsx_table_uses_manual_range_only(self) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dados"
        ws["A1"] = "ruido fora do range"
        ws["D5"] = ""
        ws["E5"] = "Jan/26"
        ws["F5"] = "Fev/26"
        ws["D6"] = "Total"
        ws["E6"] = 10
        ws["F6"] = 20

        data = BytesIO()
        wb.save(data)

        parsed = parse_xlsx_table(
            data.getvalue(),
            file_name="manual.xlsx",
            formula_mode="auto",
            cell_range="Dados!D5:F6",
        )

        self.assertEqual(parsed.sheet_name, "Dados")
        self.assertEqual(parsed.used_range, (5, 4, 6, 6))
        self.assertEqual(parsed.orientation, "single_series_row_categories_columns")
        self.assertEqual(parsed.categories, ["Jan/26", "Fev/26"])
        self.assertEqual(parsed.series, ["Total"])
        self.assertEqual(parsed.values, [[10, 20]])

    def test_dynamic_range_grows_months_and_rows_but_stops_before_auxiliary_column(self) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dados"
        ws.append(["Atributo", "Jan/26", "Fev/26", "Mar/26", "Auxiliar"])
        ws.append(["NPS", 40, 41, 42, "=SUM(B2:D2)"])
        ws.append(["CSAT", 80, 81, 82, "=SUM(B3:D3)"])
        ws.append(["Conversao", 10, 11, 12, "=SUM(B4:D4)"])

        data = BytesIO()
        wb.save(data)
        wb.close()

        parsed = parse_xlsx_table(
            data.getvalue(),
            file_name="mensal.xlsx",
            cell_range="Dados!A1:C3",
            range_mode="dynamic",
        )

        self.assertEqual(parsed.used_range, (1, 1, 4, 4))
        self.assertEqual(parsed.categories, ["Jan/26", "Fev/26", "Mar/26"])
        self.assertEqual(parsed.series, ["NPS", "CSAT", "Conversao"])
        self.assertEqual(parsed.values[-1], [10, 11, 12])

    def test_exact_range_does_not_grow(self) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Atributo", "Jan/26", "Fev/26", "Mar/26"])
        ws.append(["NPS", 40, 41, 42])
        ws.append(["CSAT", 80, 81, 82])

        data = BytesIO()
        wb.save(data)
        wb.close()

        parsed = parse_xlsx_table(
            data.getvalue(),
            file_name="mensal.xlsx",
            cell_range="A1:C2",
            range_mode="exact",
        )

        self.assertEqual(parsed.used_range, (1, 1, 2, 3))
        self.assertEqual(parsed.categories, ["Jan/26", "Fev/26"])
        self.assertEqual(parsed.series, ["NPS"])

    def test_dynamic_range_uses_excel_table_boundary_and_ignores_side_data(self) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Atributo", "Valor", "Meta", "Auxiliar"])
        ws.append(["NPS", 40, 45, 999])
        ws.append(["CSAT", 80, 85, 999])
        ws.add_table(Table(displayName="TabelaReal", ref="A1:C3"))

        data = BytesIO()
        wb.save(data)
        wb.close()

        parsed = parse_xlsx_table(
            data.getvalue(),
            file_name="estruturada.xlsx",
            cell_range="A1:B2",
            range_mode="dynamic",
        )

        self.assertEqual(parsed.used_range, (1, 1, 3, 3))
        self.assertNotIn("Auxiliar", [str(value) for row in parsed.preview_rows for value in row])

    def test_excel_date_period_labels_keep_month_year_text(self) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dados"
        ws.append(["Serie", datetime(2026, 11, 25), "Dez/25", datetime(2026, 1, 26)])
        ws.append(["NPS", 1, 2, 3])

        data = BytesIO()
        wb.save(data)
        wb.close()

        parsed = parse_xlsx_table(data.getvalue(), file_name="periods.xlsx")

        self.assertEqual(parsed.categories, ["Nov/25", "Dez/25", "Jan/26"])

    def test_numeric_year_headers_are_parsed_as_periods(self) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([None, 2025, 2026])
        ws.append(["a", 50, 20])

        data = BytesIO()
        wb.save(data)
        wb.close()

        parsed = parse_xlsx_table(data.getvalue(), file_name="y.xlsx")

        self.assertEqual(parsed.orientation, "single_series_row_categories_columns")
        self.assertEqual(parsed.categories, ["2025", "2026"])
        self.assertEqual(parsed.series, ["a"])
        self.assertEqual(parsed.values, [[50, 20]])


if __name__ == "__main__":
    unittest.main()
