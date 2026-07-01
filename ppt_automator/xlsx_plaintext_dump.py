from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO
from pathlib import PurePosixPath
from typing import Any, BinaryIO
from zipfile import ZipFile
import json
import posixpath
import re
import xml.etree.ElementTree as ET

import openpyxl

from .ppt_discovery import read_bytes


InputFile = str | bytes | bytearray | BinaryIO
SHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
OFFICE_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS = {"s": SHEET_NS, "r": OFFICE_REL_NS}


@dataclass(frozen=True)
class XlsxCellDump:
    sheet: str
    cell: str
    type: str
    raw: str
    display: str
    formula: str
    number_format: str
    merged_range: str = ""
    comment: str = ""
    style: dict[str, Any] = field(default_factory=dict)


@dataclass(frozen=True)
class XlsxSheetDump:
    sheet: str
    used_range: str
    merged_cells: list[str]
    cells: list[XlsxCellDump]


@dataclass(frozen=True)
class XlsxWorkbookDump:
    file_name: str
    sheets: list[XlsxSheetDump]

    def as_prompt_text(self, max_cells_per_sheet: int = 800) -> str:
        chunks = [f"WORKBOOK {self.file_name}"]
        for sheet in self.sheets:
            chunks.append(f"SHEET {sheet.sheet} used_range={sheet.used_range} merged={sheet.merged_cells}")
            for cell in sheet.cells[:max_cells_per_sheet]:
                chunks.append(json.dumps(cell.__dict__, ensure_ascii=False))
            if len(sheet.cells) > max_cells_per_sheet:
                chunks.append(f"... {len(sheet.cells) - max_cells_per_sheet} cells omitted")
        return "\n".join(chunks)


def dump_xlsx_workbook(workbook_file: InputFile, file_name: str = "") -> XlsxWorkbookDump:
    workbook_bytes = read_bytes(workbook_file)
    raw_by_sheet = _raw_cells_by_sheet(workbook_bytes)
    wb = openpyxl.load_workbook(BytesIO(workbook_bytes), data_only=False, read_only=False)
    sheets: list[XlsxSheetDump] = []
    try:
        for ws in wb.worksheets:
            merged_ranges = [str(item) for item in ws.merged_cells.ranges]
            merged_by_cell = _merged_lookup(ws)
            raw_lookup = raw_by_sheet.get(ws.title, {})
            cells: list[XlsxCellDump] = []
            for row in ws.iter_rows():
                for cell in row:
                    raw_info = raw_lookup.get(cell.coordinate, {})
                    formula = ""
                    value = cell.value
                    if isinstance(value, str) and value.startswith("="):
                        formula = value
                        raw = str(raw_info.get("raw") or "")
                        display = raw or value
                        cell_type = "formula"
                    else:
                        raw = _raw_value(raw_info, value)
                        display = _display_value(value)
                        cell_type = _cell_type(cell, value)
                    if not raw and not display and not formula and cell.coordinate not in merged_by_cell:
                        continue
                    comment = cell.comment.text if cell.comment is not None else ""
                    cells.append(
                        XlsxCellDump(
                            sheet=ws.title,
                            cell=cell.coordinate,
                            type=cell_type,
                            raw=raw,
                            display=display,
                            formula=formula,
                            number_format=cell.number_format,
                            merged_range=merged_by_cell.get(cell.coordinate, ""),
                            comment=comment,
                            style={
                                "bold": bool(cell.font and cell.font.bold),
                                "italic": bool(cell.font and cell.font.italic),
                                "fill": getattr(cell.fill.fgColor, "rgb", "") if cell.fill else "",
                            },
                        )
                    )
            sheets.append(
                XlsxSheetDump(
                    sheet=ws.title,
                    used_range=ws.calculate_dimension(),
                    merged_cells=merged_ranges,
                    cells=cells,
                )
            )
    finally:
        wb.close()
    return XlsxWorkbookDump(file_name=file_name, sheets=sheets)


def dump_xlsx_zip_entries(datasources_zip: InputFile, entries: list[Any]) -> list[XlsxWorkbookDump]:
    output: list[XlsxWorkbookDump] = []
    with ZipFile(BytesIO(read_bytes(datasources_zip))) as zf:
        for entry in entries:
            output.append(dump_xlsx_workbook(zf.read(entry.zip_path), file_name=entry.zip_path))
    return output


def _raw_cells_by_sheet(workbook_bytes: bytes) -> dict[str, dict[str, dict[str, str]]]:
    output: dict[str, dict[str, dict[str, str]]] = {}
    shared_strings = _shared_strings(workbook_bytes)
    with ZipFile(BytesIO(workbook_bytes)) as zf:
        sheet_paths = _sheet_paths(zf)
        for sheet_name, sheet_path in sheet_paths.items():
            if sheet_path not in zf.namelist():
                continue
            root = ET.fromstring(zf.read(sheet_path))
            cells: dict[str, dict[str, str]] = {}
            for c in root.findall(".//s:c", NS):
                ref = c.attrib.get("r", "")
                if not ref:
                    continue
                formula = _node_text(c.find("./s:f", NS))
                raw = _node_text(c.find("./s:v", NS))
                cell_type = c.attrib.get("t", "")
                if cell_type == "s" and raw.isdigit() and int(raw) < len(shared_strings):
                    raw = shared_strings[int(raw)]
                if cell_type == "inlineStr":
                    raw = "".join(_node_text(node) for node in c.findall(".//s:t", NS))
                cells[ref] = {"raw": raw, "formula": formula, "type": cell_type}
            output[sheet_name] = cells
    return output


def _sheet_paths(zf: ZipFile) -> dict[str, str]:
    if "xl/workbook.xml" not in zf.namelist() or "xl/_rels/workbook.xml.rels" not in zf.namelist():
        return {}
    rel_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    rels = {
        rel.attrib.get("Id", ""): rel.attrib.get("Target", "")
        for rel in rel_root.findall(f"{{{REL_NS}}}Relationship")
    }
    workbook_root = ET.fromstring(zf.read("xl/workbook.xml"))
    output: dict[str, str] = {}
    for sheet in workbook_root.findall(".//s:sheets/s:sheet", NS):
        name = sheet.attrib.get("name", "")
        rel_id = sheet.attrib.get(f"{{{OFFICE_REL_NS}}}id", "")
        target = rels.get(rel_id, "")
        if not name or not target:
            continue
        path = posixpath.normpath(posixpath.join("xl", target))
        output[name] = path
    return output


def _shared_strings(workbook_bytes: bytes) -> list[str]:
    with ZipFile(BytesIO(workbook_bytes)) as zf:
        if "xl/sharedStrings.xml" not in zf.namelist():
            return []
        root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
    return ["".join(_node_text(node) for node in si.findall(".//s:t", NS)) for si in root.findall("./s:si", NS)]


def _merged_lookup(ws: Any) -> dict[str, str]:
    output: dict[str, str] = {}
    for merged in ws.merged_cells.ranges:
        merged_text = str(merged)
        for row in ws[merged_text]:
            for cell in row:
                output[cell.coordinate] = merged_text
    return output


def _raw_value(raw_info: dict[str, str], value: Any) -> str:
    raw = raw_info.get("raw", "")
    if raw:
        return raw
    if value is None:
        return ""
    if isinstance(value, float):
        return format(value, ".15g")
    return str(value)


def _display_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return format(value, ".15g")
    return str(value)


def _cell_type(cell: Any, value: Any) -> str:
    if value is None:
        return "blank"
    data_type = getattr(cell, "data_type", "")
    if data_type == "n":
        return "number"
    if data_type == "b":
        return "boolean"
    if data_type == "d":
        return "date"
    return "text"


def _node_text(node: ET.Element | None) -> str:
    return "" if node is None or node.text is None else str(node.text)
