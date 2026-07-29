from __future__ import annotations

from dataclasses import dataclass, field, replace
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any, BinaryIO
from zipfile import ZIP_DEFLATED, ZipFile
import re
import unicodedata
import xml.etree.ElementTree as ET

import openpyxl
from openpyxl.utils.cell import range_boundaries

from .core import prepare_workbook_values, read_bytes


InputFile = str | Path | bytes | bytearray | BinaryIO


@dataclass(frozen=True)
class ParsedXlsxTable:
    source_id: str
    file_name: str
    sheet_name: str
    orientation: str
    categories: list[str]
    series: list[str]
    values: list[list[Any]]
    used_range: tuple[int, int, int, int] | None = None
    metadata: dict[str, str] = field(default_factory=dict)
    preview_rows: list[list[Any]] = field(default_factory=list)
    series_number_formats: list[str] = field(default_factory=list)
    # Abas existentes no arquivo. No ZIP com varias abas, cada ParsedXlsxTable
    # representa uma delas; a lista completa sustenta avisos e selecao manual.
    sheet_names: list[str] = field(default_factory=list)
    # Blocos retangulares separados encontrados na aba lida. Mais de um significa
    # que varias tabelas foram somadas num retangulo so, gerando numero errado.
    table_blocks: int = 1

    @property
    def graph_id(self) -> str:
        return self.source_id


SHEET_SEPARATOR = "#"


def zip_path_of(file_name: str) -> str:
    """Caminho do arquivo dentro do ZIP, sem o sufixo de aba."""
    return str(file_name or "").split(SHEET_SEPARATOR, 1)[0]


def source_sheet_of(file_name: str) -> str:
    """Aba embutida no identificador da fonte, vazio quando nao ha."""
    parts = str(file_name or "").split(SHEET_SEPARATOR, 1)
    return parts[1] if len(parts) == 2 else ""


def workbook_sheet_names(workbook_bytes: bytes) -> list[str]:
    workbook = openpyxl.load_workbook(
        BytesIO(_openpyxl_readable_copy(workbook_bytes)), data_only=True, read_only=True
    )
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def parse_datasource_zip(
    datasources_zip: InputFile,
    formula_mode: str = "auto",
    include_names: set[str] | None = None,
) -> list[ParsedXlsxTable]:
    """Cada ABA vira uma fonte independente.

    Planilha de relatorio real costuma ser um workbook unico com dezenas de abas
    alimentando dezenas de graficos (um caso medido tinha 178 abas para 639
    objetos). Ler so a primeira aba fazia esses decks nao casarem com nada.

    Workbook de uma aba so mantem o identificador antigo (o caminho no ZIP),
    para nao invalidar mapeamentos ja aprendidos. Com varias abas, o
    identificador vira 'arquivo.xlsx#Aba'.
    """
    output: list[ParsedXlsxTable] = []
    with ZipFile(BytesIO(read_bytes(datasources_zip))) as zf:
        for name in zf.namelist():
            if not name.lower().endswith(".xlsx"):
                continue
            if include_names is not None and zip_path_of(name) not in include_names:
                continue
            raw = zf.read(name)
            try:
                sheets = workbook_sheet_names(raw)
            except Exception:
                sheets = []
            if len(sheets) <= 1:
                output.append(parse_xlsx_table(raw, file_name=name, formula_mode=formula_mode))
                continue
            # Preparar o workbook (sanitizar, avaliar formulas, abrir) custa o
            # mesmo para uma aba ou para todas. Fazer isso por aba multiplicaria
            # o custo pelo numero de abas - com 178 abas, minutos em vez de
            # segundos. Prepara uma vez e percorre as abas.
            try:
                readable_bytes = _openpyxl_readable_copy(raw)
                calculated_bytes = prepare_workbook_values(readable_bytes, formula_mode=formula_mode)
                data_wb = openpyxl.load_workbook(BytesIO(calculated_bytes), data_only=True, read_only=True)
                formula_wb = openpyxl.load_workbook(BytesIO(readable_bytes), data_only=False, read_only=True)
            except Exception:
                output.append(parse_xlsx_table(raw, file_name=name, formula_mode=formula_mode))
                continue
            try:
                for sheet in sheets:
                    try:
                        output.append(
                            _parse_worksheet(
                                data_wb[sheet],
                                formula_wb[sheet] if sheet in formula_wb.sheetnames else data_wb[sheet],
                                file_name=f"{name}{SHEET_SEPARATOR}{sheet}",
                                sheet_names=sheets,
                            )
                        )
                    except Exception:
                        # Uma aba ilegivel (protegida, corrompida) nao pode
                        # derrubar a leitura das outras 177.
                        continue
            finally:
                data_wb.close()
                formula_wb.close()
    return output


def _parse_worksheet(data_ws, formula_ws, file_name: str, sheet_names: list[str]) -> ParsedXlsxTable:
    """Monta a tabela de UMA aba ja aberta, sem reabrir o workbook."""
    raw_rows, raw_format_rows = _worksheet_rows_and_formats(data_ws)
    formula_all_rows = [list(row) for row in formula_ws.iter_rows(values_only=True)]
    trimmed_rows, used_range = _trim_table(raw_rows)
    formula_rows = _slice_rows(formula_all_rows, used_range)
    format_rows = _slice_rows(raw_format_rows, used_range)
    metadata = _extract_metadata(trimmed_rows)
    parsed = _parse_rectangular_table(
        trimmed_rows,
        formula_rows,
        format_rows,
        file_name=file_name,
        sheet_name=data_ws.title,
        used_range=used_range,
        metadata=metadata,
    )
    return replace(
        parsed,
        sheet_names=list(sheet_names),
        table_blocks=count_table_blocks(trimmed_rows),
    )


def parse_xlsx_table(
    workbook_file: InputFile,
    file_name: str = "",
    formula_mode: str = "auto",
    cell_range: str = "",
    sheet: str = "",
) -> ParsedXlsxTable:
    original_bytes = read_bytes(workbook_file)
    readable_bytes = _openpyxl_readable_copy(original_bytes)
    calculated_bytes = prepare_workbook_values(readable_bytes, formula_mode=formula_mode)
    data_wb = openpyxl.load_workbook(BytesIO(calculated_bytes), data_only=True, read_only=True)
    formula_wb = openpyxl.load_workbook(BytesIO(readable_bytes), data_only=False, read_only=True)
    # Aba explicita ganha do que estiver no cell_range: quem passou o nome sabe
    # de qual aba precisa.
    data_ws, range_ref = _select_worksheet(data_wb, cell_range, sheet=sheet)
    formula_ws = formula_wb[data_ws.title] if data_ws.title in formula_wb.sheetnames else formula_wb.worksheets[0]
    if range_ref:
        trimmed_rows, used_range = _rows_from_range(data_ws, range_ref)
        formula_rows, _formula_used_range = _rows_from_range(formula_ws, range_ref)
        format_rows = _format_rows(data_ws, used_range)
    else:
        raw_rows, raw_format_rows = _worksheet_rows_and_formats(data_ws)
        formula_all_rows = [list(row) for row in formula_ws.iter_rows(values_only=True)]
        trimmed_rows, used_range = _trim_table(raw_rows)
        formula_rows = _slice_rows(formula_all_rows, used_range)
        format_rows = _slice_rows(raw_format_rows, used_range)
    metadata = _extract_metadata(trimmed_rows)
    parsed = _parse_rectangular_table(
        trimmed_rows,
        formula_rows,
        format_rows,
        file_name=file_name,
        sheet_name=data_ws.title,
        used_range=used_range,
        metadata=metadata,
    )
    parsed = replace(
        parsed,
        sheet_names=list(data_wb.sheetnames),
        table_blocks=count_table_blocks(trimmed_rows),
    )
    data_wb.close()
    formula_wb.close()
    return parsed


def _openpyxl_readable_copy(workbook_bytes: bytes) -> bytes:
    """Remove apenas metadados de pivot que impedem o openpyxl de ler o XLSX.

    Tabelas dinamicas nao sao fonte de dados deste produto; os valores visiveis
    das celulas continuam no worksheet. A cirurgia acontece numa copia em
    memoria, nunca no arquivo enviado nem no original do usuario.
    """
    with ZipFile(BytesIO(workbook_bytes)) as source:
        names = source.namelist()
        has_pivot_parts = any(
            name.startswith(("xl/pivotCache/", "xl/pivotTables/"))
            for name in names
        )
        if not has_pivot_parts:
            return workbook_bytes
        output = BytesIO()
        with ZipFile(output, "w", ZIP_DEFLATED) as target:
            for info in source.infolist():
                name = info.filename
                if name.startswith(("xl/pivotCache/", "xl/pivotTables/")):
                    continue
                data = source.read(name)
                if name == "xl/workbook.xml":
                    data = _remove_xml_children_by_local_name(data, {"pivotCaches"})
                elif name.endswith(".rels"):
                    data = _remove_pivot_relationships(data)
                elif name == "[Content_Types].xml":
                    data = _remove_pivot_content_types(data)
                target.writestr(info, data)
        return output.getvalue()


def _remove_xml_children_by_local_name(xml_bytes: bytes, names: set[str]) -> bytes:
    root = ET.fromstring(xml_bytes)
    for parent in root.iter():
        for child in list(parent):
            if child.tag.rsplit("}", 1)[-1] in names:
                parent.remove(child)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _remove_pivot_relationships(xml_bytes: bytes) -> bytes:
    root = ET.fromstring(xml_bytes)
    for relationship in list(root):
        rel_type = str(relationship.attrib.get("Type") or "").lower()
        if "pivottable" in rel_type or "pivotcache" in rel_type:
            root.remove(relationship)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _remove_pivot_content_types(xml_bytes: bytes) -> bytes:
    root = ET.fromstring(xml_bytes)
    for child in list(root):
        part_name = str(child.attrib.get("PartName") or "").lower()
        content_type = str(child.attrib.get("ContentType") or "").lower()
        if "/pivottable" in part_name or "/pivotcache" in part_name or "pivot" in content_type:
            root.remove(child)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _select_worksheet(workbook: Any, cell_range: str, sheet: str = "") -> tuple[Any, str]:
    sheet_name, range_ref = _split_range_ref(cell_range)
    sheet_name = sheet or sheet_name
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Aba '{sheet_name}' nao encontrada no XLSX.")
        return workbook[sheet_name], range_ref
    return workbook.worksheets[0], range_ref


def _split_range_ref(cell_range: str) -> tuple[str, str]:
    text = (cell_range or "").strip().replace("$", "")
    if not text:
        return "", ""
    if "!" in text:
        sheet_name, range_ref = text.split("!", 1)
        sheet_name = sheet_name.strip().strip("'")
    else:
        sheet_name, range_ref = "", text
    try:
        range_boundaries(range_ref)
    except Exception as exc:
        raise ValueError("Range invalido. Use algo como D5:G12 ou Planilha1!D5:G12.") from exc
    return sheet_name, range_ref


def _rows_from_range(worksheet: Any, range_ref: str) -> tuple[list[list[Any]], tuple[int, int, int, int]]:
    min_col, min_row, max_col, max_row = range_boundaries(range_ref)
    rows = [
        list(row)
        for row in worksheet.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        )
    ]
    trimmed_rows, trimmed_range = _trim_table(rows)
    if trimmed_range is None:
        return [], (min_row, min_col, max_row, max_col)
    trim_min_row, trim_min_col, trim_max_row, trim_max_col = trimmed_range
    absolute_range = (
        min_row + trim_min_row - 1,
        min_col + trim_min_col - 1,
        min_row + trim_max_row - 1,
        min_col + trim_max_col - 1,
    )
    return trimmed_rows, absolute_range


def _slice_rows(rows: list[list[Any]], used_range: tuple[int, int, int, int] | None) -> list[list[Any]]:
    if used_range is None:
        return []
    min_row, min_col, max_row, max_col = used_range
    output = []
    for row in rows[min_row - 1 : max_row]:
        output.append(list(row[min_col - 1 : max_col]))
    return output


def _format_rows(worksheet: Any, used_range: tuple[int, int, int, int] | None) -> list[list[str]]:
    if used_range is None:
        return []
    min_row, min_col, max_row, max_col = used_range
    return [
        [str(cell.number_format or "") for cell in row]
        for row in worksheet.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
        )
    ]


def _worksheet_rows_and_formats(
    worksheet: Any,
) -> tuple[list[list[Any]], list[list[str]]]:
    value_rows: list[list[Any]] = []
    format_rows: list[list[str]] = []
    for row in worksheet.iter_rows():
        value_rows.append([cell.value for cell in row])
        format_rows.append([str(cell.number_format or "") for cell in row])
    return value_rows, format_rows


def _parse_rectangular_table(
    rows: list[list[Any]],
    formula_rows: list[list[Any]],
    format_rows: list[list[str]],
    file_name: str,
    sheet_name: str,
    used_range: tuple[int, int, int, int] | None,
    metadata: dict[str, str],
) -> ParsedXlsxTable:
    if not rows:
        return ParsedXlsxTable(
            source_id=_graph_id(Path(file_name).stem),
            file_name=file_name,
            sheet_name=sheet_name,
            orientation="empty",
            categories=[],
            series=[],
            values=[],
            used_range=used_range,
            metadata=metadata,
        )

    header_row_index, header_start_col, header_end_col = _find_header_row(rows)
    if header_row_index is None:
        key_value = _parse_key_value_rows(rows)
        if key_value:
            categories, values = key_value
            return ParsedXlsxTable(
                source_id=_graph_id(Path(file_name).stem),
                file_name=file_name,
                sheet_name=sheet_name,
                orientation="key_value_rows",
                categories=categories,
                series=["Valor"],
                values=values,
                used_range=used_range,
                metadata=metadata,
                preview_rows=_preview_rows(categories, ["Valor"], values, "categories_rows_series_columns"),
            )
        return ParsedXlsxTable(
            source_id=_graph_id(Path(file_name).stem),
            file_name=file_name,
            sheet_name=sheet_name,
            orientation="unknown",
            categories=[],
            series=[],
            values=[],
            used_range=used_range,
            metadata=metadata,
            preview_rows=rows[:10],
        )

    label_col = _find_label_col(rows, header_row_index, header_start_col)
    header_values = [_text(value) for value in rows[header_row_index][header_start_col : header_end_col + 1]]
    data_items: list[tuple[str, list[Any], list[Any], list[str]]] = []
    for row_index in range(header_row_index + 1, len(rows)):
        row = rows[row_index]
        values = list(row[header_start_col : header_end_col + 1])
        if not _has_numeric_or_text(values):
            continue
        label = _text(row[label_col] if label_col is not None and label_col < len(row) else "")
        formula_values = []
        if row_index < len(formula_rows):
            formula_row = formula_rows[row_index]
            formula_values = list(formula_row[header_start_col : header_end_col + 1])
        format_values: list[str] = []
        if row_index < len(format_rows):
            format_row = format_rows[row_index]
            format_values = list(format_row[header_start_col : header_end_col + 1])
        if not label and _looks_like_nps_row(values, formula_values, data_items):
            label = "NPS"
        data_items.append((label, values, formula_values, format_values))

    categories_in_header = _period_score(header_values) >= 0.45
    labels = [item[0] for item in data_items]
    labels_are_categories = _period_score(labels) >= 0.45

    if categories_in_header and len(data_items) == 1:
        orientation = "single_series_row_categories_columns"
        series = [labels[0] or "Valor"]
        values = [data_items[0][1]]
        categories = header_values
        series_number_formats = [_dominant_number_format(data_items[0][3])]
    elif categories_in_header and not labels_are_categories:
        orientation = "series_rows_categories_columns"
        categories = header_values
        series = [
            _series_label(label, values, formulas, index)
            for index, (label, values, formulas, _formats) in enumerate(data_items)
        ]
        values = [item[1] for item in data_items]
        series_number_formats = [_dominant_number_format(item[3]) for item in data_items]
    else:
        orientation = "categories_rows_series_columns"
        categories = [
            label or f"Linha {index + 1}"
            for index, (label, _values, _formulas, _formats) in enumerate(data_items)
        ]
        series = header_values
        values = [item[1] for item in data_items]
        series_number_formats = [
            _dominant_number_format(
                [item[3][column] for item in data_items if column < len(item[3])]
            )
            for column in range(len(series))
        ]

    return ParsedXlsxTable(
        source_id=_graph_id(Path(file_name).stem),
        file_name=file_name,
        sheet_name=sheet_name,
        orientation=orientation,
        categories=categories,
        series=series,
        values=values,
        used_range=used_range,
        metadata=metadata,
        preview_rows=_preview_rows(categories, series, values, orientation),
        series_number_formats=series_number_formats,
    )


def _dominant_number_format(formats: list[str]) -> str:
    meaningful = [str(fmt or "").strip() for fmt in formats if str(fmt or "").strip().lower() not in {"", "general"}]
    if not meaningful:
        return ""
    counts: dict[str, int] = {}
    for fmt in meaningful:
        counts[fmt] = counts.get(fmt, 0) + 1
    return max(counts, key=lambda fmt: (counts[fmt], -meaningful.index(fmt)))


def _parse_key_value_rows(rows: list[list[Any]]) -> tuple[list[str], list[list[Any]]] | None:
    if len(rows) < 2:
        return None
    max_cols = max((len(row) for row in rows), default=0)
    if max_cols < 2:
        return None

    best: tuple[int, int, list[str], list[list[Any]]] | None = None
    for label_col in range(max_cols - 1):
        value_col = label_col + 1
        categories: list[str] = []
        values: list[list[Any]] = []
        for row in rows:
            label = _text(row[label_col] if label_col < len(row) else "")
            if not label or _to_number(label) is not None:
                continue
            value = row[value_col] if value_col < len(row) else None
            categories.append(label)
            values.append([value])
        value_count = sum(1 for row in values if _text(row[0]) or _to_number(row[0]) is not None)
        if len(categories) >= 2 and value_count >= 1:
            score = len(categories) + value_count
            if best is None or score > best[0]:
                best = (score, value_col, categories, values)
    if best is None:
        return None
    return best[2], best[3]


def _find_header_row(rows: list[list[Any]]) -> tuple[int | None, int, int]:
    best: tuple[int | None, int, int, int] = (None, 0, 0, -1)
    for row_index, row in enumerate(rows):
        header_cols = [col for col, value in enumerate(row) if _is_header_cell(value)]
        if len(header_cols) < 2:
            continue
        start = header_cols[0]
        end = header_cols[-1]
        if start == 0 and len(header_cols) > 1:
            if not _looks_like_period(row[0]):
                start = header_cols[1]
        data_score = 0
        for next_row in rows[row_index + 1 : row_index + 5]:
            data_score += sum(2 for value in next_row[start : end + 1] if _to_number(value) is not None)
            data_score += sum(1 for value in next_row[start : end + 1] if _is_value(value))
        header_numeric_penalty = sum(3 for value in row[start : end + 1] if _to_number(value) is not None)
        score = len(header_cols) + data_score - header_numeric_penalty
        if score > best[3]:
            best = (row_index, start, end, score)
    return best[0], best[1], best[2]


def _find_label_col(rows: list[list[Any]], header_row_index: int, header_start_col: int) -> int | None:
    best_col = None
    best_score = 0
    for col in range(header_start_col):
        score = 0
        for row in rows[header_row_index + 1 :]:
            value = row[col] if col < len(row) else None
            if _text(value) and _to_number(value) is None:
                score += 1
        if score > best_score:
            best_col = col
            best_score = score
    if best_col is not None:
        return best_col
    return header_start_col - 1 if header_start_col > 0 else None


def _trim_table(rows: list[list[Any]]) -> tuple[list[list[Any]], tuple[int, int, int, int] | None]:
    positions: list[tuple[int, int]] = []
    for r, row in enumerate(rows):
        for c, value in enumerate(row):
            if _text(value) or _is_value(value):
                positions.append((r, c))
    if not positions:
        return [], None
    min_row = min(r for r, _c in positions)
    max_row = max(r for r, _c in positions)
    min_col = min(c for _r, c in positions)
    max_col = max(c for _r, c in positions)
    trimmed = []
    for row in rows[min_row : max_row + 1]:
        trimmed.append(list(row[min_col : max_col + 1]))
    return trimmed, (min_row + 1, min_col + 1, max_row + 1, max_col + 1)


def count_table_blocks(rows: list[list[Any]]) -> int:
    """Quantas tabelas de dados separadas existem nesta aba.

    Serve so para avisar o usuario: o parser junta a aba inteira num retangulo,
    entao duas tabelas empilhadas viram numeros errados sem ninguem perceber.

    Criterio deliberadamente conservador. Separador e apenas LINHA totalmente
    vazia, nunca coluna: coluna vazia de espacamento entre rotulos e dados e
    comum em planilha exportada e nao significa segunda tabela. Cada bloco
    tambem precisa ter ao menos duas linhas e duas colunas para contar, senao
    titulo solto e linha de nota seriam contados como tabela.

    Falso negativo (duas tabelas lado a lado) e preferivel a falso positivo:
    aviso que dispara em todo arquivo vira ruido e faz ignorarem os avisos reais.
    """
    filled: list[list[bool]] = [
        [bool(_text(value) or _is_value(value)) for value in row] for row in rows
    ]
    if not any(any(row) for row in filled):
        return 0

    blocks = 0
    run: list[list[bool]] = []
    for row in filled + [[]]:
        if any(row):
            run.append(row)
            continue
        if run:
            if _looks_like_table(run):
                blocks += 1
            run = []
    return blocks


def _looks_like_table(run: list[list[bool]]) -> bool:
    if len(run) < 2:
        return False
    width = max((len(row) for row in run), default=0)
    occupied_columns = sum(
        1 for column in range(width) if any(column < len(row) and row[column] for row in run)
    )
    return occupied_columns >= 2


def _preview_rows(categories: list[str], series: list[str], values: list[list[Any]], orientation: str) -> list[list[Any]]:
    if orientation in {"series_rows_categories_columns", "single_series_row_categories_columns"}:
        return [["", *categories], *[[series[i] if i < len(series) else "", *row] for i, row in enumerate(values)]]
    return [["", *series], *[[categories[i] if i < len(categories) else "", *row] for i, row in enumerate(values)]]


def _series_label(label: str, values: list[Any], formulas: list[Any], index: int) -> str:
    if label:
        return label
    if _looks_like_nps_row(values, formulas, []):
        return "NPS"
    return f"Serie {index + 1}"


def _looks_like_nps_row(values: list[Any], formulas: list[Any], previous: list[tuple]) -> bool:
    formula_text = " ".join(_text(value) for value in formulas)
    if re.search(r"\([A-Z]+\d+-[A-Z]+\d+\)\*100", formula_text, flags=re.IGNORECASE):
        return True
    numeric = [_to_number(value) for value in values if _to_number(value) is not None]
    if not numeric:
        return False
    previous_values = [
        number
        for item in previous
        for row_values in [item[1]]
        for number in (_to_number(value) for value in row_values)
        if number is not None
    ]
    previous_percent_like = previous_values and sum(1 for value in previous_values if -1 <= value <= 1) >= len(previous_values) * 0.65
    return previous_percent_like and sum(1 for value in numeric if abs(value) > 1) >= len(numeric) * 0.65


def _extract_metadata(rows: list[list[Any]]) -> dict[str, str]:
    metadata: dict[str, str] = {}
    aliases = {
        "GRAPH ID": "graph_id",
        "GRAPHID": "graph_id",
        "ID GRAFICO": "graph_id",
        "PPT TAG": "ppt_tag",
        "TAG PPT": "ppt_tag",
        "VARIAVEL": "variable",
        "VAR ANALISE": "variable",
    }
    for row in rows[:20]:
        cells = [_text(value) for value in row]
        for index, cell in enumerate(cells):
            if not cell:
                continue
            key_text = cell
            value_text = cells[index + 1] if index + 1 < len(cells) else ""
            if ":" in cell:
                key_text, value_text = [part.strip() for part in cell.split(":", 1)]
            key = aliases.get(_norm(key_text))
            if key and value_text:
                metadata[key] = value_text
    metadata.update(_extract_context_metadata(rows))
    return metadata


def _extract_context_metadata(rows: list[list[Any]]) -> dict[str, str]:
    header_row_index, header_start_col, _header_end_col = _find_header_row(rows)
    context: dict[str, str] = {}
    if header_row_index is None:
        header_row_index = min(len(rows), 3)

    title = _first_title_before_header(rows, header_row_index)
    row_group_label = _first_row_group_label(rows, header_row_index, header_start_col)
    context_text = _join_context([title, row_group_label])
    if title:
        context["table_title"] = title
    if row_group_label:
        context["row_group_label"] = row_group_label
    if context_text:
        context["context_text"] = context_text
        context["context_tokens"] = " ".join(_norm(context_text).split()[:24])
    return context


def _first_title_before_header(rows: list[list[Any]], header_row_index: int) -> str:
    for row in rows[: max(header_row_index, 0)]:
        values = [_text(value) for value in row if _text(value)]
        if len(values) == 1 and _looks_like_context_label(values[0]):
            return values[0]
    return ""


def _first_row_group_label(rows: list[list[Any]], header_row_index: int, header_start_col: int) -> str:
    if header_start_col <= 0:
        return ""
    for row in rows[header_row_index + 1 :]:
        leading_values = [_text(value) for value in row[:header_start_col] if _text(value)]
        if not leading_values:
            continue
        label = max(leading_values, key=len)
        if _looks_like_context_label(label):
            return label
    return ""


def _looks_like_context_label(value: Any) -> bool:
    text = _norm(value)
    if not text:
        return False
    if _to_number(text) is not None:
        return False
    if text in {"TOTAL", "SERIE", "SERIES", "CATEGORIA", "CATEGORIAS", "VALOR"}:
        return False
    return len(text) >= 4


def _join_context(values: list[str]) -> str:
    output: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = _text(value)
        key = _norm(text)
        if text and key not in seen:
            output.append(text)
            seen.add(key)
    return " | ".join(output)


def _period_score(values: list[Any]) -> float:
    clean = [_text(value) for value in values if _text(value)]
    if not clean:
        return 0.0
    return sum(1 for value in clean if _looks_like_period(value)) / len(clean)


def _looks_like_period(value: Any) -> bool:
    text = _norm(value)
    if re.fullmatch(r"(JAN|FEV|FEB|MAR|ABR|APR|MAI|MAY|JUN|JUL|AGO|AUG|SET|SEP|OUT|OCT|NOV|DEZ|DEC)\s*\d{2,4}", text):
        return True
    if re.fullmatch(r"\d{1,2}[/-]\d{2,4}", text):
        return True
    return False


def _is_likely_header_label(value: Any) -> bool:
    text = _norm(value)
    return text in {"SERIE", "SERIES", "CATEGORIA", "CATEGORIAS", "MES", "MESES"}


def _is_header_cell(value: Any) -> bool:
    text = _text(value)
    if not text:
        return False
    return _to_number(text) is None


def _has_numeric_or_text(values: list[Any]) -> bool:
    return any(_is_value(value) for value in values)


def _is_value(value: Any) -> bool:
    if isinstance(value, (int, float)):
        return True
    text = _text(value)
    if not text:
        return False
    return _to_number(text) is not None or len(text) > 0


def _to_number(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    text = _text(value)
    if not text:
        return None
    text = text.replace("%", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _graph_id(value: Any) -> str:
    text = _text(value)
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    text = re.sub(r"(?:^|[\s_-])(?:slide|s)0*\d+$", "", text, flags=re.IGNORECASE).strip()
    return re.sub(r"\D", "", text)


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return _date_period_label(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _date_period_label(value: datetime | date) -> str:
    months = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
    year_suffix = value.day if value.day >= 13 else value.year % 100
    return f"{months[value.month - 1]}/{year_suffix:02d}"


def _norm(value: Any) -> str:
    text = _text(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.upper()
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()
