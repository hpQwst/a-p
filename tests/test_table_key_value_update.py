from __future__ import annotations

from io import BytesIO
import unittest
import xml.etree.ElementTree as ET

import openpyxl

from ppt_automator.ppt_discovery import PptTarget
from ppt_automator.ppt_table_writer import update_table_slide_xml
from ppt_automator.table_normalizer import normalize_to_target
from ppt_automator.xlsx_parser import ParsedXlsxTable, parse_xlsx_table


NS = {
    "p": "http://schemas.openxmlformats.org/presentationml/2006/main",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
}


class TableKeyValueUpdateTests(unittest.TestCase):
    def test_key_value_xlsx_updates_second_column_without_none_text(self) -> None:
        source = parse_xlsx_table(_key_value_workbook(), file_name="t.xlsx")
        target = PptTarget(
            slide_index=0,
            slide_number=1,
            slide_path="ppt/slides/slide1.xml",
            shape_name="8282462966",
            shape_id="1",
            object_type="table",
            left_in=0,
            top_in=0,
            width_in=1,
            height_in=1,
            table_cells=[["Base:", ""], ["Total", ""], ["Natura", ""], ["Avon", ""]],
        )

        plan = normalize_to_target(target, source)
        updated = update_table_slide_xml(_slide_with_table(), target, plan)
        values = _table_values(updated)

        self.assertEqual(source.orientation, "key_value_rows")
        self.assertEqual(plan.values, [["Base:", ""], ["Total", 50], ["Natura", 20], ["Avon", 30]])
        self.assertEqual(values, [["Base:", ""], ["Total", "50"], ["Natura", "20"], ["Avon", "30"]])
        self.assertNotIn("None", updated.decode("utf-8"))
        self.assertEqual(_paragraph_child_names(updated, row_index=2, cell_index=2), ["pPr", "r", "endParaRPr"])

    def test_key_value_rows_keep_xlsx_order_and_names(self) -> None:
        source = parse_xlsx_table(_reordered_key_value_workbook(), file_name="t.xlsx")
        target = PptTarget(
            slide_index=0,
            slide_number=1,
            slide_path="ppt/slides/slide1.xml",
            shape_name="8282462966",
            shape_id="1",
            object_type="table",
            left_in=0,
            top_in=0,
            width_in=1,
            height_in=1,
            table_cells=[["Base:", ""], ["Total antigo", ""], ["Natura antiga", ""], ["Avon antiga", ""]],
        )

        plan = normalize_to_target(target, source)

        self.assertEqual(
            plan.values,
            [["Base:", ""], ["AVON oficial", 30], ["TOTAL oficial", 50], ["NATURA oficial", 20]],
        )

    def test_monthly_table_adds_rows_and_columns_without_growing_past_frame(self) -> None:
        source = ParsedXlsxTable(
            source_id="mensal",
            file_name="mensal.xlsx",
            sheet_name="Dados",
            orientation="series_rows_categories_columns",
            categories=["2025", "2026", "2027"],
            series=["NPS", "CSAT"],
            values=[[40, 41, 42], [80, 81, 82]],
        )
        target = PptTarget(
            slide_index=0,
            slide_number=1,
            slide_path="ppt/slides/slide1.xml",
            shape_name="Tabela mensal",
            shape_id="2",
            object_type="table",
            left_in=0,
            top_in=0,
            width_in=1,
            height_in=1,
            table_cells=[["2025", "2026"], ["", ""]],
        )

        plan = normalize_to_target(target, source)
        updated = update_table_slide_xml(_slide_with_small_month_table(), target, plan)

        self.assertEqual(
            plan.table_matrix,
            [[None, "2025", "2026", "2027"], ["NPS", 40, 41, 42], ["CSAT", 80, 81, 82]],
        )
        self.assertEqual(
            _table_values(updated),
            [["", "2025", "2026", "2027"], ["NPS", "40", "41", "42"], ["CSAT", "80", "81", "82"]],
        )
        self.assertEqual(sum(_grid_widths(updated)), 1_000)
        self.assertEqual(sum(_row_heights(updated)), 1_000)
        self.assertEqual(len(_grid_widths(updated)), 4)
        self.assertEqual(len(_row_heights(updated)), 3)
        self.assertEqual(updated.decode("utf-8").count("ABCDEF"), 12)


def _key_value_workbook() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Base:"
    ws["B1"] = None
    ws["A2"] = "Total"
    ws["B2"] = 50
    ws["A3"] = "Natura"
    ws["B3"] = 20
    ws["A4"] = "Avon"
    ws["B4"] = 30
    data = BytesIO()
    wb.save(data)
    return data.getvalue()


def _reordered_key_value_workbook() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["", "Valor"])
    ws.append(["AVON oficial", 30])
    ws.append(["TOTAL oficial", 50])
    ws.append(["NATURA oficial", 20])
    data = BytesIO()
    wb.save(data)
    return data.getvalue()


def _slide_with_table() -> bytes:
    rows = "\n".join(
        f"""
        <a:tr h="370840">
          <a:tc><a:txBody><a:bodyPr/><a:lstStyle/><a:p><a:r><a:t>{label}</a:t></a:r></a:p></a:txBody></a:tc>
          <a:tc><a:txBody><a:bodyPr/><a:lstStyle/><a:p><a:pPr/><a:endParaRPr lang="pt-BR" sz="800"/></a:p></a:txBody></a:tc>
        </a:tr>
        """
        for label in ["Base:", "Total", "Natura", "Avon"]
    )
    return f"""
    <p:sld xmlns:p="{NS['p']}" xmlns:a="{NS['a']}">
      <p:cSld>
        <p:spTree>
          <p:graphicFrame>
            <p:nvGraphicFramePr><p:cNvPr id="1" name="8282462966"/></p:nvGraphicFramePr>
            <a:graphic><a:graphicData><a:tbl>{rows}</a:tbl></a:graphicData></a:graphic>
          </p:graphicFrame>
        </p:spTree>
      </p:cSld>
    </p:sld>
    """.encode("utf-8")


def _slide_with_small_month_table() -> bytes:
    rows = "\n".join(
        f"""
        <a:tr h="500">
          <a:tc><a:txBody><a:bodyPr/><a:lstStyle/><a:p><a:r><a:t>{left}</a:t></a:r></a:p></a:txBody><a:tcPr><a:solidFill><a:srgbClr val="ABCDEF"/></a:solidFill></a:tcPr></a:tc>
          <a:tc><a:txBody><a:bodyPr/><a:lstStyle/><a:p><a:r><a:t>{right}</a:t></a:r></a:p></a:txBody><a:tcPr><a:solidFill><a:srgbClr val="ABCDEF"/></a:solidFill></a:tcPr></a:tc>
        </a:tr>
        """
        for left, right in [("2025", "2026"), ("", "")]
    )
    return f"""
    <p:sld xmlns:p="{NS['p']}" xmlns:a="{NS['a']}">
      <p:cSld>
        <p:spTree>
          <p:graphicFrame>
            <p:nvGraphicFramePr><p:cNvPr id="2" name="Tabela mensal"/></p:nvGraphicFramePr>
            <p:xfrm><a:off x="0" y="0"/><a:ext cx="1000" cy="1000"/></p:xfrm>
            <a:graphic><a:graphicData><a:tbl>
              <a:tblGrid><a:gridCol w="500"/><a:gridCol w="500"/></a:tblGrid>
              {rows}
            </a:tbl></a:graphicData></a:graphic>
          </p:graphicFrame>
        </p:spTree>
      </p:cSld>
    </p:sld>
    """.encode("utf-8")


def _table_values(slide_xml: bytes) -> list[list[str]]:
    root = ET.fromstring(slide_xml)
    return [
        ["".join(text.text or "" for text in cell.findall(".//a:t", NS)) for cell in row.findall("./a:tc", NS)]
        for row in root.findall(".//a:tbl/a:tr", NS)
    ]


def _grid_widths(slide_xml: bytes) -> list[int]:
    root = ET.fromstring(slide_xml)
    return [int(column.attrib["w"]) for column in root.findall(".//a:tblGrid/a:gridCol", NS)]


def _row_heights(slide_xml: bytes) -> list[int]:
    root = ET.fromstring(slide_xml)
    return [int(row.attrib["h"]) for row in root.findall(".//a:tbl/a:tr", NS)]


def _paragraph_child_names(slide_xml: bytes, row_index: int, cell_index: int) -> list[str]:
    root = ET.fromstring(slide_xml)
    row = root.findall(".//a:tbl/a:tr", NS)[row_index - 1]
    cell = row.findall("./a:tc", NS)[cell_index - 1]
    paragraph = cell.find("./a:txBody/a:p", NS)
    assert paragraph is not None
    return [child.tag.split("}", 1)[-1] for child in list(paragraph)]


if __name__ == "__main__":
    unittest.main()
