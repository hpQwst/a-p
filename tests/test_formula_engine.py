from __future__ import annotations

from io import BytesIO
import unittest

import openpyxl

from ppt_automator.core import FormulaCalculationError, prepare_workbook_values


class FormulaEngineTests(unittest.TestCase):
    def test_calculates_supported_formulas_without_office(self) -> None:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet["A1"] = 10
        worksheet["A2"] = 5
        worksheet["A3"] = "=SUM(A1:A2)"
        worksheet["A4"] = "=IF(A3>10,A3*2,0)"

        output = prepare_workbook_values(_workbook_bytes(workbook), formula_mode="auto")
        result = openpyxl.load_workbook(BytesIO(output), data_only=True)
        try:
            self.assertEqual(result.active["A3"].value, 15)
            self.assertEqual(result.active["A4"].value, 30)
        finally:
            result.close()

    def test_calculates_sumproduct_used_by_real_datasources(self) -> None:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet["A1"] = 2
        worksheet["A2"] = 3
        worksheet["B1"] = 10
        worksheet["B2"] = 20
        worksheet["C1"] = "=SUMPRODUCT(A1:A2,B1:B2)"

        output = prepare_workbook_values(_workbook_bytes(workbook), formula_mode="auto")
        result = openpyxl.load_workbook(BytesIO(output), data_only=True)
        try:
            self.assertEqual(result.active["C1"].value, 80)
        finally:
            result.close()

    def test_calculates_sqrt_used_by_mb2_datasources(self) -> None:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet["A1"] = 100
        worksheet["A2"] = "=1.96*SQRT(0.25/A1)*100"
        worksheet["A3"] = "=RAIZ(81)"

        output = prepare_workbook_values(_workbook_bytes(workbook), formula_mode="auto")
        result = openpyxl.load_workbook(BytesIO(output), data_only=True)
        try:
            self.assertAlmostEqual(result.active["A2"].value, 9.8)
            self.assertEqual(result.active["A3"].value, 9)
        finally:
            result.close()

    def test_rejects_unknown_formula_instead_of_guessing_a_value(self) -> None:
        workbook = openpyxl.Workbook()
        workbook.active["A1"] = "=VLOOKUP(1,A1:B2,2,FALSE)"

        with self.assertRaisesRegex(FormulaCalculationError, "nao suportada ou insegura"):
            prepare_workbook_values(_workbook_bytes(workbook), formula_mode="internal")

    def test_rejects_expression_injection(self) -> None:
        workbook = openpyxl.Workbook()
        workbook.active["A1"] = '=__import__("os").system("echo unsafe")'

        with self.assertRaises(FormulaCalculationError):
            prepare_workbook_values(_workbook_bytes(workbook), formula_mode="internal")


def _workbook_bytes(workbook: openpyxl.Workbook) -> bytes:
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


if __name__ == "__main__":
    unittest.main()
