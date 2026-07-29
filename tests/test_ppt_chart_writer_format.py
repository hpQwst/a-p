from __future__ import annotations

from dataclasses import replace
from io import BytesIO
import unittest
import xml.etree.ElementTree as ET
from zipfile import ZIP_DEFLATED, ZipFile

import openpyxl

from ppt_automator.ppt_chart_writer import (
    _chart_value_text,
    _updated_chart_xml_bytes,
    resolved_series_number_formats,
)
from ppt_automator.ppt_discovery import NS, PptTarget
from ppt_automator.table_normalizer import TransformPlan
from ppt_automator.xlsx_parser import ParsedXlsxTable, parse_xlsx_table


class PptChartWriterFormatTests(unittest.TestCase):
    def test_decimal_visual_format_does_not_scale_value_as_percentage(self) -> None:
        self.assertEqual(_chart_value_text(15.990453460620525, numeric=True, number_format="0.0"), "15.9904534606")

    def test_value_is_written_verbatim_even_with_percent_format(self) -> None:
        # Regra: o valor gravado e sempre verbatim (precisao total para o "Editar
        # dados"); um formatCode "%" so afeta a EXIBICAO no PowerPoint, nunca a
        # escala do numero armazenado. Nao dividimos mais por 100.
        self.assertEqual(_chart_value_text(15.990453460620525, numeric=True, number_format="0.0%"), "15.9904534606")

    def test_textual_percent_is_stripped_without_rescaling(self) -> None:
        self.assertEqual(_chart_value_text("30,8%", numeric=True, number_format="0.0%"), "30.8")

    def test_mixed_chart_preserves_percent_bars_and_numeric_line(self) -> None:
        target = _mixed_chart_target()
        plan = _mixed_chart_plan(target)

        self.assertEqual(
            resolved_series_number_formats(target, plan),
            ["0%", "0%", "0%", "#,##0.0"],
        )

    def test_manual_override_wins_per_series(self) -> None:
        target = _mixed_chart_target()
        plan = replace(
            _mixed_chart_plan(target),
            series_format_overrides={"Detrator": "number", "NPS": "percent"},
        )

        self.assertEqual(
            resolved_series_number_formats(target, plan),
            ["0.0", "0%", "0%", "0%"],
        )

    def test_xlsx_number_formats_are_detected_per_series(self) -> None:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["", "Nov/25", "Dez/25"])
        sheet.append(["Detrator", 0.20, 0.19])
        sheet.append(["Promotor", 0.65, 0.67])
        sheet.append(["NPS", 44.7, 50.5])
        for row in (2, 3):
            for column in (2, 3):
                sheet.cell(row, column).number_format = "0%"
        for column in (2, 3):
            sheet.cell(4, column).number_format = "#,##0.0"
        output = BytesIO()
        workbook.save(output)

        parsed = parse_xlsx_table(output.getvalue(), file_name="dados.xlsx")

        self.assertEqual(parsed.series, ["Detrator", "Promotor", "NPS"])
        self.assertEqual(parsed.series_number_formats, ["0%", "0%", "#,##0.0"])

    def test_chart_xml_receives_distinct_formats_without_scaling_values(self) -> None:
        target = _mixed_chart_target()
        plan = _mixed_chart_plan(target)
        package = BytesIO()
        with ZipFile(package, "w", ZIP_DEFLATED) as zf:
            zf.writestr(target.chart_xml, _chart_xml_with_four_series())
        package.seek(0)

        with ZipFile(package, "r") as zf:
            updated = _updated_chart_xml_bytes(zf, target, plan)

        root = ET.fromstring(updated)
        series = root.findall(".//c:ser", NS)
        label_formats = [
            ser.find("./c:dLbls/c:numFmt", NS).attrib["formatCode"]
            for ser in series
        ]
        cache_formats = [
            ser.find("./c:val/c:numRef/c:numCache/c:formatCode", NS).text
            for ser in series
        ]
        first_values = [
            ser.find("./c:val/c:numRef/c:numCache/c:pt/c:v", NS).text
            for ser in series
        ]

        self.assertEqual(label_formats, ["0%", "0%", "0%", "#,##0.0"])
        self.assertEqual(cache_formats, label_formats)
        self.assertEqual(first_values, ["0.2", "0.15", "0.65", "44.7"])


def _mixed_chart_target() -> PptTarget:
    return PptTarget(
        slide_index=0,
        slide_number=1,
        slide_path="ppt/slides/slide1.xml",
        shape_name="Grafico 1",
        shape_id="1",
        object_type="chart",
        left_in=0,
        top_in=0,
        width_in=1,
        height_in=1,
        chart_xml="ppt/charts/chart1.xml",
        sheet_name="Sheet1",
        expected_orientation="series_rows_categories_columns",
        expected_categories=["Nov/25", "Dez/25"],
        expected_series=["Detrator", "Neutro", "Promotor", ""],
        expected_values=[
            [0.19, 0.20],
            [0.13, 0.15],
            [0.68, 0.65],
            [49.0, 45.0],
        ],
        series_value_formats=["0%", "0%", "0%", "#,##0.0"],
    )


def _mixed_chart_plan(target: PptTarget) -> TransformPlan:
    source = ParsedXlsxTable(
        source_id="dados",
        file_name="dados.xlsx",
        sheet_name="Sheet1",
        orientation="series_rows_categories_columns",
        categories=["Nov/25", "Dez/25"],
        series=["Detrator", "Neutro", "Promotor", "NPS"],
        values=[
            [0.20, 0.19],
            [0.15, 0.16],
            [0.65, 0.67],
            [44.7, 50.5],
        ],
        series_number_formats=["###0%", "###0%", "###0%", "#,##0.0"],
    )
    return TransformPlan(
        target=target,
        datasource=source,
        action="align",
        orientation_xlsx=source.orientation,
        orientation_ppt=target.expected_orientation,
        categories=source.categories,
        series=source.series,
        values=source.values,
        confidence=1.0,
        reason="teste",
    )


def _chart_xml_with_four_series() -> str:
    series_xml = "".join(
        f'<c:ser><c:idx val="{index}"/><c:order val="{index}"/></c:ser>'
        for index in range(4)
    )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<c:chartSpace xmlns:c="http://schemas.openxmlformats.org/drawingml/2006/chart">'
        f"<c:chart><c:plotArea><c:barChart>{series_xml}</c:barChart></c:plotArea></c:chart>"
        "</c:chartSpace>"
    )


if __name__ == "__main__":
    unittest.main()
