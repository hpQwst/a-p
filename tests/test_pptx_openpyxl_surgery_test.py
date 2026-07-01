from __future__ import annotations

from io import BytesIO
import unittest

import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo

from scripts.pptx_openpyxl_surgery_test import update_embedded_workbook_openpyxl
from scripts.pptx_openpyxl_surgery_test import update_embedded_workbook_raw_values_only
from scripts.pptx_openpyxl_surgery_test import update_embedded_workbook_sheet_values_only
from scripts.pptx_openpyxl_surgery_test import update_embedded_workbook_xml


class PptxOpenpyxlSurgeryScriptTests(unittest.TestCase):
    def test_updates_existing_sheet_and_resizes_existing_table(self) -> None:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Planilha1"
        for row in [[" ", "Old"], ["A", 1]]:
            worksheet.append(row)
        table = Table(displayName="Tabela1", ref="A1:B2")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        worksheet.add_table(table)
        stream = BytesIO()
        workbook.save(stream)
        workbook.close()

        updated = update_embedded_workbook_openpyxl(
            stream.getvalue(),
            "Planilha1",
            [[" ", "TOTAL", "REDE 1"], ["Cristal", 15.990453460620525, 21.598968407479045]],
        )

        result = openpyxl.load_workbook(BytesIO(updated), data_only=True)
        self.assertEqual(result.sheetnames, ["Planilha1"])
        ws = result["Planilha1"]
        self.assertEqual(ws["A1"].value, " ")
        self.assertEqual(ws["B1"].value, "TOTAL")
        self.assertEqual(ws["A2"].value, "Cristal")
        self.assertAlmostEqual(ws["B2"].value, 15.990453460620525)
        self.assertEqual(ws.tables["Tabela1"].ref, "A1:C2")
        result.close()

    def test_sheet_values_writer_only_updates_existing_worksheet_values(self) -> None:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Planilha1"
        for row in [[" ", "Old"], ["A", 1]]:
            worksheet.append(row)
        table = Table(displayName="Tabela1", ref="A1:B2")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        worksheet.add_table(table)
        stream = BytesIO()
        workbook.save(stream)
        workbook.close()
        original = stream.getvalue()

        updated = update_embedded_workbook_sheet_values_only(
            original,
            "Planilha1",
            [[" ", "TOTAL"], ["Cristal", 15.990453460620525]],
        )

        from zipfile import ZipFile

        with ZipFile(BytesIO(original)) as before, ZipFile(BytesIO(updated)) as after:
            self.assertEqual(before.namelist(), after.namelist())
            changed = [name for name in before.namelist() if before.read(name) != after.read(name)]
            self.assertEqual(changed, ["xl/worksheets/sheet1.xml"])

        result = openpyxl.load_workbook(BytesIO(updated), data_only=True)
        ws = result["Planilha1"]
        self.assertEqual(ws["B1"].value, "Old")
        self.assertEqual(ws["A2"].value, "A")
        self.assertAlmostEqual(ws["B2"].value, 15.990453460620525)
        self.assertEqual(ws.tables["Tabela1"].ref, "A1:B2")
        result.close()

    def test_raw_values_writer_only_replaces_numeric_v_text(self) -> None:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Planilha1"
        for row in [[" ", "Old"], ["A", 1]]:
            worksheet.append(row)
        table = Table(displayName="Tabela1", ref="A1:B2")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        worksheet.add_table(table)
        stream = BytesIO()
        workbook.save(stream)
        workbook.close()
        original = stream.getvalue()

        updated = update_embedded_workbook_raw_values_only(
            original,
            "Planilha1",
            [[" ", "TOTAL"], ["Cristal", 15.990453460620525]],
        )

        from zipfile import ZipFile

        with ZipFile(BytesIO(original)) as before, ZipFile(BytesIO(updated)) as after:
            self.assertEqual(before.namelist(), after.namelist())
            changed = [name for name in before.namelist() if before.read(name) != after.read(name)]
            self.assertEqual(changed, ["xl/worksheets/sheet1.xml"])
            before_sheet = before.read("xl/worksheets/sheet1.xml")
            after_sheet = after.read("xl/worksheets/sheet1.xml")
            self.assertIn(b'<c r="A2" t="inlineStr"><is><t>A</t></is></c>', after_sheet)
            self.assertNotEqual(before_sheet, after_sheet)

        result = openpyxl.load_workbook(BytesIO(updated), data_only=True)
        ws = result["Planilha1"]
        self.assertEqual(ws["B1"].value, "Old")
        self.assertEqual(ws["A2"].value, "A")
        self.assertAlmostEqual(ws["B2"].value, 15.990453460620525)
        self.assertEqual(ws.tables["Tabela1"].ref, "A1:B2")
        result.close()

    def test_xml_writer_preserves_workbook_parts_and_updates_values(self) -> None:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Planilha1"
        for row in [[" ", "Old"], ["A", 1]]:
            worksheet.append(row)
        table = Table(displayName="Tabela1", ref="A1:B2")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        worksheet.add_table(table)
        stream = BytesIO()
        workbook.save(stream)
        workbook.close()
        original = stream.getvalue()

        updated = update_embedded_workbook_xml(
            original,
            "Planilha1",
            [[" ", "TOTAL", "REDE 1"], ["Cristal", 15.990453460620525, 21.598968407479045]],
        )

        from zipfile import ZipFile

        with ZipFile(BytesIO(original)) as before, ZipFile(BytesIO(updated)) as after:
            self.assertEqual(before.namelist(), after.namelist())
            changed = [name for name in before.namelist() if before.read(name) != after.read(name)]
            self.assertIn("xl/worksheets/sheet1.xml", changed)
            self.assertIn("xl/tables/table1.xml", changed)
            self.assertLessEqual(set(changed), {"xl/worksheets/sheet1.xml", "xl/tables/table1.xml", "xl/sharedStrings.xml"})

        result = openpyxl.load_workbook(BytesIO(updated), data_only=True)
        ws = result["Planilha1"]
        self.assertEqual(ws["A1"].value, " ")
        self.assertEqual(ws["B1"].value, "TOTAL")
        self.assertEqual(ws["A2"].value, "Cristal")
        self.assertAlmostEqual(ws["C2"].value, 21.598968407479045)
        self.assertEqual(ws.tables["Tabela1"].ref, "A1:C2")
        result.close()


if __name__ == "__main__":
    unittest.main()
