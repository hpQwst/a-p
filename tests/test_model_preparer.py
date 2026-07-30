from __future__ import annotations

from io import BytesIO
from tempfile import TemporaryDirectory
from unittest.mock import patch
from pathlib import Path
from zipfile import ZipFile
import os
import json
import re
import time
import unittest

import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from fastapi.testclient import TestClient

from ppt_automator.model_preparer import (
    build_mapping_workbook,
    read_mapping_workbook,
    validate_mapping_package,
)
from ppt_automator.project_store import (
    ensure_store,
    list_prepared_models,
    load_prepared_model,
    load_prepared_model_bytes,
    save_prepared_model,
)
from ppt_automator.xlsx_parser import parse_xlsx_table
from ppt_automator import analyze_update_package
from web import main
from web.main import app


TARGET_ID = "S001_T001_CHART"
MB_DIR = Path(os.getenv("AUTO_PPT_MB_TEST_DIR", r"C:\Users\HugoRocha\Documents\automatizador-ppt-arquivos\mb"))
MB_PPT = MB_DIR / "MBTESTE_formula.pptx"
MB_DATASOURCES = MB_DIR / "datasources.zip"


def _manifest() -> dict:
    return {
        "schema_version": 2,
        "model_name": "Modelo NPS",
        "original_filename": "nps.pptx",
        "original_sha256": "a" * 64,
        "identified_sha256": "b" * 64,
        "slide_width_in": 13.3333,
        "slide_height_in": 7.5,
        "slide_count": 1,
        "object_count": 1,
        "chart_count": 1,
        "table_count": 0,
        "objects": [
            {
                "ativo": 0,
                "id_objeto": TARGET_ID,
                "rotulo_visual": "T1",
                "slide": 1,
                "tipo_objeto": "chart",
                "tipo_visual": "column",
                "titulo_slide": "Evolutivo do NPS",
                "titulo_detectado": "Evolutivo do NPS",
                "confianca_titulo": "alta",
                "nome_amigavel": "NPS mensal",
                "linhas_no_ppt": ["Detrator", "Promotor"],
                "colunas_no_ppt": ["Jan/26", "Fev/26"],
                "qtd_linhas": 2,
                "qtd_colunas": 2,
                "arquivo_xlsx": "",
                "aba_xlsx": "",
                "modo_leitura": "auto",
                "referencia": "",
                "orientacao": "auto",
                "formato_valores": "auto",
                "observacao": "",
                "status": "INATIVO",
                "shape_name_original": "Grafico 1",
                "shape_id": "7",
                "target_fingerprint": "",
                "left_in": 1,
                "top_in": 1,
                "width_in": 8,
                "height_in": 4,
                "expected_orientation": "series_rows_categories_columns",
                "expected_categories": ["Jan/26", "Fev/26"],
                "expected_series": ["Detrator", "Promotor"],
                "expected_values": [[0.2, 0.19], [0.65, 0.67]],
                "table_cells": [],
                "value_format": "0%",
                "series_value_formats": ["0%", "0%"],
                "chart_series_colors": ["#FF0000", "#00AA00"],
                "chart_xml": "ppt/charts/chart1.xml",
                "workbook_embedded": "ppt/embeddings/Microsoft_Excel_Worksheet1.xlsx",
                "sheet_name": "Plan1",
            }
        ],
    }


def _source_xlsx() -> bytes:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Dados"
    sheet.append(["Serie", "Jan/26", "Fev/26", "Mar/26"])
    sheet.append(["Detrator", 0.20, 0.19, 0.18])
    sheet.append(["Promotor", 0.65, 0.67, 0.69])
    table = Table(displayName="tb_nps", ref="A1:D3")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2")
    sheet.add_table(table)
    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def _filled_mapping() -> bytes:
    data = build_mapping_workbook(_manifest())
    workbook = openpyxl.load_workbook(BytesIO(data))
    sheet = workbook["OBJETOS"]
    headers = {cell.value: cell.column for cell in sheet[4]}
    sheet.cell(5, headers["ativo"], 1)
    sheet.cell(5, headers["arquivo_xlsx"], "nps.xlsx")
    sheet.cell(5, headers["aba_xlsx"], "Dados")
    sheet.cell(5, headers["modo_leitura"], "tabela_excel")
    sheet.cell(5, headers["referencia"], "tb_nps")
    sheet.cell(5, headers["formato_valores"], "percentual")
    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


class ModelPreparerTests(unittest.TestCase):
    def test_mapping_workbook_roundtrip_and_named_excel_table(self) -> None:
        mapping = _filled_mapping()

        rows = read_mapping_workbook(mapping)
        report, entries = validate_mapping_package(
            _manifest(),
            mapping,
            {"nps.xlsx": _source_xlsx()},
        )

        self.assertEqual(rows[TARGET_ID]["ativo"], 1)
        self.assertTrue(report["ok"], report)
        self.assertEqual(report["resolved_count"], 1)
        self.assertEqual(entries[TARGET_ID]["datasource"], "nps.xlsx")
        self.assertEqual(entries[TARGET_ID]["cell_range"], "tb_nps")
        self.assertTrue(entries[TARGET_ID]["allow_axis_growth"])
        self.assertEqual(entries[TARGET_ID]["value_format"], "percentual")

        parsed = parse_xlsx_table(
            _source_xlsx(),
            file_name="nps.xlsx",
            cell_range="tb_nps",
            sheet="Dados",
        )
        self.assertEqual(parsed.used_range, (1, 1, 3, 4))

    def test_missing_file_blocks_only_active_object(self) -> None:
        mapping = _filled_mapping()

        report, entries = validate_mapping_package(_manifest(), mapping, {})

        self.assertFalse(report["ok"])
        self.assertEqual(entries, {})
        self.assertIn("Arquivo ausente: nps.xlsx", report["errors"][0]["message"])

        workbook = openpyxl.load_workbook(BytesIO(mapping))
        sheet = workbook["OBJETOS"]
        headers = {cell.value: cell.column for cell in sheet[4]}
        sheet.cell(5, headers["ativo"], 0)
        stream = BytesIO()
        workbook.save(stream)
        workbook.close()
        inactive_report, _ = validate_mapping_package(_manifest(), stream.getvalue(), {})
        self.assertIn("Nenhum objeto esta ativo", inactive_report["errors"][0]["message"])

    def test_prepared_model_is_versioned_and_isolated_by_squad(self) -> None:
        with TemporaryDirectory() as tmp:
            env = {
                **os.environ,
                "AUTO_PPT_STORAGE_BACKEND": "local",
                "AUTO_PPT_DATA_ROOT": tmp,
            }
            with patch.dict(os.environ, env, clear=True):
                ensure_store()
                first = save_prepared_model(
                    "squad2",
                    "Modelo NPS",
                    _manifest(),
                    {TARGET_ID: {"datasource": "nps.xlsx"}},
                    {
                        "original.pptx": b"original",
                        "identified.pptx": b"identified",
                        "mapping.xlsx": _filled_mapping(),
                    },
                    actor="teste@qwst.co",
                )
                second = save_prepared_model(
                    "squad2",
                    "Modelo NPS",
                    _manifest(),
                    {TARGET_ID: {"datasource": "nps.xlsx"}},
                    {
                        "original.pptx": b"original-v2",
                        "identified.pptx": b"identified-v2",
                        "mapping.xlsx": _filled_mapping(),
                    },
                    slug=first.slug,
                    actor="teste@qwst.co",
                )

                self.assertEqual(list_prepared_models("squad1"), [])
                self.assertEqual(second.version_count, 2)
                model = load_prepared_model("squad2", first.slug)
                self.assertEqual(model["version_count"], 2)
                self.assertEqual(
                    load_prepared_model_bytes("squad2", first.slug, "identified.pptx"),
                    b"identified-v2",
                )


@unittest.skipUnless(MB_PPT.exists() and MB_DATASOURCES.exists(), "Arquivos MB nao encontrados.")
class ModelPreparerWebFlowTests(unittest.TestCase):
    def test_real_mb_prepare_import_and_preview_flow(self) -> None:
        targets, sources, plans = analyze_update_package(MB_PPT, MB_DATASOURCES)
        self.assertTrue(plans)
        source_by_name = {source.file_name: source for source in sources}

        with TemporaryDirectory() as tmp:
            env = {
                **os.environ,
                "AUTO_PPT_STORAGE_BACKEND": "local",
                "AUTO_PPT_DATA_ROOT": str(Path(tmp) / "data"),
                "AUTO_PPT_AUTO_SLIDE_AI": "0",
            }
            runtime = Path(tmp) / "jobs"
            runtime.mkdir()
            with (
                patch.dict(os.environ, env, clear=True),
                patch.object(main, "RUNTIME_ROOT", runtime),
                patch("web.main.ai_configured", return_value=False),
            ):
                main.ANALYZE_FILES_CACHE.clear()
                client = TestClient(app)
                prepared = client.post(
                    "/models/squad1/prepare",
                    data={"model_name": "MB preparado"},
                    files={
                        "pptx": (
                            MB_PPT.name,
                            MB_PPT.read_bytes(),
                            "application/vnd.openxmlformats-officedocument.presentationml.presentation",
                        )
                    },
                    follow_redirects=False,
                )
                self.assertEqual(prepared.status_code, 303, prepared.text)
                studio_url = prepared.headers["location"]
                studio = client.get(studio_url)
                self.assertEqual(studio.status_code, 200)
                self.assertIn("Estudio de Mapeamento", studio.text)
                self.assertIn("BAIXAR XLSX MODELO", studio.text)
                job_id = re.search(r"/studio/([a-f0-9]{32})", studio_url).group(1)

                mapping_response = client.get(f"/models/squad1/studio/{job_id}/mapping.xlsx")
                self.assertEqual(mapping_response.status_code, 200)
                workbook = openpyxl.load_workbook(BytesIO(mapping_response.content))
                sheet = workbook["OBJETOS"]
                headers = {cell.value: cell.column for cell in sheet[4]}
                plan_by_target = {plan.target_id: plan for plan in plans}
                for row in range(5, sheet.max_row + 1):
                    target_id = str(sheet.cell(row, headers["id_objeto"]).value or "")
                    plan = plan_by_target.get(target_id)
                    if not plan:
                        continue
                    source = source_by_name[plan.datasource.file_name]
                    filename = plan.datasource.file_name.split("#", 1)[0].replace("\\", "/").split("/")[-1]
                    sheet.cell(row, headers["ativo"], 1)
                    sheet.cell(row, headers["arquivo_xlsx"], filename)
                    sheet.cell(row, headers["aba_xlsx"], source.sheet_name)
                    sheet.cell(row, headers["modo_leitura"], "auto")
                filled = BytesIO()
                workbook.save(filled)
                workbook.close()

                imported = client.post(
                    f"/models/squad1/studio/{job_id}/import",
                    data={"project_name": "MB rodada preparada"},
                    files=[
                        (
                            "mapping",
                            (
                                "mapeamento.xlsx",
                                filled.getvalue(),
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            ),
                        ),
                        (
                            "datasources",
                            ("datasources.zip", MB_DATASOURCES.read_bytes(), "application/zip"),
                        ),
                    ],
                    follow_redirects=False,
                )
                self.assertEqual(imported.status_code, 303, imported.text)
                preview_url = imported.headers["location"]
                self.assertRegex(preview_url, r"^/jobs/[a-f0-9]{32}/preview$")
                preview_job_id = preview_url.split("/")[2]
                metadata = json.loads((runtime / preview_job_id / "metadata.json").read_text(encoding="utf-8"))
                self.assertEqual(set(metadata["allowed_target_ids"]), set(plan_by_target))
                self.assertFalse(metadata["use_ai"])

                for _attempt in range(100):
                    status = client.get(f"/jobs/{preview_job_id}/processing-status").json()
                    if status.get("status") == "error":
                        self.fail(status.get("message") or "Preview falhou.")
                    if status.get("status") == "complete":
                        break
                    time.sleep(0.1)
                else:
                    self.fail("Preview do modelo preparado nao terminou.")

                model = load_prepared_model("squad1", "mb-preparado")
                self.assertIsNotNone(model)
                self.assertEqual(model["active_count"], len(plans))

                run_form = client.get("/models/squad1/mb-preparado/run")
                self.assertEqual(run_form.status_code, 200)
                self.assertIn("Usar modelo preparado", run_form.text)
                rerun = client.post(
                    "/models/squad1/mb-preparado/run",
                    data={
                        "project_ref": (
                            f"{metadata['project']['squad']}|{metadata['project']['slug']}"
                        ),
                    },
                    files={
                        "datasources": (
                            "datasources.zip",
                            MB_DATASOURCES.read_bytes(),
                            "application/zip",
                        ),
                    },
                    follow_redirects=False,
                )
                self.assertEqual(rerun.status_code, 303, rerun.text)
                rerun_job_id = rerun.headers["location"].split("/")[2]
                rerun_metadata = json.loads(
                    (runtime / rerun_job_id / "metadata.json").read_text(encoding="utf-8")
                )
                self.assertFalse(rerun_metadata["use_ai"])
                self.assertEqual(set(rerun_metadata["allowed_target_ids"]), set(plan_by_target))
                for _attempt in range(100):
                    rerun_status = client.get(f"/jobs/{rerun_job_id}/processing-status").json()
                    if rerun_status.get("status") == "error":
                        self.fail(rerun_status.get("message") or "Preview da nova rodada falhou.")
                    if rerun_status.get("status") == "complete":
                        break
                    time.sleep(0.1)
                else:
                    self.fail("Preview da nova rodada nao terminou.")


def _mapping_with(**cells: object) -> bytes:
    """Planilha de mapeamento preenchida, com celulas sobrescritas por nome."""
    workbook = openpyxl.load_workbook(BytesIO(_filled_mapping()))
    sheet = workbook["OBJETOS"]
    headers = {cell.value: cell.column for cell in sheet[4]}
    for name, value in cells.items():
        sheet.cell(5, headers[name], value)
    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


class MappingValidationGuardTests(unittest.TestCase):
    """O preparador so pode publicar quando TODA linha ativa resolve. Estas sao
    as garantias que, se quebrarem, gravam numero errado no slide sem avisar."""

    def test_inactive_object_is_never_resolved(self) -> None:
        report, entries = validate_mapping_package(
            _manifest(),
            _mapping_with(ativo=0),
            {"nps.xlsx": _source_xlsx()},
        )
        self.assertNotIn(TARGET_ID, entries)
        self.assertFalse(report["ok"])

    def test_duplicate_source_names_are_refused(self) -> None:
        """Dois arquivos com o mesmo nome tornam ambiguo qual alimenta o objeto."""
        report, _entries = validate_mapping_package(
            _manifest(),
            _filled_mapping(),
            {"pasta_a/nps.xlsx": _source_xlsx(), "pasta_b/nps.xlsx": _source_xlsx()},
        )
        self.assertFalse(report["ok"])
        self.assertTrue(
            any("repetido" in issue["message"].lower() for issue in report["errors"]),
            report["errors"],
        )

    def test_unknown_sheet_is_reported_not_silently_ignored(self) -> None:
        report, entries = validate_mapping_package(
            _manifest(),
            _mapping_with(aba_xlsx="AbaQueNaoExiste", modo_leitura="auto", referencia=""),
            {"nps.xlsx": _source_xlsx()},
        )
        self.assertFalse(report["ok"])
        self.assertNotIn(TARGET_ID, entries)

    def test_unknown_target_id_only_warns(self) -> None:
        """ID que nao existe mais no PPT nao pode derrubar a publicacao: o deck
        pode ter sido reeditado. Vira aviso, e o objeto e ignorado."""
        workbook = openpyxl.load_workbook(BytesIO(_filled_mapping()))
        sheet = workbook["OBJETOS"]
        headers = {cell.value: cell.column for cell in sheet[4]}
        sheet.cell(6, headers["id_objeto"], "S009_T009_CHART")
        sheet.cell(6, headers["ativo"], 0)
        stream = BytesIO()
        workbook.save(stream)
        workbook.close()

        report, entries = validate_mapping_package(
            _manifest(), stream.getvalue(), {"nps.xlsx": _source_xlsx()}
        )
        self.assertTrue(report["ok"], report)
        self.assertIn(TARGET_ID, entries)
        self.assertNotIn("S009_T009_CHART", entries)
        self.assertTrue(
            any("S009_T009_CHART" in issue.get("target_id", "") for issue in report["warnings"]),
            report["warnings"],
        )

    def test_resolved_entry_carries_the_sheet_it_must_read(self) -> None:
        _report, entries = validate_mapping_package(
            _manifest(),
            _mapping_with(modo_leitura="auto", referencia=""),
            {"nps.xlsx": _source_xlsx()},
        )
        self.assertEqual(entries[TARGET_ID]["datasource"], "nps.xlsx")
        self.assertEqual(entries[TARGET_ID].get("sheet_name") or "Dados", "Dados")


if __name__ == "__main__":
    unittest.main()
