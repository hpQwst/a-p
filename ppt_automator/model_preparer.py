from __future__ import annotations

from dataclasses import replace
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable
from zipfile import ZIP_DEFLATED, ZipFile
import hashlib
import json
import re
import unicodedata
import xml.etree.ElementTree as ET

import openpyxl
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from .learned_mapping import mapping_entry_learning_fields
from .ppt_discovery import PptTarget, discover_ppt_targets
from .ppt_target_renamer import rename_targets_in_slide_xml
from .table_normalizer import normalize_to_target
from .target_labeler import target_aliases, target_fingerprint, visual_label
from .xlsx_parser import parse_xlsx_table, workbook_sheet_names


MODEL_SCHEMA_VERSION = 2
MAPPING_SHEET = "OBJETOS"
MAPPING_HEADER_ROW = 4
EDITABLE_COLUMNS = {
    "ativo",
    "nome_amigavel",
    "arquivo_xlsx",
    "aba_xlsx",
    "modo_leitura",
    "referencia",
    "orientacao",
    "formato_valores",
    "observacao",
}
MAPPING_COLUMNS = [
    "ativo",
    "id_objeto",
    "slide",
    "tipo_objeto",
    "tipo_visual",
    "titulo_slide",
    "titulo_detectado",
    "confianca_titulo",
    "nome_amigavel",
    "linhas_no_ppt",
    "colunas_no_ppt",
    "qtd_linhas",
    "qtd_colunas",
    "arquivo_xlsx",
    "aba_xlsx",
    "modo_leitura",
    "referencia",
    "orientacao",
    "formato_valores",
    "observacao",
    "status",
]
ALLOWED_READ_MODES = {"auto", "tabela_excel", "dinamico", "exato"}
ALLOWED_ORIENTATIONS = {"auto", "manter", "transpor"}
ALLOWED_VALUE_FORMATS = {"auto", "percentual", "numero", "milhares"}


def prepare_model_assets(
    pptx_bytes: bytes,
    *,
    model_name: str,
    original_filename: str,
) -> tuple[dict[str, Any], bytes]:
    targets = discover_ppt_targets(
        pptx_bytes,
        numeric_only=False,
        include_text_shapes=False,
    )
    targets = [target for target in targets if target.object_type in {"chart", "table"}]
    if not targets:
        raise ValueError("O PPT nao possui graficos ou tabelas PowerPoint mapeaveis.")

    slide_width, slide_height = _presentation_size(pptx_bytes)
    objects = [_target_manifest(target, targets) for target in targets]
    identified = _identified_pptx(pptx_bytes, targets)
    manifest = {
        "schema_version": MODEL_SCHEMA_VERSION,
        "model_name": str(model_name or Path(original_filename).stem).strip(),
        "original_filename": Path(original_filename or "modelo.pptx").name,
        "original_sha256": hashlib.sha256(pptx_bytes).hexdigest(),
        "identified_sha256": hashlib.sha256(identified).hexdigest(),
        "slide_width_in": slide_width,
        "slide_height_in": slide_height,
        "slide_count": max((target.slide_number for target in targets), default=0),
        "object_count": len(objects),
        "chart_count": sum(1 for item in objects if item["tipo_objeto"] == "chart"),
        "table_count": sum(1 for item in objects if item["tipo_objeto"] == "table"),
        "objects": objects,
    }
    return manifest, identified


def build_mapping_workbook(manifest: dict[str, Any], overrides: dict[str, dict] | None = None) -> bytes:
    overrides = overrides or {}
    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "INSTRUCOES"
    objects_sheet = workbook.create_sheet(MAPPING_SHEET)
    structure_sheet = workbook.create_sheet("ESTRUTURA")
    examples_sheet = workbook.create_sheet("EXEMPLOS")
    formats_sheet = workbook.create_sheet("FORMATOS")
    metadata_sheet = workbook.create_sheet("_METADADOS")
    lists_sheet = workbook.create_sheet("_LISTAS")

    _style_instructions(instructions, manifest)
    _write_objects_sheet(objects_sheet, manifest, overrides)
    _write_structure_sheet(structure_sheet, manifest)
    _write_examples_sheet(examples_sheet)
    _write_formats_sheet(formats_sheet)
    _write_metadata_sheet(metadata_sheet, manifest)
    _write_lists_sheet(lists_sheet)
    metadata_sheet.sheet_state = "hidden"
    lists_sheet.sheet_state = "hidden"
    workbook.active = 0

    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def read_mapping_workbook(mapping_bytes: bytes) -> dict[str, dict[str, Any]]:
    workbook = openpyxl.load_workbook(BytesIO(mapping_bytes), data_only=False, read_only=False)
    try:
        if MAPPING_SHEET not in workbook.sheetnames:
            raise ValueError("O XLSX precisa ter a aba OBJETOS.")
        sheet = workbook[MAPPING_SHEET]
        header_row = _find_mapping_header(sheet)
        header_map = {
            _header_key(cell.value): cell.column
            for cell in sheet[header_row]
            if _header_key(cell.value)
        }
        if "id_objeto" not in header_map or "ativo" not in header_map:
            raise ValueError("A aba OBJETOS precisa manter as colunas id_objeto e ativo.")

        entries: dict[str, dict[str, Any]] = {}
        for row in range(header_row + 1, sheet.max_row + 1):
            target_id = _cell_text(sheet.cell(row, header_map["id_objeto"]).value)
            if not target_id:
                continue
            if target_id in entries:
                raise ValueError(f"O id_objeto '{target_id}' aparece mais de uma vez.")
            entry = {
                column: sheet.cell(row, header_map[column]).value
                for column in MAPPING_COLUMNS
                if column in header_map
            }
            entry["id_objeto"] = target_id
            entry["ativo"] = _active_value(entry.get("ativo"))
            for key in MAPPING_COLUMNS:
                if key not in {"ativo", "slide", "qtd_linhas", "qtd_colunas"}:
                    entry[key] = _cell_text(entry.get(key))
            filename, embedded_sheet = _split_source_reference(entry.get("arquivo_xlsx", ""))
            entry["arquivo_xlsx"] = filename
            if embedded_sheet and not entry.get("aba_xlsx"):
                entry["aba_xlsx"] = embedded_sheet
            entry["modo_leitura"] = _normalized_choice(
                entry.get("modo_leitura"),
                ALLOWED_READ_MODES,
                "auto",
                f"modo_leitura invalido em {target_id}",
            )
            entry["orientacao"] = _normalized_choice(
                entry.get("orientacao"),
                ALLOWED_ORIENTATIONS,
                "auto",
                f"orientacao invalida em {target_id}",
            )
            entry["formato_valores"] = _normalized_choice(
                entry.get("formato_valores"),
                ALLOWED_VALUE_FORMATS,
                "auto",
                f"formato_valores invalido em {target_id}",
            )
            entries[target_id] = entry
        return entries
    finally:
        workbook.close()


def validate_mapping_package(
    manifest: dict[str, Any],
    mapping_bytes: bytes,
    source_files: dict[str, bytes],
) -> tuple[dict[str, Any], dict[str, dict[str, Any]]]:
    rows = read_mapping_workbook(mapping_bytes)
    targets = {
        target.target_id: target
        for target in discover_ppt_targets(
            _manifest_identified_bytes(manifest),
            numeric_only=False,
            include_text_shapes=False,
        )
        if target.object_type in {"chart", "table"}
    } if manifest.get("_identified_pptx_bytes") else {}
    if not targets:
        targets = {
            item["id_objeto"]: _target_from_manifest(item)
            for item in manifest.get("objects") or []
        }
    expected_ids = set(targets)
    errors: list[dict[str, Any]] = []
    reviews: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []
    resolved: dict[str, dict[str, Any]] = {}

    normalized_sources: dict[str, tuple[str, bytes]] = {}
    duplicate_names: set[str] = set()
    for raw_name, data in source_files.items():
        name = Path(str(raw_name or "")).name
        key = name.casefold()
        if key in normalized_sources:
            duplicate_names.add(name)
        normalized_sources[key] = (name, data)
    for name in sorted(duplicate_names):
        errors.append(_issue("ERROR", "", f"Arquivo repetido: {name}. Renomeie para nomes unicos."))

    active_rows = {target_id: row for target_id, row in rows.items() if row.get("ativo") == 1}
    for target_id in sorted(set(rows) - expected_ids):
        warnings.append(_issue("WARNING", target_id, "ID nao existe neste PPT e sera ignorado."))
    missing_rows = sorted(expected_ids - set(rows))
    if missing_rows:
        warnings.append(
            _issue(
                "WARNING",
                "",
                f"{len(missing_rows)} objeto(s) do PPT nao aparecem no XLSX e ficarao inativos.",
            )
        )

    used_files: set[str] = set()
    for target_id, row in active_rows.items():
        target = targets.get(target_id)
        if target is None:
            continue
        filename = Path(row.get("arquivo_xlsx") or "").name
        if not filename:
            errors.append(_issue("ERROR", target_id, "Preencha arquivo_xlsx para o objeto ativo."))
            continue
        source_pair = normalized_sources.get(filename.casefold())
        if source_pair is None:
            errors.append(_issue("ERROR", target_id, f"Arquivo ausente: {filename}."))
            continue
        uploaded_name, source_bytes = source_pair
        used_files.add(uploaded_name.casefold())
        try:
            sheet_names = workbook_sheet_names(source_bytes)
        except Exception as exc:
            errors.append(_issue("ERROR", target_id, f"Nao foi possivel abrir {uploaded_name}: {exc}"))
            continue
        sheet = str(row.get("aba_xlsx") or "").strip()
        if not sheet:
            if len(sheet_names) == 1:
                sheet = sheet_names[0]
            else:
                errors.append(
                    _issue(
                        "ERROR",
                        target_id,
                        f"{uploaded_name} tem {len(sheet_names)} abas. Preencha aba_xlsx. "
                        f"Disponiveis: {', '.join(sheet_names[:12])}.",
                    )
                )
                continue
        if sheet not in sheet_names:
            errors.append(
                _issue(
                    "ERROR",
                    target_id,
                    f"Aba '{sheet}' nao existe em {uploaded_name}. "
                    f"Disponiveis: {', '.join(sheet_names[:12])}.",
                )
            )
            continue
        try:
            cell_range, range_mode, range_note = _resolve_read_contract(
                source_bytes,
                sheet=sheet,
                mode=row.get("modo_leitura") or "auto",
                reference=row.get("referencia") or "",
            )
            source_name = uploaded_name if len(sheet_names) == 1 else f"{uploaded_name}#{sheet}"
            source = parse_xlsx_table(
                source_bytes,
                file_name=source_name,
                formula_mode="auto",
                cell_range=cell_range,
                range_mode=range_mode,
                sheet=sheet,
            )
            allow_axis_growth = _allows_axis_growth(
                row.get("modo_leitura") or "auto",
                cell_range,
                range_mode,
            )
            if allow_axis_growth:
                source = replace(
                    source,
                    metadata={**source.metadata, "_allow_axis_growth": "1"},
                )
            if not source.values or not any(
                value is not None and str(value).strip()
                for source_row in source.values
                for value in source_row
            ):
                raise ValueError("a tabela selecionada nao possui dados.")
            plan = normalize_to_target(
                target,
                source,
                confidence=1.0,
                match_reason="Mapeamento preparado pelo cliente.",
                axis_mode=row.get("orientacao") or "auto",
            )
            if source.table_blocks > 1 and not cell_range:
                reviews.append(
                    _issue(
                        "REVIEW",
                        target_id,
                        "A aba parece ter mais de um bloco. Defina referencia ou use uma Tabela do Excel.",
                    )
                )
            for message in plan.warnings:
                reviews.append(_issue("REVIEW", target_id, message))
            learning = mapping_entry_learning_fields(target, source)
            resolved[target_id] = {
                "target_id": target_id,
                "object_type": target.object_type,
                "slide": target.slide_number,
                "shape_name": target.shape_name,
                "shape_id": target.shape_id,
                "target_aliases": sorted(target_aliases(target)),
                "target_fingerprint": target_fingerprint(target),
                "datasource": source.file_name,
                "datasource_basename": Path(source.file_name).name,
                "cell_range": cell_range,
                "range_mode": range_mode,
                "read_mode": row.get("modo_leitura") or "auto",
                "allow_axis_growth": allow_axis_growth,
                "orientation": row.get("orientacao") or "auto",
                "value_format": row.get("formato_valores") or "auto",
                "friendly_name": row.get("nome_amigavel") or "",
                "notes": row.get("observacao") or "",
                "action": plan.action,
                "confidence": 1.0,
                "reason": f"Preparador de Modelo validou {source.file_name}. {range_note}".strip(),
                **learning,
            }
        except Exception as exc:
            errors.append(_issue("ERROR", target_id, f"Fonte invalida: {exc}"))

    for key, (name, _data) in normalized_sources.items():
        if key not in used_files:
            warnings.append(_issue("WARNING", "", f"Arquivo enviado mas nao usado: {name}."))
    if not active_rows:
        errors.append(_issue("ERROR", "", "Nenhum objeto esta ativo. Marque ativo=1 no que deve atualizar."))

    report = {
        "schema_version": MODEL_SCHEMA_VERSION,
        "object_count": len(expected_ids),
        "mapping_row_count": len(rows),
        "active_count": len(active_rows),
        "resolved_count": len(resolved),
        "source_count": len(normalized_sources),
        "error_count": len(errors),
        "review_count": len(reviews),
        "warning_count": len(warnings),
        "errors": errors,
        "reviews": reviews,
        "warnings": warnings,
        "ok": not errors,
    }
    return report, resolved


def mapping_entries_from_rows(
    manifest: dict[str, Any],
    rows: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    output = []
    for item in manifest.get("objects") or []:
        target_id = str(item.get("id_objeto") or "")
        merged = dict(item)
        merged.update(rows.get(target_id) or {})
        merged["id_objeto"] = target_id
        output.append(merged)
    return output


def datasource_zip(source_files: dict[str, bytes]) -> bytes:
    stream = BytesIO()
    with ZipFile(stream, "w", ZIP_DEFLATED) as archive:
        for name, data in sorted(source_files.items(), key=lambda item: item[0].casefold()):
            archive.writestr(Path(name).name, data)
    return stream.getvalue()


def _target_manifest(target: PptTarget, all_targets: Iterable[PptTarget]) -> dict[str, Any]:
    if target.object_type == "chart":
        if target.expected_orientation == "series_rows_categories_columns":
            rows = list(target.expected_series)
            columns = list(target.expected_categories)
        else:
            rows = list(target.expected_categories)
            columns = list(target.expected_series)
    else:
        cells = [list(row) for row in target.table_cells]
        columns = list(cells[0]) if cells else []
        rows = [row[0] for row in cells[1:] if row] if len(cells) > 1 else []
    detected_title = _detected_title(target)
    slide_title = next(
        (
            _detected_title(candidate)
            for candidate in all_targets
            if candidate.slide_number == target.slide_number and _detected_title(candidate)
        ),
        "",
    )
    return {
        "ativo": 0,
        "id_objeto": target.target_id,
        "rotulo_visual": visual_label(target.target_id),
        "slide": target.slide_number,
        "tipo_objeto": target.object_type,
        "tipo_visual": target.chart_kind or ("tabela" if target.object_type == "table" else ""),
        "titulo_slide": slide_title,
        "titulo_detectado": detected_title,
        "confianca_titulo": _title_confidence(target, detected_title),
        "nome_amigavel": detected_title,
        "linhas_no_ppt": rows,
        "colunas_no_ppt": columns,
        "qtd_linhas": len(rows),
        "qtd_colunas": len(columns),
        "arquivo_xlsx": "",
        "aba_xlsx": "",
        "modo_leitura": "auto",
        "referencia": "",
        "orientacao": "auto",
        "formato_valores": "auto",
        "observacao": "",
        "status": "INATIVO",
        "shape_name_original": target.shape_name,
        "shape_id": target.shape_id,
        "target_fingerprint": target_fingerprint(target),
        "left_in": round(target.left_in, 4),
        "top_in": round(target.top_in, 4),
        "width_in": round(target.width_in, 4),
        "height_in": round(target.height_in, 4),
        "expected_orientation": target.expected_orientation,
        "expected_categories": list(target.expected_categories),
        "expected_series": list(target.expected_series),
        "expected_values": [list(row) for row in target.expected_values],
        "table_cells": [list(row) for row in target.table_cells],
        "value_format": target.value_format,
        "series_value_formats": list(target.series_value_formats),
        "chart_series_colors": list(target.chart_series_colors),
        "chart_xml": target.chart_xml,
        "workbook_embedded": target.workbook_embedded,
        "sheet_name": target.sheet_name,
    }


def _target_from_manifest(item: dict[str, Any]) -> PptTarget:
    return PptTarget(
        slide_index=max(int(item.get("slide") or 1) - 1, 0),
        slide_number=int(item.get("slide") or 1),
        slide_path=f"ppt/slides/slide{int(item.get('slide') or 1)}.xml",
        shape_name=str(item.get("id_objeto") or ""),
        shape_id=str(item.get("shape_id") or ""),
        object_type=str(item.get("tipo_objeto") or ""),
        left_in=float(item.get("left_in") or 0),
        top_in=float(item.get("top_in") or 0),
        width_in=float(item.get("width_in") or 0),
        height_in=float(item.get("height_in") or 0),
        title=str(item.get("titulo_detectado") or ""),
        chart_xml=str(item.get("chart_xml") or ""),
        workbook_embedded=str(item.get("workbook_embedded") or ""),
        sheet_name=str(item.get("sheet_name") or ""),
        expected_orientation=str(item.get("expected_orientation") or ""),
        expected_categories=list(item.get("expected_categories") or []),
        expected_series=list(item.get("expected_series") or []),
        expected_values=[list(row) for row in item.get("expected_values") or []],
        value_format=str(item.get("value_format") or ""),
        series_value_formats=list(item.get("series_value_formats") or []),
        chart_kind=str(item.get("tipo_visual") or ""),
        chart_series_colors=list(item.get("chart_series_colors") or []),
        table_cells=[list(row) for row in item.get("table_cells") or []],
        target_key=str(item.get("id_objeto") or ""),
    )


def _identified_pptx(pptx_bytes: bytes, targets: list[PptTarget]) -> bytes:
    by_slide: dict[str, list[PptTarget]] = {}
    for target in targets:
        by_slide.setdefault(target.slide_path, []).append(target)
    output = BytesIO()
    with ZipFile(BytesIO(pptx_bytes)) as source, ZipFile(output, "w", ZIP_DEFLATED) as destination:
        for info in source.infolist():
            data = source.read(info.filename)
            if info.filename in by_slide:
                data = rename_targets_in_slide_xml(data, by_slide[info.filename])
            destination.writestr(info, data)
    return output.getvalue()


def _presentation_size(pptx_bytes: bytes) -> tuple[float, float]:
    try:
        with ZipFile(BytesIO(pptx_bytes)) as archive:
            root = ET.fromstring(archive.read("ppt/presentation.xml"))
        namespace = "http://schemas.openxmlformats.org/presentationml/2006/main"
        size = root.find(f"{{{namespace}}}sldSz")
        if size is not None:
            return round(int(size.attrib["cx"]) / 914400, 4), round(int(size.attrib["cy"]) / 914400, 4)
    except Exception:
        pass
    return 13.3333, 7.5


def _detected_title(target: PptTarget) -> str:
    if target.title:
        return target.title
    segments = [
        re.sub(r"\s+", " ", segment).strip(" .:-")
        for segment in re.split(r"\s+\|\s+", target.nearby_text or "")
    ]
    candidates = [
        segment
        for segment in segments
        if 4 <= len(segment) <= 100
        and "<" not in segment
        and ">" not in segment
        and "nome_" not in segment.casefold()
    ]
    if not candidates:
        return ""
    def score(value: str) -> tuple[int, int, int]:
        letters = [char for char in value if char.isalpha()]
        uppercase = bool(letters) and sum(char.isupper() for char in letters) / len(letters) >= 0.85
        title_words = len(value.split())
        return (
            4 if uppercase else 0,
            2 if 2 <= title_words <= 12 else 0,
            -abs(len(value) - 32),
        )
    return max(candidates, key=score)


def _title_confidence(target: PptTarget, detected_title: str) -> str:
    if target.title and target.nearby_text and target.title in target.nearby_text:
        return "alta"
    if target.title:
        return "media"
    if detected_title:
        return "baixa"
    return "ausente"


def _style_instructions(sheet, manifest: dict[str, Any]) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:H1")
    sheet["A1"] = "PREPARADOR DE MODELO — INSTRUCOES"
    sheet["A1"].font = Font(name="Aptos Display", size=18, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="166047")
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 34
    lines = [
        ("Modelo", manifest.get("model_name", "")),
        ("Arquivo", manifest.get("original_filename", "")),
        ("Objetos encontrados", manifest.get("object_count", 0)),
        ("1", "Abra a aba OBJETOS. Cada linha representa um grafico ou tabela do PowerPoint."),
        ("2", "Marque ativo=1 somente no que deve ser atualizado."),
        ("3", "Preencha arquivo_xlsx e aba_xlsx. Tambem aceitamos arquivo.xlsx#Aba."),
        ("4", "Prefira modo_leitura=auto. Use tabela_excel, dinamico ou exato quando precisar limitar auxiliares."),
        ("5", "Envie este XLSX junto com todos os arquivos citados. A validacao abre cada fonte antes do preview."),
    ]
    for row, (label, text) in enumerate(lines, start=3):
        sheet.cell(row, 1, label).font = Font(bold=True, color="166047")
        sheet.cell(row, 2, text)
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        sheet.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
        sheet.row_dimensions[row].height = 30 if row >= 6 else 22
    for column, width in {"A": 24, "B": 24, "C": 18, "D": 18, "E": 18, "F": 18, "G": 18, "H": 18}.items():
        sheet.column_dimensions[column].width = width


def _write_objects_sheet(sheet, manifest: dict[str, Any], overrides: dict[str, dict]) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(MAPPING_COLUMNS))
    sheet.cell(1, 1, "OBJETOS DO POWERPOINT")
    sheet.cell(1, 1).font = Font(name="Aptos Display", size=16, bold=True, color="FFFFFF")
    sheet.cell(1, 1).fill = PatternFill("solid", fgColor="166047")
    sheet.cell(1, 1).alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 30
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(MAPPING_COLUMNS))
    sheet.cell(2, 1, "Cinza = gerado pelo sistema. Verde = campo para preencher. IDs nao devem ser alterados.")
    sheet.cell(2, 1).font = Font(italic=True, color="58636E")
    for col, name in enumerate(MAPPING_COLUMNS, start=1):
        cell = sheet.cell(MAPPING_HEADER_ROW, col, name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="24313A")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[MAPPING_HEADER_ROW].height = 42

    objects = manifest.get("objects") or []
    for row_index, item in enumerate(objects, start=MAPPING_HEADER_ROW + 1):
        merged = dict(item)
        merged.update(overrides.get(str(item.get("id_objeto") or "")) or {})
        for col_index, column in enumerate(MAPPING_COLUMNS, start=1):
            value: Any = merged.get(column, "")
            if column in {"linhas_no_ppt", "colunas_no_ppt"} and isinstance(value, list):
                value = "\n".join(str(part) for part in value)
            cell = sheet.cell(row_index, col_index, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if column in EDITABLE_COLUMNS:
                cell.fill = PatternFill("solid", fgColor="EAF7F1")
                cell.protection = Protection(locked=False)
            else:
                cell.fill = PatternFill("solid", fgColor="F1F3F4")
                cell.protection = Protection(locked=True)
        status_col = MAPPING_COLUMNS.index("status") + 1
        active_col = _excel_column(MAPPING_COLUMNS.index("ativo") + 1)
        file_col = _excel_column(MAPPING_COLUMNS.index("arquivo_xlsx") + 1)
        sheet.cell(
            row_index,
            status_col,
            f'=IF({active_col}{row_index}=0,"IGNORADO",IF({file_col}{row_index}="","FALTA ARQUIVO","PRONTO"))',
        )
        sheet.row_dimensions[row_index].height = 44

    last_row = max(MAPPING_HEADER_ROW + len(objects), MAPPING_HEADER_ROW + 1)
    if objects:
        table = Table(displayName="ObjetosDoPowerPoint", ref=f"A{MAPPING_HEADER_ROW}:{_excel_column(len(MAPPING_COLUMNS))}{last_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium4",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
    sheet.auto_filter.ref = f"A{MAPPING_HEADER_ROW}:{_excel_column(len(MAPPING_COLUMNS))}{last_row}"
    sheet.freeze_panes = "A5"
    widths = {
        "ativo": 9, "id_objeto": 22, "slide": 8, "tipo_objeto": 13, "tipo_visual": 18,
        "titulo_slide": 28, "titulo_detectado": 34, "confianca_titulo": 16,
        "nome_amigavel": 30, "linhas_no_ppt": 34, "colunas_no_ppt": 34,
        "qtd_linhas": 11, "qtd_colunas": 12, "arquivo_xlsx": 30, "aba_xlsx": 24,
        "modo_leitura": 18, "referencia": 22, "orientacao": 14,
        "formato_valores": 18, "observacao": 36, "status": 18,
    }
    for index, name in enumerate(MAPPING_COLUMNS, start=1):
        sheet.column_dimensions[_excel_column(index)].width = widths.get(name, 18)

    _list_validation(sheet, "ativo", ["0", "1"], last_row)
    _list_validation(sheet, "modo_leitura", sorted(ALLOWED_READ_MODES), last_row)
    _list_validation(sheet, "orientacao", sorted(ALLOWED_ORIENTATIONS), last_row)
    _list_validation(sheet, "formato_valores", sorted(ALLOWED_VALUE_FORMATS), last_row)
    status_letter = _excel_column(MAPPING_COLUMNS.index("status") + 1)
    sheet.conditional_formatting.add(
        f"{status_letter}{MAPPING_HEADER_ROW + 1}:{status_letter}{last_row}",
        FormulaRule(
            formula=[f'{status_letter}{MAPPING_HEADER_ROW + 1}="FALTA ARQUIVO"'],
            fill=PatternFill("solid", fgColor="FDE8E7"),
        ),
    )
    sheet.protection.sheet = True
    sheet.protection.autoFilter = True
    sheet.protection.sort = True
    sheet.protection.selectLockedCells = False
    sheet.protection.selectUnlockedCells = True


def _write_structure_sheet(sheet, manifest: dict[str, Any]) -> None:
    headers = [
        "id_objeto", "slide", "tipo", "titulo_detectado", "linhas_do_grafico_antes",
        "colunas_do_grafico_antes", "shape_name_original", "shape_id",
    ]
    sheet.append(headers)
    for item in manifest.get("objects") or []:
        sheet.append([
            item.get("id_objeto"), item.get("slide"), item.get("tipo_objeto"),
            item.get("titulo_detectado"),
            "\n".join(str(value) for value in item.get("linhas_no_ppt") or []),
            "\n".join(str(value) for value in item.get("colunas_no_ppt") or []),
            item.get("shape_name_original"), item.get("shape_id"),
        ])
    _style_simple_sheet(sheet, len(headers))


def _write_examples_sheet(sheet) -> None:
    sheet.append(["situacao", "ativo", "arquivo_xlsx", "aba_xlsx", "modo_leitura", "referencia", "quando_usar"])
    sheet.append(["Recomendado", 1, "nps.xlsx", "Mensal", "auto", "", "Uma Tabela do Excel ou um bloco limpo na aba."])
    sheet.append(["Tabela nomeada", 1, "vendas.xlsx", "Base", "tabela_excel", "tb_vendas", "Melhor opcao para crescer todo mes e ignorar auxiliares."])
    sheet.append(["Range dinamico", 1, "pesquisa.xlsx", "Resultado", "dinamico", "A2:F8", "Ancora fixa; novas linhas/meses contiguos entram."])
    sheet.append(["Range exato", 1, "indicadores.xlsx", "Painel", "exato", "B3:G9", "Nunca sair do intervalo informado."])
    sheet.append(["Ignorar objeto", 0, "", "", "auto", "", "Objeto permanece intacto no PowerPoint."])
    _style_simple_sheet(sheet, 7)


def _write_formats_sheet(sheet) -> None:
    sheet.append(["campo", "valor", "efeito"])
    rows = [
        ("modo_leitura", "auto", "Usa Tabela do Excel unica; senao le o bloco da aba."),
        ("modo_leitura", "tabela_excel", "Referencia deve ser o nome da Tabela do Excel."),
        ("modo_leitura", "dinamico", "Expande range-semente para novos meses/linhas, sem puxar auxiliares laterais."),
        ("modo_leitura", "exato", "Usa somente o range informado."),
        ("orientacao", "auto", "Sistema decide manter ou transpor pela estrutura."),
        ("orientacao", "manter", "Registra preferencia por manter orientacao."),
        ("orientacao", "transpor", "Registra preferencia por transpor orientacao."),
        ("formato_valores", "auto", "Preserva percentual quando fonte e grafico indicam percentual."),
        ("formato_valores", "percentual", "Forca tratamento de percentual."),
        ("formato_valores", "numero", "Forca numero comum."),
        ("formato_valores", "milhares", "Aplica exibicao em milhares quando suportado."),
    ]
    for row in rows:
        sheet.append(row)
    _style_simple_sheet(sheet, 3)


def _write_metadata_sheet(sheet, manifest: dict[str, Any]) -> None:
    sheet.append(["chave", "valor"])
    sheet.append(["schema_version", MODEL_SCHEMA_VERSION])
    sheet.append(["original_filename", manifest.get("original_filename", "")])
    sheet.append(["original_sha256", manifest.get("original_sha256", "")])
    payload = json.dumps(manifest, ensure_ascii=False, separators=(",", ":"))
    for index in range(0, len(payload), 30000):
        sheet.append([f"manifest_json_{index // 30000:04d}", payload[index:index + 30000]])


def _write_lists_sheet(sheet) -> None:
    lists = {
        "ativo": ["0", "1"],
        "modo_leitura": sorted(ALLOWED_READ_MODES),
        "orientacao": sorted(ALLOWED_ORIENTATIONS),
        "formato_valores": sorted(ALLOWED_VALUE_FORMATS),
    }
    for column, (name, values) in enumerate(lists.items(), start=1):
        sheet.cell(1, column, name)
        for row, value in enumerate(values, start=2):
            sheet.cell(row, column, value)


def _style_simple_sheet(sheet, column_count: int) -> None:
    sheet.sheet_view.showGridLines = False
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="24313A")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.freeze_panes = "A2"
    for column in range(1, column_count + 1):
        sheet.column_dimensions[_excel_column(column)].width = 28
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def _list_validation(sheet, column_name: str, values: list[str], last_row: int) -> None:
    column = _excel_column(MAPPING_COLUMNS.index(column_name) + 1)
    validation = DataValidation(type="list", formula1=f'"{",".join(values)}"', allow_blank=False)
    validation.error = f"Escolha: {', '.join(values)}"
    validation.errorTitle = "Valor invalido"
    sheet.add_data_validation(validation)
    validation.add(f"{column}{MAPPING_HEADER_ROW + 1}:{column}{max(last_row, MAPPING_HEADER_ROW + 500)}")


def _find_mapping_header(sheet) -> int:
    for row in range(1, min(sheet.max_row, 20) + 1):
        values = {_header_key(cell.value) for cell in sheet[row]}
        if "id_objeto" in values:
            return row
    raise ValueError("Nao encontrei o cabecalho da aba OBJETOS.")


def _resolve_read_contract(source_bytes: bytes, *, sheet: str, mode: str, reference: str) -> tuple[str, str, str]:
    normalized_mode = _normalized_choice(mode, ALLOWED_READ_MODES, "auto", "modo_leitura invalido")
    reference = str(reference or "").strip().replace("$", "")
    tables = _excel_tables(source_bytes, sheet)
    if normalized_mode == "tabela_excel":
        table_name = reference or (next(iter(tables)) if len(tables) == 1 else "")
        if not table_name:
            raise ValueError("informe o nome da Tabela do Excel em referencia.")
        if table_name.casefold() not in {name.casefold() for name in tables}:
            raise ValueError(
                f"Tabela do Excel '{table_name}' nao encontrada. Disponiveis: {', '.join(tables) or 'nenhuma'}."
            )
        actual_name = next(name for name in tables if name.casefold() == table_name.casefold())
        return actual_name, "exact", f"Tabela do Excel {actual_name}."
    if normalized_mode == "dinamico":
        if not reference:
            raise ValueError("modo dinamico exige um range-semente, como A2:F8.")
        _validate_reference(reference, allow_table=False)
        return reference, "dynamic", f"Range dinamico {reference}."
    if normalized_mode == "exato":
        if not reference:
            raise ValueError("modo exato exige um range, como A2:F8.")
        _validate_reference(reference, allow_table=False)
        return reference, "exact", f"Range exato {reference}."
    if reference:
        if reference.casefold() in {name.casefold() for name in tables}:
            actual_name = next(name for name in tables if name.casefold() == reference.casefold())
            return actual_name, "exact", f"Auto escolheu Tabela do Excel {actual_name}."
        _validate_reference(reference, allow_table=False)
        return reference, "dynamic", f"Auto expandira o range-semente {reference}."
    if len(tables) == 1:
        table_name = next(iter(tables))
        return table_name, "exact", f"Auto escolheu Tabela do Excel {table_name}."
    return "", "exact", "Auto leu o bloco principal da aba."


def _allows_axis_growth(read_mode: str, cell_range: str, range_mode: str) -> bool:
    mode = str(read_mode or "auto").strip().lower()
    reference = str(cell_range or "").strip()
    if mode == "dinamico" or str(range_mode or "").strip().lower() == "dynamic":
        return bool(reference)
    if mode == "tabela_excel":
        return bool(reference)
    if mode == "auto" and reference and not re.fullmatch(
        r"(?:'[^']+'|[^!]+)!?\$?[A-Za-z]{1,3}\$?\d+(?::\$?[A-Za-z]{1,3}\$?\d+)?",
        reference,
    ):
        return True
    return False


def _excel_tables(source_bytes: bytes, sheet_name: str) -> dict[str, str]:
    workbook = openpyxl.load_workbook(BytesIO(source_bytes), data_only=False, read_only=False)
    try:
        if sheet_name not in workbook.sheetnames:
            return {}
        sheet = workbook[sheet_name]
        return {
            str(getattr(table, "name", "") or ""): str(getattr(table, "ref", "") or "")
            for table in sheet.tables.values()
            if str(getattr(table, "name", "") or "")
        }
    finally:
        workbook.close()


def _validate_reference(reference: str, *, allow_table: bool = True) -> None:
    text = reference.split("!", 1)[-1].strip().strip("'")
    if re.fullmatch(r"[A-Za-z]{1,4}\d{1,7}(:[A-Za-z]{1,4}\d{1,7})?", text):
        return
    if allow_table and re.fullmatch(r"[A-Za-z_\\][A-Za-z0-9_.\\]*", text):
        return
    raise ValueError("referencia invalida. Use A2:F8 ou o nome de uma Tabela do Excel.")


def _split_source_reference(value: str) -> tuple[str, str]:
    text = str(value or "").strip().replace("\\", "/")
    if "#" not in text:
        return Path(text).name, ""
    filename, sheet = text.split("#", 1)
    return Path(filename).name, sheet.strip()


def _normalized_choice(value: Any, allowed: set[str], default: str, message: str) -> str:
    normalized = _norm(value).lower().replace(" ", "_")
    if not normalized:
        return default
    aliases = {
        "automatico": "auto",
        "automático": "auto",
        "tabela": "tabela_excel",
        "dinâmico": "dinamico",
        "exact": "exato",
        "keep": "manter",
        "transpose": "transpor",
        "percent": "percentual",
    }
    normalized = aliases.get(normalized, normalized)
    if normalized not in allowed:
        raise ValueError(f"{message}: {value}.")
    return normalized


def _active_value(value: Any) -> int:
    if isinstance(value, bool):
        return 1 if value else 0
    text = _norm(value).lower()
    if text in {"1", "sim", "s", "yes", "true", "ativo"}:
        return 1
    if text in {"", "0", "nao", "não", "n", "no", "false", "inativo"}:
        return 0
    raise ValueError(f"ativo deve ser 1 ou 0, recebido: {value}.")


def _issue(level: str, target_id: str, message: str) -> dict[str, str]:
    return {"level": level, "target_id": target_id, "message": message}


def _header_key(value: Any) -> str:
    text = _norm(value).lower()
    text = re.sub(r"[^a-z0-9]+", "_", text).strip("_")
    aliases = {
        "id": "id_objeto",
        "objeto_id": "id_objeto",
        "arquivo": "arquivo_xlsx",
        "nome_do_arquivo": "arquivo_xlsx",
        "aba": "aba_xlsx",
        "range": "referencia",
        "intervalo": "referencia",
        "ativo_1_0": "ativo",
    }
    return aliases.get(text, text)


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _norm(value: Any) -> str:
    text = "" if value is None else str(value).strip()
    text = unicodedata.normalize("NFKD", text)
    return "".join(char for char in text if not unicodedata.combining(char))


def _excel_column(number: int) -> str:
    from openpyxl.utils import get_column_letter

    return get_column_letter(number)


def _manifest_identified_bytes(manifest: dict[str, Any]) -> bytes:
    value = manifest.get("_identified_pptx_bytes")
    return bytes(value) if isinstance(value, (bytes, bytearray)) else b""
