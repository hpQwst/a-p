from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path, PurePosixPath
from typing import Any, Iterable
from urllib.parse import unquote, urlparse
from zipfile import ZipFile
import posixpath
import re
import xml.etree.ElementTree as ET

import openpyxl

from .archive_safety import validate_pptx_bytes
from .engine import _workbook_matrix
from .ppt_chart_writer import resolved_series_number_formats
from .ppt_discovery import NS, PptTarget, read_bytes
from .table_normalizer import TransformPlan


REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
CHART_NS = "http://schemas.openxmlformats.org/drawingml/2006/chart"
REGRESSION_NS = {
    **NS,
    "c": CHART_NS,
    "r": REL_NS,
}


class PptxRegressionError(AssertionError):
    pass


@dataclass(frozen=True)
class PptxRegressionReport:
    slide_count: int
    package_parts: int
    xml_parts: int
    relationships_checked: int
    geometry_objects_checked: int
    charts_checked: int
    tables_checked: int


@dataclass(frozen=True)
class RenderRegressionReport:
    slides_checked: int
    pixels_checked_outside_targets: int
    pixels_changed_outside_targets: int
    worst_outside_change_ratio: float


def validate_generated_pptx(
    original_pptx: bytes | bytearray | str | Path,
    generated_pptx: bytes | bytearray | str | Path,
    plans: Iterable[TransformPlan],
    targets: Iterable[PptTarget] | None = None,
) -> PptxRegressionReport:
    """Valida um PPT gerado contra o template e os planos aplicados.

    A checagem e deliberadamente mais forte do que "o ZIP abre": garante a
    integridade OPC, preservacao byte-a-byte das partes que nao deveriam mudar,
    geometria intacta, matrizes completas nos workbooks dos graficos e ausencia
    de truncamento nas tabelas do slide.
    """

    original_bytes = read_bytes(original_pptx)
    generated_bytes = read_bytes(generated_pptx)
    plan_list = list(plans)
    target_list = list(targets) if targets is not None else [plan.target for plan in plan_list]
    validate_pptx_bytes(generated_bytes)

    with ZipFile(BytesIO(original_bytes)) as original, ZipFile(BytesIO(generated_bytes)) as generated:
        bad_member = generated.testzip()
        if bad_member:
            raise PptxRegressionError(f"PPTX gerado tem CRC invalido em {bad_member}.")

        original_names = set(original.namelist())
        generated_names = set(generated.namelist())
        if original_names != generated_names:
            missing = sorted(original_names - generated_names)
            added = sorted(generated_names - original_names)
            raise PptxRegressionError(
                f"Partes OPC mudaram. Ausentes={missing[:8]} adicionadas={added[:8]}."
            )

        allowed_changes = _allowed_changed_parts(plan_list, target_list)
        unexpected_changes = [
            name
            for name in sorted(original_names - allowed_changes)
            if original.read(name) != generated.read(name)
        ]
        if unexpected_changes:
            raise PptxRegressionError(
                "Partes fora dos targets foram alteradas: "
                + ", ".join(unexpected_changes[:12])
            )

        xml_parts = _validate_all_xml(generated)
        relationships_checked = _validate_relationship_targets(generated)
        geometry_objects_checked = _validate_slide_geometry(original, generated)
        charts_checked, tables_checked = _validate_plan_outputs(generated, plan_list)
        slide_count = len(
            [
                name
                for name in generated_names
                if re.fullmatch(r"ppt/slides/slide\d+\.xml", name)
            ]
        )

    return PptxRegressionReport(
        slide_count=slide_count,
        package_parts=len(generated_names),
        xml_parts=xml_parts,
        relationships_checked=relationships_checked,
        geometry_objects_checked=geometry_objects_checked,
        charts_checked=charts_checked,
        tables_checked=tables_checked,
    )


def validate_rendered_slides(
    original_pptx: bytes | bytearray | str | Path,
    original_render_dir: str | Path,
    generated_render_dir: str | Path,
    plans: Iterable[TransformPlan],
    *,
    pixel_threshold: int = 8,
    outside_change_ratio_limit: float = 0.00005,
    target_padding_in: float = 0.12,
) -> RenderRegressionReport:
    """Compara renders e rejeita mudancas visuais fora dos objetos atualizados."""

    try:
        from PIL import Image, ImageChops, ImageDraw
    except ModuleNotFoundError as exc:
        raise PptxRegressionError(
            "A regressao visual requer Pillow. Instale-o apenas no ambiente "
            "local de validacao com: python -m pip install Pillow."
        ) from exc

    plan_list = list(plans)
    original_images = _rendered_slide_paths(Path(original_render_dir))
    generated_images = _rendered_slide_paths(Path(generated_render_dir))
    if original_images.keys() != generated_images.keys():
        raise PptxRegressionError(
            "Renders original/gerado tem conjuntos de slides diferentes: "
            f"{sorted(original_images)} != {sorted(generated_images)}."
        )
    if not original_images:
        raise PptxRegressionError("Nenhum slide PNG foi renderizado.")

    slide_width_in, slide_height_in = _slide_size_inches(read_bytes(original_pptx))
    plans_by_slide: dict[int, list[TransformPlan]] = {}
    for plan in plan_list:
        plans_by_slide.setdefault(plan.target.slide_number, []).append(plan)

    checked_pixels = 0
    changed_pixels = 0
    worst_ratio = 0.0
    for slide_number in sorted(original_images):
        with Image.open(original_images[slide_number]) as before_source:
            before = before_source.convert("RGB")
        with Image.open(generated_images[slide_number]) as after_source:
            after = after_source.convert("RGB")
        if before.size != after.size:
            raise PptxRegressionError(
                f"Render do slide {slide_number} mudou de tamanho: "
                f"{before.size} != {after.size}."
            )

        outside_mask = Image.new("1", before.size, 1)
        draw = ImageDraw.Draw(outside_mask)
        for plan in plans_by_slide.get(slide_number, []):
            target = plan.target
            left = max(0.0, target.left_in - target_padding_in)
            top = max(0.0, target.top_in - target_padding_in)
            right = min(slide_width_in, target.left_in + target.width_in + target_padding_in)
            bottom = min(slide_height_in, target.top_in + target.height_in + target_padding_in)
            draw.rectangle(
                (
                    round(left / slide_width_in * before.width),
                    round(top / slide_height_in * before.height),
                    round(right / slide_width_in * before.width),
                    round(bottom / slide_height_in * before.height),
                ),
                fill=0,
            )

        difference = ImageChops.difference(before, after).convert("L")
        changed = difference.point(lambda value: 255 if value > pixel_threshold else 0, mode="1")
        changed_outside = ImageChops.logical_and(changed, outside_mask)
        outside_count = outside_mask.histogram()[1]
        changed_count = changed_outside.histogram()[1]
        ratio = (changed_count / outside_count) if outside_count else 0.0
        if ratio > outside_change_ratio_limit:
            raise PptxRegressionError(
                f"Slide {slide_number} mudou fora dos targets: "
                f"{changed_count}/{outside_count} pixels ({ratio:.4%}); "
                f"limite {outside_change_ratio_limit:.4%}."
            )
        checked_pixels += outside_count
        changed_pixels += changed_count
        worst_ratio = max(worst_ratio, ratio)

    return RenderRegressionReport(
        slides_checked=len(original_images),
        pixels_checked_outside_targets=checked_pixels,
        pixels_changed_outside_targets=changed_pixels,
        worst_outside_change_ratio=worst_ratio,
    )


def _rendered_slide_paths(directory: Path) -> dict[int, Path]:
    output: dict[int, Path] = {}
    for path in directory.glob("slide-*.png"):
        match = re.fullmatch(r"slide-(\d+)\.png", path.name, flags=re.IGNORECASE)
        if match:
            output[int(match.group(1))] = path
    return output


def _slide_size_inches(pptx_bytes: bytes) -> tuple[float, float]:
    with ZipFile(BytesIO(pptx_bytes)) as archive:
        root = ET.fromstring(archive.read("ppt/presentation.xml"))
    size = root.find("./p:sldSz", REGRESSION_NS)
    if size is None:
        return 13.333333, 7.5
    width = int(size.attrib.get("cx") or 12192000) / 914400
    height = int(size.attrib.get("cy") or 6858000) / 914400
    return width, height


def _allowed_changed_parts(
    plans: list[TransformPlan],
    targets: list[PptTarget],
) -> set[str]:
    allowed = {target.slide_path for target in targets if target.slide_path}
    for plan in plans:
        target = plan.target
        allowed.update(
            part
            for part in (
                target.slide_path,
                target.chart_xml,
                target.workbook_embedded,
            )
            if part
        )
    return allowed


def _validate_all_xml(archive: ZipFile) -> int:
    count = 0
    for name in archive.namelist():
        if not name.lower().endswith((".xml", ".rels")):
            continue
        try:
            ET.fromstring(archive.read(name))
        except ET.ParseError as exc:
            raise PptxRegressionError(f"XML invalido em {name}: {exc}.") from exc
        count += 1
    return count


def _validate_relationship_targets(archive: ZipFile) -> int:
    names = set(archive.namelist())
    checked = 0
    for rels_name in sorted(name for name in names if name.endswith(".rels")):
        root = ET.fromstring(archive.read(rels_name))
        source_part = _source_part_for_relationships(rels_name)
        source_dir = posixpath.dirname(source_part)
        for relationship in root.findall(f"./{{{REL_NS}}}Relationship"):
            if str(relationship.attrib.get("TargetMode") or "").lower() == "external":
                continue
            target = unquote(str(relationship.attrib.get("Target") or "")).split("#", 1)[0]
            if not target:
                continue
            if urlparse(target).scheme:
                continue
            if target.startswith("/"):
                resolved = target.lstrip("/")
            else:
                resolved = posixpath.normpath(posixpath.join(source_dir, target))
            if resolved.startswith("../") or PurePosixPath(resolved).is_absolute():
                raise PptxRegressionError(
                    f"Relacionamento interno inseguro em {rels_name}: {target}."
                )
            if resolved not in names:
                raise PptxRegressionError(
                    f"Relacionamento quebrado em {rels_name}: {target} -> {resolved}."
                )
            checked += 1
    return checked


def _source_part_for_relationships(rels_name: str) -> str:
    if rels_name == "_rels/.rels":
        return ""
    marker = "/_rels/"
    if marker not in rels_name or not rels_name.endswith(".rels"):
        return ""
    prefix, filename = rels_name.split(marker, 1)
    return f"{prefix}/{filename[:-5]}"


def _validate_slide_geometry(original: ZipFile, generated: ZipFile) -> int:
    slide_names = sorted(
        name
        for name in original.namelist()
        if re.fullmatch(r"ppt/slides/slide\d+\.xml", name)
    )
    checked = 0
    for slide_name in slide_names:
        before = _slide_geometry(original.read(slide_name))
        after = _slide_geometry(generated.read(slide_name))
        if before.keys() != after.keys():
            raise PptxRegressionError(f"Objetos do slide mudaram em {slide_name}.")
        for shape_id, geometry in before.items():
            if after[shape_id] != geometry:
                raise PptxRegressionError(
                    f"Geometria do objeto {shape_id} mudou em {slide_name}."
                )
            checked += 1
    return checked


def _slide_geometry(slide_xml: bytes) -> dict[str, tuple[Any, ...]]:
    root = ET.fromstring(slide_xml)
    output: dict[str, tuple[Any, ...]] = {}
    shape_specs = (
        ("p:sp", "./p:nvSpPr/p:cNvPr", "./p:spPr/a:xfrm"),
        ("p:pic", "./p:nvPicPr/p:cNvPr", "./p:spPr/a:xfrm"),
        ("p:graphicFrame", "./p:nvGraphicFramePr/p:cNvPr", "./p:xfrm"),
        ("p:cxnSp", "./p:nvCxnSpPr/p:cNvPr", "./p:spPr/a:xfrm"),
        ("p:grpSp", "./p:nvGrpSpPr/p:cNvPr", "./p:grpSpPr/a:xfrm"),
    )
    for shape_query, identity_query, geometry_query in shape_specs:
        for shape in root.findall(f".//{shape_query}", REGRESSION_NS):
            identity = shape.find(identity_query, REGRESSION_NS)
            if identity is None:
                continue
            shape_id = str(identity.attrib.get("id") or "")
            geometry = shape.find(geometry_query, REGRESSION_NS)
            output[shape_id] = _element_signature(geometry)
    return output


def _element_signature(element: ET.Element | None) -> tuple[Any, ...]:
    if element is None:
        return ()
    return (
        element.tag.rsplit("}", 1)[-1],
        tuple(sorted(element.attrib.items())),
        tuple(_element_signature(child) for child in list(element)),
    )


def _validate_plan_outputs(
    generated: ZipFile,
    plans: list[TransformPlan],
) -> tuple[int, int]:
    charts_checked = 0
    tables_checked = 0
    for plan in plans:
        if plan.object_type == "chart":
            _validate_chart_output(generated, plan)
            charts_checked += 1
        elif plan.object_type == "table":
            _validate_table_output(generated, plan)
            tables_checked += 1
    return charts_checked, tables_checked


def _validate_chart_output(archive: ZipFile, plan: TransformPlan) -> None:
    target = plan.target
    if not target.chart_xml or target.chart_xml not in archive.namelist():
        raise PptxRegressionError(f"Chart XML ausente para {plan.target_id}.")
    if not target.workbook_embedded or target.workbook_embedded not in archive.namelist():
        raise PptxRegressionError(f"Workbook embutido ausente para {plan.target_id}.")

    chart_root = ET.fromstring(archive.read(target.chart_xml))
    series_elements = chart_root.findall(".//c:ser", REGRESSION_NS)
    if len(series_elements) < len(plan.series):
        raise PptxRegressionError(
            f"Grafico {plan.target_id} truncou series: "
            f"{len(series_elements)} no PPT, {len(plan.series)} no plano."
        )
    actual_series = [
        _chart_series_name(series)
        for series in series_elements[: len(plan.series)]
    ]
    if actual_series != [str(value) for value in plan.series]:
        raise PptxRegressionError(
            f"Series divergentes em {plan.target_id}: {actual_series} != {plan.series}."
        )
    expected_formats = resolved_series_number_formats(target, plan)
    for index, series in enumerate(series_elements[: len(plan.series)]):
        expected_format = expected_formats[index] if index < len(expected_formats) else ""
        cache_format = series.find("./c:val/c:numRef/c:numCache/c:formatCode", REGRESSION_NS)
        actual_format = str(cache_format.text or "") if cache_format is not None else ""
        if actual_format != expected_format:
            raise PptxRegressionError(
                f"Formato numerico da serie {index + 1} de {plan.target_id} divergiu: "
                f"{actual_format!r} != {expected_format!r}."
            )
        label_format = series.find("./c:dLbls/c:numFmt", REGRESSION_NS)
        actual_label_format = (
            str(label_format.attrib.get("formatCode") or "")
            if label_format is not None
            else ""
        )
        if actual_label_format != expected_format:
            raise PptxRegressionError(
                f"Formato do rotulo da serie {index + 1} de {plan.target_id} divergiu: "
                f"{actual_label_format!r} != {expected_format!r}."
            )

    workbook = openpyxl.load_workbook(
        BytesIO(archive.read(target.workbook_embedded)),
        data_only=True,
        read_only=True,
    )
    try:
        worksheet = (
            workbook[target.sheet_name]
            if target.sheet_name and target.sheet_name in workbook.sheetnames
            else workbook.worksheets[0]
        )
        expected = _workbook_matrix(plan)
        for row_index, row in enumerate(expected, 1):
            for column_index, value in enumerate(row, 1):
                actual = worksheet.cell(row=row_index, column=column_index).value
                if not _cell_values_equal(actual, value):
                    raise PptxRegressionError(
                        f"Workbook de {plan.target_id} divergiu em "
                        f"{worksheet.title}!{worksheet.cell(row_index, column_index).coordinate}: "
                        f"{actual!r} != {value!r}."
                    )
    finally:
        workbook.close()


def _chart_series_name(series: ET.Element) -> str:
    cache_value = series.find("./c:tx//c:v", REGRESSION_NS)
    return str(cache_value.text or "") if cache_value is not None else ""


def _validate_table_output(archive: ZipFile, plan: TransformPlan) -> None:
    target = plan.target
    if target.slide_path not in archive.namelist():
        raise PptxRegressionError(f"Slide ausente para tabela {plan.target_id}.")
    root = ET.fromstring(archive.read(target.slide_path))
    frame = _find_target_frame(root, target)
    if frame is None:
        raise PptxRegressionError(f"Tabela {plan.target_id} nao encontrada no slide.")
    table = frame.find(".//a:tbl", REGRESSION_NS)
    if table is None:
        raise PptxRegressionError(f"Objeto {plan.target_id} deixou de ser tabela.")

    expected = plan.table_matrix or plan.values
    actual = [
        [
            "".join(text.text or "" for text in cell.findall(".//a:t", REGRESSION_NS))
            for cell in row.findall("./a:tc", REGRESSION_NS)
        ]
        for row in table.findall("./a:tr", REGRESSION_NS)
    ]
    if len(actual) < len(expected):
        raise PptxRegressionError(
            f"Tabela {plan.target_id} truncou linhas: {len(actual)} < {len(expected)}."
        )
    for row_index, expected_row in enumerate(expected):
        if len(actual[row_index]) < len(expected_row):
            raise PptxRegressionError(
                f"Tabela {plan.target_id} truncou colunas na linha {row_index + 1}."
            )
        for column_index, value in enumerate(expected_row):
            expected_text = _format_table_value(value, plan.number_format)
            if actual[row_index][column_index] != expected_text:
                raise PptxRegressionError(
                    f"Tabela {plan.target_id} divergiu em "
                    f"L{row_index + 1}C{column_index + 1}: "
                    f"{actual[row_index][column_index]!r} != {expected_text!r}."
                )
    if any(value == "None" for row in actual for value in row):
        raise PptxRegressionError(f"Tabela {plan.target_id} gravou texto 'None'.")


def _find_target_frame(root: ET.Element, target: PptTarget) -> ET.Element | None:
    for frame in root.findall(".//p:graphicFrame", REGRESSION_NS):
        identity = frame.find("./p:nvGraphicFramePr/p:cNvPr", REGRESSION_NS)
        if identity is None:
            continue
        if str(identity.attrib.get("id") or "") == str(target.shape_id):
            return frame
        if str(identity.attrib.get("name") or "") in {
            target.shape_name,
            target.target_id,
        }:
            return frame
    return None


def _format_table_value(value: Any, number_format: str) -> str:
    if value is None:
        return ""
    if number_format == "thousands_pt_br":
        try:
            number = float(value)
            if number.is_integer():
                return f"{int(number):,}".replace(",", ".")
        except (TypeError, ValueError):
            pass
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).replace(".", ",")
    return str(value)


def _cell_values_equal(actual: Any, expected: Any) -> bool:
    if actual in (None, "") and expected in (None, ""):
        return True
    if isinstance(actual, (int, float)) and isinstance(expected, (int, float)):
        tolerance = max(1e-9, abs(float(expected)) * 1e-9)
        return abs(float(actual) - float(expected)) <= tolerance
    return str(actual) == str(expected)
