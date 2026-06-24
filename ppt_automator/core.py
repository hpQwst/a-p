from __future__ import annotations

from dataclasses import dataclass, field
from difflib import SequenceMatcher
from io import BytesIO
from pathlib import Path
from typing import Any, BinaryIO, Iterable
from zipfile import ZIP_DEFLATED, ZipFile
import posixpath
import re
import shutil
import subprocess
import tempfile
import unicodedata
import xml.etree.ElementTree as ET

import openpyxl
from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter, range_boundaries


PML_NS = "http://schemas.openxmlformats.org/presentationml/2006/main"
DML_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
CHART_NS = "http://schemas.openxmlformats.org/drawingml/2006/chart"
REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
R_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
MC_NS = "http://schemas.openxmlformats.org/markup-compatibility/2006"
C16R2_NS = "http://schemas.microsoft.com/office/drawing/2015/06/chart"
C15_NS = "http://schemas.microsoft.com/office/drawing/2012/chart"
C16_NS = "http://schemas.microsoft.com/office/drawing/2014/chart"
C14_NS = "http://schemas.microsoft.com/office/drawing/2007/8/2/chart"
C16R3_NS = "http://schemas.microsoft.com/office/drawing/2017/03/chart"

NS = {
    "p": PML_NS,
    "a": DML_NS,
    "c": CHART_NS,
    "rel": REL_NS,
    "r": R_NS,
}

for prefix, uri in {
    "p": PML_NS,
    "a": DML_NS,
    "c": CHART_NS,
    "r": R_NS,
    "mc": MC_NS,
    "c14": C14_NS,
    "c15": C15_NS,
    "c16": C16_NS,
    "c16r2": C16R2_NS,
    "c16r3": C16R3_NS,
}.items():
    ET.register_namespace(prefix, uri)


InputFile = str | Path | bytes | bytearray | BinaryIO
FormulaMode = str


@dataclass(frozen=True)
class MappingRow:
    graph_id: str
    var_analise: str = ""
    tipo: str = ""
    periodo_agregado: str = ""
    metrica: str = ""
    ano: str = ""
    tipo_grafico: str = ""
    abertura: str = ""
    remove_linha: bool = False
    remove_coluna: bool = False
    atualizar_grafico: bool = True
    numero_slide: int | None = None
    nome_grafico: str = ""
    nome_original: str = ""


@dataclass(frozen=True)
class SourceTable:
    graph_id: str
    file_name: str
    sheet_name: str
    question: str
    headers: list[str]
    rows: list[str]
    values: list[list[Any]]
    respondents: dict[str, Any] = field(default_factory=dict)
    metadata: dict[str, str] = field(default_factory=dict)
    source_key: str = ""


@dataclass(frozen=True)
class ChartTarget:
    graph_id: str
    slide_number: int
    chart_path: str
    embedded_workbook_path: str
    sheet_name: str
    headers: list[str]
    rows: list[str]
    left_in: float
    top_in: float
    width_in: float
    height_in: float
    nearby_text: str = ""
    slide_text: str = ""


@dataclass(frozen=True)
class SourceMatch:
    source: SourceTable
    score: float
    reason: str
    header_score: float = 0.0
    row_score: float = 0.0
    variable_score: float = 0.0
    metadata_score: float = 0.0


@dataclass(frozen=True)
class ChartJob:
    graph_id: str
    status: str
    message: str
    mapping: MappingRow | None = None
    source: SourceTable | None = None
    target: ChartTarget | None = None
    headers: list[str] = field(default_factory=list)
    rows: list[str] = field(default_factory=list)
    values: list[list[Any]] = field(default_factory=list)
    match_score: float = 0.0
    match_reason: str = ""
    match_candidates: list[tuple[str, float, str]] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return self.status == "ok"


@dataclass(frozen=True)
class _SlideTextBox:
    text: str
    left_in: float
    top_in: float
    width_in: float
    height_in: float


def read_bytes(file: InputFile) -> bytes:
    if isinstance(file, (bytes, bytearray)):
        return bytes(file)
    if isinstance(file, (str, Path)):
        return Path(file).read_bytes()
    if hasattr(file, "seek"):
        file.seek(0)
    data = file.read()
    return data if isinstance(data, bytes) else bytes(data)


class FormulaCalculationError(RuntimeError):
    """Raised when a workbook contains formulas but no calculator can resolve them."""


def load_mapping(mapping_file: InputFile, formula_mode: FormulaMode = "auto") -> list[MappingRow]:
    workbook_bytes = prepare_workbook_values(read_bytes(mapping_file), formula_mode=formula_mode)
    wb = openpyxl.load_workbook(BytesIO(workbook_bytes), data_only=True, read_only=True)
    ws = wb["MAPEAMENTO"] if "MAPEAMENTO" in wb.sheetnames else wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    headers = [str(v).strip() if v is not None else "" for v in next(rows)]
    index = {name: i for i, name in enumerate(headers)}
    output: list[MappingRow] = []

    for raw in rows:
        graph_id = _graph_id(_cell(raw, index, "graph_id"))
        if not graph_id:
            continue
        output.append(
            MappingRow(
                graph_id=graph_id,
                var_analise=_text(_cell(raw, index, "var_analise")),
                tipo=_text(_cell(raw, index, "tipo")),
                periodo_agregado=_text(_cell(raw, index, "periodo_agregado")),
                metrica=_text(_cell(raw, index, "metrica")),
                ano=_text(_cell(raw, index, "ano")),
                tipo_grafico=_text(_cell(raw, index, "tipo_grafico")),
                abertura=_text(_cell(raw, index, "abertura")),
                remove_linha=_truthy(_cell(raw, index, "remove_linha")),
                remove_coluna=_truthy(_cell(raw, index, "remove_coluna")),
                atualizar_grafico=_truthy(_cell(raw, index, "atualizar_grafico"), default=True),
                numero_slide=_int_or_none(_cell(raw, index, "numero_slide")),
                nome_grafico=_text(_cell(raw, index, "nome_grafico")),
                nome_original=_graph_id(_cell(raw, index, "nome_original")),
            )
        )
    return output


def load_datasource_tables(datasources_zip: InputFile, formula_mode: FormulaMode = "auto") -> list[SourceTable]:
    output: list[SourceTable] = []
    with ZipFile(BytesIO(read_bytes(datasources_zip))) as zf:
        names = [n for n in zf.namelist() if n.lower().endswith(".xlsx")]
        for index, name in enumerate(names, 1):
            graph_id = _graph_id(Path(name).stem)
            workbook_bytes = prepare_workbook_values(zf.read(name), formula_mode=formula_mode)
            wb = openpyxl.load_workbook(BytesIO(workbook_bytes), data_only=True, read_only=True)
            ws = wb.worksheets[0]
            source_key = graph_id or f"source:{index}:{name}"
            output.append(_read_source_table(graph_id, name, ws, source_key=source_key))
    return output


def load_datasources(datasources_zip: InputFile, formula_mode: FormulaMode = "auto") -> dict[str, SourceTable]:
    output: dict[str, SourceTable] = {}
    for source in load_datasource_tables(datasources_zip, formula_mode=formula_mode):
        output[source.graph_id or source.source_key] = source
    return output


def read_source_table_from_workbook(
    workbook_file: InputFile,
    file_name: str,
    graph_id: str = "",
    formula_mode: FormulaMode = "auto",
    source_key: str = "",
) -> SourceTable:
    graph_id = _graph_id(graph_id)
    workbook_bytes = prepare_workbook_values(read_bytes(workbook_file), formula_mode=formula_mode)
    wb = openpyxl.load_workbook(BytesIO(workbook_bytes), data_only=True, read_only=True)
    ws = wb.worksheets[0]
    return _read_source_table(graph_id, file_name, ws, source_key=source_key or graph_id or file_name)


def load_ppt_targets(pptx_file: InputFile) -> dict[str, ChartTarget]:
    ppt_bytes = read_bytes(pptx_file)
    output: dict[str, ChartTarget] = {}
    with ZipFile(BytesIO(ppt_bytes)) as zf:
        slides = _sorted_slide_paths(zf)
        for slide_number, slide_path in enumerate(slides, 1):
            slide_rels_path = _rels_path(slide_path)
            if slide_rels_path not in zf.namelist():
                continue
            slide_rels = _relationship_map(zf.read(slide_rels_path))
            slide_root = ET.fromstring(zf.read(slide_path))
            text_boxes = _slide_text_boxes(slide_root)
            slide_text = _join_unique([box.text for box in text_boxes])
            for frame in slide_root.findall(".//p:graphicFrame", NS):
                cnv = frame.find("./p:nvGraphicFramePr/p:cNvPr", NS)
                if cnv is None:
                    continue
                graph_id = _graph_id(cnv.attrib.get("name"))
                if not graph_id:
                    continue
                chart_el = frame.find(".//c:chart", NS)
                if chart_el is None:
                    continue
                rel_id = chart_el.attrib.get(f"{{{R_NS}}}id")
                if not rel_id or rel_id not in slide_rels:
                    continue
                chart_path = _norm_join(slide_path, slide_rels[rel_id]["target"])
                chart_rels_path = _rels_path(chart_path)
                if chart_rels_path not in zf.namelist():
                    continue
                chart_rels = _relationship_map(zf.read(chart_rels_path))
                workbook_path = ""
                for rel in chart_rels.values():
                    if rel["type"].endswith("/package"):
                        workbook_path = _norm_join(chart_path, rel["target"])
                        break
                if not workbook_path:
                    continue
                headers, labels, sheet_name = _read_embedded_chart_workbook(zf, workbook_path)
                left, top, width, height = _graphic_frame_dimensions(frame)
                nearby_text = _nearby_text(left, top, width, height, text_boxes)
                output[graph_id] = ChartTarget(
                    graph_id=graph_id,
                    slide_number=slide_number,
                    chart_path=chart_path,
                    embedded_workbook_path=workbook_path,
                    sheet_name=sheet_name,
                    headers=headers,
                    rows=labels,
                    left_in=left,
                    top_in=top,
                    width_in=width,
                    height_in=height,
                    nearby_text=nearby_text,
                    slide_text=slide_text,
                )
    return output


def build_chart_jobs(
    pptx_file: InputFile,
    mapping_file: InputFile,
    datasources_zip: InputFile,
    formula_mode: FormulaMode = "auto",
    respect_update_flag: bool = False,
    auto_match_sources: bool = True,
    min_match_score: float = 0.62,
    ambiguous_score_gap: float = 0.08,
) -> list[ChartJob]:
    mappings = load_mapping(mapping_file, formula_mode=formula_mode)
    source_tables = load_datasource_tables(datasources_zip, formula_mode=formula_mode)
    sources = {source.graph_id: source for source in source_tables if source.graph_id}
    targets = load_ppt_targets(pptx_file)
    jobs: list[ChartJob] = []
    used_source_keys: set[str] = set()

    for mapping in mappings:
        source = sources.get(mapping.graph_id)
        target = targets.get(mapping.graph_id)
        match_score = 1.0 if source is not None else 0.0
        match_reason = "Nome do arquivo bate com o graph_id." if source is not None else ""
        candidates: list[tuple[str, float, str]] = []
        if respect_update_flag and not mapping.atualizar_grafico:
            jobs.append(
                ChartJob(
                    graph_id=mapping.graph_id,
                    status="skipped",
                    message="Marcado como nao atualizar_grafico no mapeamento.",
                    mapping=mapping,
                    source=source,
                    target=target,
                    match_score=match_score,
                    match_reason=match_reason,
                )
            )
            continue
        if source is None and target is not None and auto_match_sources:
            matches = suggest_source_matches(
                mapping,
                target,
                source_tables,
                used_source_keys=used_source_keys,
            )
            candidates = [(m.source.file_name, m.score, m.reason) for m in matches[:5]]
            if matches:
                best = matches[0]
                second_score = matches[1].score if len(matches) > 1 else 0.0
                if best.score >= min_match_score and best.score - second_score >= ambiguous_score_gap:
                    source = best.source
                    match_score = best.score
                    match_reason = f"Datasource encontrado automaticamente: {best.reason}"
                elif best.score >= min_match_score:
                    jobs.append(
                        ChartJob(
                            graph_id=mapping.graph_id,
                            status="ambiguous_source",
                            message=(
                                "Encontrei datasources parecidos demais. "
                                f"Melhores: {_candidate_summary(candidates[:3])}."
                            ),
                            mapping=mapping,
                            target=target,
                            match_score=best.score,
                            match_reason=best.reason,
                            match_candidates=candidates,
                        )
                    )
                    continue
        if source is None:
            message = "Datasource nao encontrado no ZIP."
            if candidates:
                message += f" Melhor candidato: {_candidate_summary(candidates[:1])}."
            jobs.append(
                ChartJob(
                    graph_id=mapping.graph_id,
                    status="missing_source",
                    message=message,
                    mapping=mapping,
                    target=target,
                    match_score=candidates[0][1] if candidates else 0.0,
                    match_reason=candidates[0][2] if candidates else "",
                    match_candidates=candidates,
                )
            )
            continue
        if target is None:
            jobs.append(
                ChartJob(
                    graph_id=mapping.graph_id,
                    status="missing_target",
                    message="Grafico com esse nome nao encontrado no PPT.",
                    mapping=mapping,
                    source=source,
                    match_score=match_score,
                    match_reason=match_reason,
                    match_candidates=candidates,
                )
            )
            continue
        headers, labels, values = _build_chart_matrix(source, target)
        if source.source_key:
            used_source_keys.add(source.source_key)
        jobs.append(
            build_chart_job(
                mapping,
                source,
                target,
                match_score=match_score,
                match_reason=match_reason,
                match_candidates=candidates,
                headers=headers,
                rows=labels,
                values=values,
            )
        )
    return jobs


def build_auto_chart_jobs(
    pptx_file: InputFile,
    datasources_zip: InputFile,
    formula_mode: FormulaMode = "auto",
    min_match_score: float = 0.68,
) -> list[ChartJob]:
    source_tables = load_datasource_tables(datasources_zip, formula_mode=formula_mode)
    targets = load_ppt_targets(pptx_file)
    pair_scores: list[tuple[float, str, SourceMatch]] = []
    candidates_by_target: dict[str, list[tuple[str, float, str]]] = {}

    for target in targets.values():
        mapping = _inferred_mapping_for_target(target)
        matches = suggest_source_matches(mapping, target, source_tables)
        candidates_by_target[target.graph_id] = [
            (match.source.file_name, match.score, match.reason) for match in matches[:5]
        ]
        for match in matches[:5]:
            pair_scores.append((match.score, target.graph_id, match))

    pair_scores.sort(key=lambda item: item[0], reverse=True)
    used_targets: set[str] = set()
    used_sources: set[str] = set()
    jobs: list[ChartJob] = []

    for score, graph_id, match in pair_scores:
        source_key = match.source.source_key or match.source.file_name
        if score < min_match_score:
            break
        if graph_id in used_targets or source_key in used_sources:
            continue
        target = targets.get(graph_id)
        if target is None:
            continue
        mapping = _inferred_mapping_for_target(target)
        jobs.append(
            build_chart_job(
                mapping,
                match.source,
                target,
                match_score=score,
                match_reason=f"Modo automatico: {match.reason}",
                match_candidates=candidates_by_target.get(graph_id, []),
            )
        )
        used_targets.add(graph_id)
        used_sources.add(source_key)

    return sorted(
        jobs,
        key=lambda job: (
            job.target.slide_number if job.target else 9999,
            job.target.top_in if job.target else 9999,
            job.target.left_in if job.target else 9999,
        ),
    )


def build_chart_job(
    mapping: MappingRow,
    source: SourceTable,
    target: ChartTarget,
    match_score: float = 1.0,
    match_reason: str = "Selecionado manualmente.",
    match_candidates: list[tuple[str, float, str]] | None = None,
    headers: list[str] | None = None,
    rows: list[str] | None = None,
    values: list[list[Any]] | None = None,
) -> ChartJob:
    if headers is None or rows is None or values is None:
        headers, rows, values = _build_chart_matrix(source, target)
    return ChartJob(
        graph_id=mapping.graph_id,
        status="ok",
        message="Pronto para atualizar.",
        mapping=mapping,
        source=source,
        target=target,
        headers=headers,
        rows=rows,
        values=values,
        match_score=match_score,
        match_reason=match_reason,
        match_candidates=match_candidates or [],
    )


def suggest_source_matches(
    mapping: MappingRow,
    target: ChartTarget,
    sources: Iterable[SourceTable],
    used_source_keys: set[str] | None = None,
) -> list[SourceMatch]:
    used_source_keys = used_source_keys or set()
    matches: list[SourceMatch] = []
    for source in sources:
        if source.source_key and source.source_key in used_source_keys:
            continue
        matches.append(_score_source_match(mapping, target, source))
    return sorted(matches, key=lambda match: match.score, reverse=True)


def generate_pptx(pptx_file: InputFile, jobs: Iterable[ChartJob]) -> bytes:
    ppt_bytes = read_bytes(pptx_file)
    replacements: dict[str, bytes] = {}

    with ZipFile(BytesIO(ppt_bytes)) as zf:
        for job in jobs:
            if not job.ok or job.target is None:
                continue
            replacements[job.target.embedded_workbook_path] = _updated_workbook_bytes(
                zf,
                job.target.embedded_workbook_path,
                job.headers,
                job.rows,
                job.values,
            )
            replacements[job.target.chart_path] = _updated_chart_xml_bytes(
                zf,
                job.target.chart_path,
                job.headers,
                job.rows,
                job.values,
            )

        output = BytesIO()
        with ZipFile(output, "w", ZIP_DEFLATED) as zout:
            for info in zf.infolist():
                data = replacements.get(info.filename)
                zout.writestr(info, data if data is not None else zf.read(info.filename))
    return output.getvalue()


def _is_formula_value(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _clean_coord(coordinate: str | None) -> str:
    if not coordinate:
        return ""
    return coordinate.replace("$", "").upper()


def _flatten(values: Iterable[Any]) -> Iterable[Any]:
    for value in values:
        if isinstance(value, (list, tuple)):
            yield from _flatten(value)
        else:
            yield value


def _to_number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "."))
    except ValueError:
        return 0.0


def _numeric_values(values: Iterable[Any]) -> list[float]:
    return [_to_number(value) for value in _flatten(values) if value not in (None, "")]


def _average(values: list[float]) -> float:
    return sum(values) / len(values) if values else 0.0


def _matches_criteria(value: Any, criteria: Any) -> bool:
    text = str(criteria).strip().strip('"')
    operators = [
        (">=", lambda a, b: a >= b),
        ("<=", lambda a, b: a <= b),
        ("!=", lambda a, b: a != b),
        ("<>", lambda a, b: a != b),
        (">", lambda a, b: a > b),
        ("<", lambda a, b: a < b),
        ("=", lambda a, b: a == b),
    ]
    for operator, compare in operators:
        if text.startswith(operator):
            expected = text[len(operator) :].strip()
            try:
                return compare(_to_number(value), _to_number(expected))
            except Exception:
                return compare(str(value), expected)
    return _norm(value) == _norm(text)


def prepare_workbook_values(workbook_bytes: bytes, formula_mode: FormulaMode = "auto") -> bytes:
    """Return workbook bytes with formulas resolved to static values when requested.

    `openpyxl` can read cached formula results, but it cannot calculate formulas.
    In auto mode we only launch Excel when formulas are actually present.
    """

    normalized_mode = (formula_mode or "cached").lower()
    if normalized_mode in {"cached", "none", "off"}:
        return workbook_bytes
    if normalized_mode not in {"auto", "excel", "libreoffice", "internal"}:
        raise ValueError("formula_mode must be 'auto', 'excel', 'libreoffice', 'internal' or 'cached'.")
    if normalized_mode == "auto" and not workbook_has_formulas(workbook_bytes):
        return workbook_bytes
    if normalized_mode == "internal":
        return calculate_workbook_with_internal_engine(workbook_bytes)
    if normalized_mode == "libreoffice":
        return calculate_workbook_with_libreoffice(workbook_bytes)
    try:
        return calculate_workbook_with_excel(workbook_bytes)
    except FormulaCalculationError:
        if normalized_mode == "auto":
            try:
                return calculate_workbook_with_libreoffice(workbook_bytes)
            except FormulaCalculationError:
                return calculate_workbook_with_internal_engine(workbook_bytes)
        raise


def workbook_has_formulas(workbook_bytes: bytes) -> bool:
    wb = openpyxl.load_workbook(BytesIO(workbook_bytes), data_only=False, read_only=True)
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    value = cell.value
                    if getattr(cell, "data_type", None) == "f":
                        return True
                    if isinstance(value, str) and value.startswith("="):
                        return True
        return False
    finally:
        wb.close()


def calculate_workbook_with_excel(workbook_bytes: bytes) -> bytes:
    """Use local Excel COM automation to calculate and replace formulas with values."""

    try:
        import pythoncom
        import win32com.client
    except Exception as exc:  # pragma: no cover - depends on host machine
        raise FormulaCalculationError(
            "Este arquivo tem formulas, mas o Excel/pywin32 nao esta disponivel para calcula-las."
        ) from exc

    with tempfile.TemporaryDirectory(prefix="ppt_automator_excel_") as tmp_dir:
        tmp = Path(tmp_dir)
        input_path = tmp / "input.xlsx"
        output_path = tmp / "calculated.xlsx"
        input_path.write_bytes(workbook_bytes)
        shutil.copy2(input_path, output_path)

        excel = None
        workbook = None
        pythoncom.CoInitialize()
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.AskToUpdateLinks = False
            workbook = excel.Workbooks.Open(str(output_path), UpdateLinks=0, ReadOnly=False)
            excel.CalculateFullRebuild()
            for sheet in workbook.Worksheets:
                used_range = sheet.UsedRange
                used_range.Value = used_range.Value
            workbook.Save()
            workbook.Close(SaveChanges=True)
            workbook = None
            return output_path.read_bytes()
        except Exception as exc:  # pragma: no cover - depends on host machine
            raise FormulaCalculationError(f"Nao consegui calcular formulas com o Excel: {exc}") from exc
        finally:
            if workbook is not None:
                try:
                    workbook.Close(SaveChanges=False)
                except Exception:
                    pass
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    pass
            pythoncom.CoUninitialize()


def calculate_workbook_with_libreoffice(workbook_bytes: bytes) -> bytes:
    """Use LibreOffice headless to recalculate workbook formula caches."""

    executable = shutil.which("soffice") or shutil.which("libreoffice")
    if not executable:
        raise FormulaCalculationError("LibreOffice nao esta disponivel para calcular formulas.")

    with tempfile.TemporaryDirectory(prefix="ppt_automator_libreoffice_") as tmp_dir:
        tmp = Path(tmp_dir)
        input_path = tmp / "input.xlsx"
        output_dir = tmp / "out"
        output_dir.mkdir()
        input_path.write_bytes(workbook_bytes)

        command = [
            executable,
            "--headless",
            "--nologo",
            "--nofirststartwizard",
            "--nolockcheck",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(output_dir),
            str(input_path),
        ]
        try:
            completed = subprocess.run(
                command,
                capture_output=True,
                text=True,
                timeout=180,
                check=False,
            )
        except Exception as exc:  # pragma: no cover - depends on host machine
            raise FormulaCalculationError(f"Nao consegui executar LibreOffice: {exc}") from exc

        if completed.returncode != 0:
            detail = (completed.stderr or completed.stdout or "").strip()
            raise FormulaCalculationError(f"LibreOffice falhou ao calcular formulas: {detail}")

        output_path = output_dir / "input.xlsx"
        if not output_path.exists():
            candidates = list(output_dir.glob("*.xlsx"))
            if not candidates:
                raise FormulaCalculationError("LibreOffice nao gerou um arquivo XLSX recalculado.")
            output_path = candidates[0]
        return output_path.read_bytes()


def calculate_workbook_with_internal_engine(workbook_bytes: bytes) -> bytes:
    """Calculate common Excel formulas without launching Excel.

    This intentionally covers the formulas most often used in support tables:
    arithmetic, cell/range references, SUM/SOMA, AVERAGE/MEDIA, MIN, MAX,
    COUNT, COUNTA, IF/SE and simple SUMIF/COUNTIF criteria.
    Unsupported formulas fall back to cached values when the workbook has them.
    """

    formula_wb = openpyxl.load_workbook(BytesIO(workbook_bytes), data_only=False)
    cached_wb = openpyxl.load_workbook(BytesIO(workbook_bytes), data_only=True)
    evaluator = _SimpleFormulaEvaluator(formula_wb, cached_wb)

    for ws in formula_wb.worksheets:
        cached_ws = cached_wb[ws.title]
        for row in ws.iter_rows():
            for cell in row:
                if _is_formula_value(cell.value):
                    try:
                        cell.value = evaluator.evaluate_cell(ws.title, cell.coordinate)
                    except Exception:
                        cell.value = cached_ws[cell.coordinate].value

    out = BytesIO()
    formula_wb.save(out)
    return out.getvalue()


class _SimpleFormulaEvaluator:
    def __init__(self, formula_wb: Any, cached_wb: Any) -> None:
        self.formula_wb = formula_wb
        self.cached_wb = cached_wb
        self.memo: dict[tuple[str, str], Any] = {}
        self.stack: set[tuple[str, str]] = set()

    def evaluate_cell(self, sheet_name: str, coordinate: str) -> Any:
        coordinate = coordinate.replace("$", "").upper()
        key = (sheet_name, coordinate)
        if key in self.memo:
            return self.memo[key]
        if key in self.stack:
            raise ValueError(f"Circular formula reference at {sheet_name}!{coordinate}")
        self.stack.add(key)
        try:
            ws = self.formula_wb[sheet_name]
            value = ws[coordinate].value
            if _is_formula_value(value):
                result = self.evaluate_formula(str(value)[1:], sheet_name)
            else:
                result = value
            self.memo[key] = result
            return result
        finally:
            self.stack.remove(key)

    def evaluate_formula(self, formula: str, current_sheet: str) -> Any:
        expression = formula.strip()
        expression = expression.replace(";", ",")
        expression = re.sub(r"(?<=\d)%", "/100", expression)
        expression = expression.replace("^", "**")
        expression = expression.replace("<>", "!=")
        expression = re.sub(r"(?<![<>=!])=(?!=)", "==", expression)

        strings: list[str] = []

        def keep_string(match: re.Match[str]) -> str:
            strings.append(match.group(0))
            return f"__STR{len(strings) - 1}__"

        expression = re.sub(r'"[^"]*"', keep_string, expression)

        refs: list[str] = []

        def keep_ref(code: str) -> str:
            refs.append(code)
            return f"__REF{len(refs) - 1}__"

        expression = self._replace_qualified_refs(expression, keep_ref)
        expression = self._replace_unqualified_refs(expression, current_sheet, keep_ref)
        expression = self._replace_functions(expression)
        expression = self._replace_boolean_names(expression)

        for i, code in enumerate(refs):
            expression = expression.replace(f"__REF{i}__", code)
        for i, value in enumerate(strings):
            expression = expression.replace(f"__STR{i}__", value)

        env = self._eval_env(current_sheet)
        return eval(expression, {"__builtins__": {}}, env)

    def _replace_qualified_refs(self, expression: str, keep_ref: Any) -> str:
        quoted = r"'([^']+)'"
        plain = r"([A-Za-z_][A-Za-z0-9_ ]*)"
        cell = r"\$?[A-Za-z]{1,3}\$?\d+"
        pattern = re.compile(rf"(?:(?:{quoted})|(?:{plain}))!({cell})(?::({cell}))?")

        def repl(match: re.Match[str]) -> str:
            sheet = match.group(1) or match.group(2)
            start = _clean_coord(match.group(3))
            end = _clean_coord(match.group(4)) if match.group(4) else None
            if end:
                return keep_ref(f'_range({sheet!r}, "{start}:{end}")')
            return keep_ref(f'_cell({sheet!r}, "{start}")')

        return pattern.sub(repl, expression)

    def _replace_unqualified_refs(self, expression: str, current_sheet: str, keep_ref: Any) -> str:
        cell = r"\$?[A-Za-z]{1,3}\$?\d+"
        range_pattern = re.compile(rf"(?<![A-Za-z0-9_])({cell}):({cell})(?![A-Za-z0-9_])")
        cell_pattern = re.compile(rf"(?<![A-Za-z0-9_])({cell})(?![A-Za-z0-9_])")

        expression = range_pattern.sub(
            lambda m: keep_ref(
                f'_range({current_sheet!r}, "{_clean_coord(m.group(1))}:{_clean_coord(m.group(2))}")'
            ),
            expression,
        )
        return cell_pattern.sub(
            lambda m: keep_ref(f'_cell({current_sheet!r}, "{_clean_coord(m.group(1))}")'),
            expression,
        )

    def _replace_functions(self, expression: str) -> str:
        function_map = {
            "SUM": "_sum",
            "SOMA": "_sum",
            "AVERAGE": "_avg",
            "MEDIA": "_avg",
            "MÉDIA": "_avg",
            "MIN": "_min",
            "MAX": "_max",
            "COUNT": "_count",
            "CONT.NÚM": "_count",
            "CONT.NUM": "_count",
            "COUNTA": "_counta",
            "CONT.VALORES": "_counta",
            "IF": "_if",
            "SE": "_if",
            "ROUND": "round",
            "ARRED": "round",
            "ARREDONDAR": "round",
            "SUMIF": "_sumif",
            "SOMASE": "_sumif",
            "COUNTIF": "_countif",
            "CONT.SE": "_countif",
            "CONTSE": "_countif",
        }

        def repl(match: re.Match[str]) -> str:
            name = match.group(1).upper()
            return function_map.get(name, name) + "("

        return re.sub(r"\b([A-Za-zÀ-Úà-ú.]+)\s*\(", repl, expression)

    def _replace_boolean_names(self, expression: str) -> str:
        replacements = {
            "TRUE": "True",
            "FALSE": "False",
            "VERDADEIRO": "True",
            "FALSO": "False",
        }
        for old, new in replacements.items():
            expression = re.sub(rf"\b{old}\b", new, expression, flags=re.IGNORECASE)
        return expression

    def _eval_env(self, current_sheet: str) -> dict[str, Any]:
        return {
            "_cell": lambda sheet, coord: self._cell_value(sheet or current_sheet, coord),
            "_range": lambda sheet, ref: self._range_values(sheet or current_sheet, ref),
            "_sum": lambda *args: sum(_numeric_values(args)),
            "_avg": lambda *args: _average(_numeric_values(args)),
            "_min": lambda *args: min(_numeric_values(args), default=0),
            "_max": lambda *args: max(_numeric_values(args), default=0),
            "_count": lambda *args: len(_numeric_values(args)),
            "_counta": lambda *args: sum(1 for value in _flatten(args) if value not in (None, "")),
            "_if": lambda condition, true_value, false_value=None: true_value if condition else false_value,
            "_sumif": self._sumif,
            "_countif": self._countif,
            "round": round,
            "abs": abs,
        }

    def _cell_value(self, sheet_name: str, coordinate: str) -> Any:
        coordinate = _clean_coord(coordinate)
        if sheet_name not in self.formula_wb.sheetnames:
            return 0
        value = self.formula_wb[sheet_name][coordinate].value
        if _is_formula_value(value):
            return self.evaluate_cell(sheet_name, coordinate)
        return 0 if value is None else value

    def _range_values(self, sheet_name: str, ref: str) -> list[Any]:
        if sheet_name not in self.formula_wb.sheetnames:
            return []
        ws = self.formula_wb[sheet_name]
        min_col, min_row, max_col, max_row = range_boundaries(ref.replace("$", ""))
        values: list[Any] = []
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                coord = f"{get_column_letter(col)}{row}"
                values.append(self._cell_value(ws.title, coord))
        return values

    def _sumif(self, criteria_range: Any, criteria: Any, sum_range: Any = None) -> float:
        criteria_values = list(_flatten([criteria_range]))
        sum_values = list(_flatten([sum_range if sum_range is not None else criteria_range]))
        total = 0.0
        for i, value in enumerate(criteria_values):
            if _matches_criteria(value, criteria):
                total += _to_number(sum_values[i] if i < len(sum_values) else 0)
        return total

    def _countif(self, criteria_range: Any, criteria: Any) -> int:
        return sum(1 for value in _flatten([criteria_range]) if _matches_criteria(value, criteria))


def _cell(row: tuple[Any, ...], index: dict[str, int], name: str) -> Any:
    i = index.get(name)
    if i is None or i >= len(row):
        return None
    return row[i]


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _graph_id(value: Any) -> str:
    text = _text(value)
    if not text:
        return ""
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    digits = re.sub(r"\D", "", text)
    if not digits:
        return ""
    return digits.zfill(10) if len(digits) < 10 else digits


def _truthy(value: Any, default: bool = False) -> bool:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    return str(value).strip().lower() in {"1", "true", "sim", "yes", "y"}


def _int_or_none(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _candidate_summary(candidates: list[tuple[str, float, str]]) -> str:
    return ", ".join(f"{Path(name).name} ({score:.0%})" for name, score, _reason in candidates)


def _inferred_mapping_for_target(target: ChartTarget) -> MappingRow:
    label = _best_target_label(target)
    return MappingRow(
        graph_id=target.graph_id,
        var_analise=label,
        nome_grafico=label,
        numero_slide=target.slide_number,
    )


def _best_target_label(target: ChartTarget) -> str:
    text = target.nearby_text or target.slide_text or " ".join(target.rows[:3])
    if not text:
        return f"Grafico {target.graph_id}"
    return text[:160]


def _score_source_match(mapping: MappingRow, target: ChartTarget, source: SourceTable) -> SourceMatch:
    if mapping.graph_id and source.graph_id and mapping.graph_id == source.graph_id:
        return SourceMatch(source, 1.0, "graph_id igual ao nome do arquivo", 1.0, 1.0, 1.0, 1.0)

    source_text = " ".join(
        [
            source.file_name,
            source.sheet_name,
            source.question,
            " ".join(source.headers),
            " ".join(source.rows[:12]),
            " ".join(source.metadata.values()),
        ]
    )
    target_text = " ".join(
        [
            target.graph_id,
            target.nearby_text,
            target.slide_text,
            " ".join(target.headers),
            " ".join(target.rows[:12]),
        ]
    )
    variable_hints = " ".join(
        [
            mapping.var_analise,
            mapping.tipo,
            mapping.metrica,
            mapping.nome_grafico,
            mapping.nome_original,
        ]
    )
    abertura_hints = " ".join([mapping.abertura, mapping.nome_grafico])

    header_score = _coverage_score(target.headers, source.headers)
    row_score = _coverage_score(target.rows, source.rows)
    variable_score = max(
        _soft_text_score(mapping.var_analise, source.question),
        _soft_text_score(variable_hints, source_text),
        _soft_text_score(target_text, source_text) * 0.7,
    )
    abertura_score = max(
        _soft_text_score(mapping.abertura, " ".join(source.headers)),
        _soft_text_score(abertura_hints, " ".join(source.headers)),
    )
    metadata_score = _metadata_match_score(mapping, target, source)

    score = (
        0.44 * header_score
        + 0.34 * row_score
        + 0.14 * variable_score
        + 0.05 * abertura_score
        + 0.03 * metadata_score
    )
    score = max(0.0, min(1.0, score))
    reason = (
        f"colunas {header_score:.0%}, linhas {row_score:.0%}, "
        f"variavel {variable_score:.0%}, abertura {abertura_score:.0%}"
    )
    if metadata_score:
        reason += f", metadados {metadata_score:.0%}"
    return SourceMatch(
        source=source,
        score=score,
        reason=reason,
        header_score=header_score,
        row_score=row_score,
        variable_score=variable_score,
        metadata_score=max(abertura_score, metadata_score),
    )


def _coverage_score(targets: list[str], choices: list[str]) -> float:
    required = [_text(value) for value in targets if _text(value)]
    if not required:
        return 0.0
    if not choices:
        return 0.0
    return sum(_best_text_score(target, choices) for target in required) / len(required)


def _best_text_score(target: str, choices: list[str]) -> float:
    return max((_soft_text_score(target, choice) for choice in choices), default=0.0)


def _soft_text_score(left: Any, right: Any) -> float:
    left_norm = _norm(left)
    right_norm = _norm(right)
    if not left_norm or not right_norm:
        return 0.0
    if left_norm == right_norm:
        return 1.0

    left_compact = left_norm.replace(" ", "")
    right_compact = right_norm.replace(" ", "")
    if left_compact == right_compact:
        return 1.0
    if len(left_compact) > 2 and len(right_compact) > 2:
        if left_compact in right_compact or right_compact in left_compact:
            return 0.92

    left_tokens = set(left_norm.split())
    right_tokens = set(right_norm.split())
    token_score = 0.0
    if left_tokens and right_tokens:
        token_score = len(left_tokens & right_tokens) / max(len(left_tokens), len(right_tokens))
    ratio_score = SequenceMatcher(None, left_compact, right_compact).ratio()
    return max(token_score, ratio_score * 0.75)


def _metadata_match_score(mapping: MappingRow, target: ChartTarget, source: SourceTable) -> float:
    if not source.metadata:
        return 0.0
    scores: list[float] = []
    graph_id = _graph_id(source.metadata.get("graph_id"))
    if graph_id:
        scores.append(1.0 if graph_id == mapping.graph_id or graph_id == target.graph_id else 0.0)
    for field_name, expected in {
        "var_analise": mapping.var_analise,
        "abertura": mapping.abertura,
        "nome_grafico": mapping.nome_grafico,
        "nome_original": mapping.nome_original,
    }.items():
        observed = source.metadata.get(field_name)
        if observed and expected:
            scores.append(_soft_text_score(expected, observed))
    return max(scores, default=0.0)


def _extract_source_metadata(rows: list[list[Any]]) -> dict[str, str]:
    aliases = {
        "GRAPH ID": "graph_id",
        "GRAPHID": "graph_id",
        "ID GRAFICO": "graph_id",
        "GRAFICO": "nome_grafico",
        "NOME GRAFICO": "nome_grafico",
        "NOME DO GRAFICO": "nome_grafico",
        "NOME ORIGINAL": "nome_original",
        "VAR ANALISE": "var_analise",
        "VARIAVEL": "var_analise",
        "ABERTURA": "abertura",
        "QUEBRA": "abertura",
        "PPT TAG": "nome_grafico",
        "TAG PPT": "nome_grafico",
    }
    metadata: dict[str, str] = {}
    for row in rows[:25]:
        cells = [_text(value) for value in row]
        for index, cell in enumerate(cells):
            if not cell:
                continue
            key_text = cell
            value_text = cells[index + 1] if index + 1 < len(cells) else ""
            if ":" in cell:
                key_text, value_text = [part.strip() for part in cell.split(":", 1)]
            key = aliases.get(_norm(key_text))
            if key and value_text:
                metadata[key] = value_text
    return metadata


def _read_source_table(graph_id: str, file_name: str, ws: Any, source_key: str = "") -> SourceTable:
    raw_rows = [list(row) for row in ws.iter_rows(values_only=True)]
    metadata = _extract_source_metadata(raw_rows)
    header_index = _find_source_header_row(raw_rows)
    if header_index is None:
        return SourceTable(
            graph_id,
            file_name,
            ws.title,
            "",
            [],
            [],
            [],
            metadata=metadata,
            source_key=source_key or graph_id or file_name,
        )

    header_row = raw_rows[header_index]
    value_start = 3
    value_end = _last_non_empty_index(header_row)
    headers = [_text(v) for v in header_row[value_start : value_end + 1]]

    question = ""
    rows: list[str] = []
    values: list[list[Any]] = []
    respondents: dict[str, Any] = {}

    for raw in raw_rows[header_index + 1 :]:
        if _row_is_empty(raw):
            continue
        if not question:
            question = _text(raw[0] if len(raw) > 0 else "")
        label = _text(raw[1] if len(raw) > 1 else raw[0] if raw else "")
        if not label:
            continue
        row_values = list(raw[value_start : value_start + len(headers)])
        if _norm(label) == "RESPONDENTES":
            respondents = {headers[i]: row_values[i] for i in range(min(len(headers), len(row_values)))}
            continue
        rows.append(label)
        values.append(row_values)

    return SourceTable(
        graph_id=graph_id,
        file_name=file_name,
        sheet_name=ws.title,
        question=question,
        headers=headers,
        rows=rows,
        values=values,
        respondents=respondents,
        metadata=metadata,
        source_key=source_key or graph_id or file_name,
    )


def _find_source_header_row(rows: list[list[Any]]) -> int | None:
    for i, row in enumerate(rows):
        populated = [_text(v) for v in row[3:] if _text(v)]
        if len(populated) >= 2:
            return i
    return None


def _read_embedded_chart_workbook(zf: ZipFile, workbook_path: str) -> tuple[list[str], list[str], str]:
    wb = openpyxl.load_workbook(BytesIO(zf.read(workbook_path)), data_only=True, read_only=True)
    ws = wb.worksheets[0]
    headers: list[str] = []
    col = 2
    while col <= ws.max_column:
        value = _text(ws.cell(row=1, column=col).value)
        if not value:
            break
        headers.append(value)
        col += 1

    labels: list[str] = []
    row = 2
    while row <= ws.max_row:
        value = _text(ws.cell(row=row, column=1).value)
        if not value:
            break
        labels.append(value)
        row += 1
    return headers, labels, ws.title


def _build_chart_matrix(source: SourceTable, target: ChartTarget) -> tuple[list[str], list[str], list[list[Any]]]:
    source_header_indexes = [_match_index(target_header, source.headers) for target_header in target.headers]
    source_row_indexes = [_match_index(target_label, source.rows) for target_label in target.rows]

    values: list[list[Any]] = []
    for target_row_index, source_row_index in enumerate(source_row_indexes):
        if source_row_index is None and target_row_index < len(source.values):
            source_row_index = target_row_index
        output_row: list[Any] = []
        for target_col_index, source_col_index in enumerate(source_header_indexes):
            if source_col_index is None and target_col_index < len(source.headers):
                source_col_index = target_col_index
            value = None
            if source_row_index is not None and source_col_index is not None:
                try:
                    value = source.values[source_row_index][source_col_index]
                except IndexError:
                    value = None
            output_row.append(_normalize_chart_value(value))
        values.append(output_row)
    return list(target.headers), list(target.rows), values


def _match_index(target: str, choices: list[str]) -> int | None:
    target_norm = _norm(target)
    if not target_norm:
        return None
    normalized = [_norm(choice) for choice in choices]
    for i, choice_norm in enumerate(normalized):
        if target_norm == choice_norm:
            return i
    for i, choice_norm in enumerate(normalized):
        if target_norm in choice_norm or choice_norm in target_norm:
            return i
    tokens = [tok for tok in target_norm.split() if len(tok) > 2]
    if tokens:
        for i, choice_norm in enumerate(normalized):
            if all(tok in choice_norm for tok in tokens):
                return i
    return None


def _normalize_chart_value(value: Any) -> Any:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        if -1 <= float(value) <= 1:
            return round(float(value) * 100, 1)
        return round(float(value), 1)
    try:
        number = float(str(value).replace(",", "."))
    except ValueError:
        return value
    if -1 <= number <= 1:
        return round(number * 100, 1)
    return round(number, 1)


def _updated_workbook_bytes(
    zf: ZipFile,
    workbook_path: str,
    headers: list[str],
    rows: list[str],
    values: list[list[Any]],
) -> bytes:
    wb = openpyxl.load_workbook(BytesIO(zf.read(workbook_path)))
    ws = wb.worksheets[0]

    max_row = max(ws.max_row, len(rows) + 1)
    max_col = max(ws.max_column, len(headers) + 1)
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.value = None

    ws.cell(row=1, column=1).value = None
    for col_idx, header in enumerate(headers, 2):
        ws.cell(row=1, column=col_idx).value = header
    for row_idx, label in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1).value = label
        for col_idx, value in enumerate(values[row_idx - 2], 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            if isinstance(value, (int, float)):
                cell.number_format = "0.0"

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _updated_chart_xml_bytes(
    zf: ZipFile,
    chart_path: str,
    headers: list[str],
    rows: list[str],
    values: list[list[Any]],
) -> bytes:
    root = ET.fromstring(zf.read(chart_path))
    series = root.findall(".//c:ser", NS)
    end_col = _excel_col(len(headers) + 1)

    for index, ser in enumerate(series[: len(rows)]):
        excel_row = index + 2
        _update_series_text(ser, f"Planilha1!$A${excel_row}", rows[index])
        _update_series_categories(ser, f"Planilha1!$B$1:${end_col}$1", headers)
        _update_series_values(ser, f"Planilha1!$B${excel_row}:${end_col}${excel_row}", values[index])

    xml = ET.tostring(root, encoding="utf-8", xml_declaration=True)
    return xml.replace(b"'", b'"', 2)


def _update_series_text(ser: ET.Element, formula: str, label: str) -> None:
    tx = ser.find("./c:tx/c:strRef", NS)
    if tx is None:
        return
    _set_formula(tx, formula)
    cache = tx.find("./c:strCache", NS)
    if cache is None:
        cache = ET.SubElement(tx, f"{{{CHART_NS}}}strCache")
    _set_cache_values(cache, [label], numeric=False)


def _update_series_categories(ser: ET.Element, formula: str, labels: list[str]) -> None:
    cat = ser.find("./c:cat/c:strRef", NS)
    if cat is None:
        return
    _set_formula(cat, formula)
    cache = cat.find("./c:strCache", NS)
    if cache is None:
        cache = ET.SubElement(cat, f"{{{CHART_NS}}}strCache")
    _set_cache_values(cache, labels, numeric=False)


def _update_series_values(ser: ET.Element, formula: str, values: list[Any]) -> None:
    val = ser.find("./c:val/c:numRef", NS)
    if val is None:
        return
    _set_formula(val, formula)
    cache = val.find("./c:numCache", NS)
    if cache is None:
        cache = ET.SubElement(val, f"{{{CHART_NS}}}numCache")
    _set_cache_values(cache, values, numeric=True)


def _set_formula(parent: ET.Element, formula: str) -> None:
    formula_el = parent.find("./c:f", NS)
    if formula_el is None:
        formula_el = ET.SubElement(parent, f"{{{CHART_NS}}}f")
    formula_el.text = formula


def _set_cache_values(cache: ET.Element, values: list[Any], numeric: bool) -> None:
    for child in list(cache):
        if child.tag in {f"{{{CHART_NS}}}ptCount", f"{{{CHART_NS}}}pt"}:
            cache.remove(child)

    insert_at = 0
    for i, child in enumerate(list(cache)):
        if child.tag == f"{{{CHART_NS}}}formatCode":
            insert_at = i + 1

    pt_count = ET.Element(f"{{{CHART_NS}}}ptCount", {"val": str(len(values))})
    cache.insert(insert_at, pt_count)
    for offset, value in enumerate(values):
        pt = ET.Element(f"{{{CHART_NS}}}pt", {"idx": str(offset)})
        v = ET.SubElement(pt, f"{{{CHART_NS}}}v")
        v.text = _chart_value_text(value, numeric=numeric)
        cache.insert(insert_at + offset + 1, pt)


def _chart_value_text(value: Any, numeric: bool) -> str:
    if value is None:
        return "0" if numeric else ""
    if isinstance(value, float):
        return f"{value:.1f}".rstrip("0").rstrip(".")
    return str(value)


def _sorted_slide_paths(zf: ZipFile) -> list[str]:
    return sorted(
        [n for n in zf.namelist() if re.fullmatch(r"ppt/slides/slide\d+\.xml", n)],
        key=lambda name: int(re.search(r"slide(\d+)\.xml", name).group(1)),
    )


def _rels_path(part_path: str) -> str:
    directory, filename = posixpath.split(part_path)
    return posixpath.join(directory, "_rels", f"{filename}.rels")


def _relationship_map(xml_bytes: bytes) -> dict[str, dict[str, str]]:
    root = ET.fromstring(xml_bytes)
    output: dict[str, dict[str, str]] = {}
    for rel in root.findall(f"{{{REL_NS}}}Relationship"):
        output[rel.attrib["Id"]] = {
            "type": rel.attrib.get("Type", ""),
            "target": rel.attrib.get("Target", ""),
        }
    return output


def _norm_join(base_part: str, target: str) -> str:
    return posixpath.normpath(posixpath.join(posixpath.dirname(base_part), target))


def _graphic_frame_dimensions(frame: ET.Element) -> tuple[float, float, float, float]:
    xfrm = frame.find("./p:xfrm", NS)
    return _xfrm_dimensions(xfrm)


def _shape_dimensions(shape: ET.Element) -> tuple[float, float, float, float]:
    xfrm = shape.find("./p:spPr/a:xfrm", NS)
    return _xfrm_dimensions(xfrm)


def _xfrm_dimensions(xfrm: ET.Element | None) -> tuple[float, float, float, float]:
    if xfrm is None:
        return 0.0, 0.0, 0.0, 0.0
    off = xfrm.find("./a:off", NS)
    ext = xfrm.find("./a:ext", NS)
    if off is None or ext is None:
        return 0.0, 0.0, 0.0, 0.0
    emu_per_inch = 914400
    return (
        round(int(off.attrib.get("x", "0")) / emu_per_inch, 3),
        round(int(off.attrib.get("y", "0")) / emu_per_inch, 3),
        round(int(ext.attrib.get("cx", "0")) / emu_per_inch, 3),
        round(int(ext.attrib.get("cy", "0")) / emu_per_inch, 3),
    )


def _slide_text_boxes(slide_root: ET.Element) -> list[_SlideTextBox]:
    boxes: list[_SlideTextBox] = []
    for shape in slide_root.findall(".//p:sp", NS):
        texts = [_text(node.text) for node in shape.findall(".//a:t", NS)]
        text = _join_unique([part for part in texts if part])
        if not text:
            continue
        left, top, width, height = _shape_dimensions(shape)
        boxes.append(
            _SlideTextBox(
                text=text,
                left_in=left,
                top_in=top,
                width_in=width,
                height_in=height,
            )
        )
    return boxes


def _nearby_text(
    left: float,
    top: float,
    width: float,
    height: float,
    text_boxes: list[_SlideTextBox],
) -> str:
    if not text_boxes:
        return ""
    chart_center_x = left + width / 2
    chart_center_y = top + height / 2
    ranked: list[tuple[float, str]] = []
    for box in text_boxes:
        box_center_x = box.left_in + box.width_in / 2
        box_center_y = box.top_in + box.height_in / 2
        distance = abs(chart_center_x - box_center_x) + abs(chart_center_y - box_center_y)
        horizontal_overlap = _overlap_ratio(left, width, box.left_in, box.width_in)
        vertical_gap = min(abs((box.top_in + box.height_in) - top), abs((top + height) - box.top_in))
        above_or_inside = box.top_in <= top + height and box.top_in + box.height_in <= top + height + 0.5
        if horizontal_overlap >= 0.2 and above_or_inside:
            distance *= 0.45
        if vertical_gap <= 0.35:
            distance *= 0.7
        ranked.append((distance, box.text))
    ranked.sort(key=lambda item: item[0])
    return _join_unique([text for _distance, text in ranked[:8]])


def _overlap_ratio(left_a: float, width_a: float, left_b: float, width_b: float) -> float:
    right_a = left_a + width_a
    right_b = left_b + width_b
    overlap = max(0.0, min(right_a, right_b) - max(left_a, left_b))
    smaller = min(width_a, width_b)
    return overlap / smaller if smaller else 0.0


def _join_unique(values: Iterable[Any], separator: str = " | ") -> str:
    parts: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = _text(value)
        key = _norm(text)
        if not text or key in seen:
            continue
        parts.append(text)
        seen.add(key)
    return separator.join(parts)


def _last_non_empty_index(row: list[Any]) -> int:
    for i in range(len(row) - 1, -1, -1):
        if _text(row[i]):
            return i
    return -1


def _row_is_empty(row: list[Any]) -> bool:
    return all(_text(v) == "" for v in row)


def _norm(value: Any) -> str:
    text = _text(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.upper()
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _excel_col(index: int) -> str:
    letters = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters
