from __future__ import annotations

import argparse
import binascii
from io import BytesIO
from pathlib import Path
import re
import struct
import sys
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile
import zlib
import copy
import xml.etree.ElementTree as ET

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from ppt_automator.ppt_chart_writer import chart_replacements
from ppt_automator.ppt_discovery import PptTarget, discover_ppt_targets
from ppt_automator.table_normalizer import TransformPlan, normalize_to_target
from ppt_automator.xlsx_parser import parse_xlsx_table


DEFAULT_SHAPE_NAME = "1130655160"
SHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
DOC_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
MC_NS = "http://schemas.openxmlformats.org/markup-compatibility/2006"
X14AC_NS = "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac"
XR_NS = "http://schemas.microsoft.com/office/spreadsheetml/2014/revision"
XR2_NS = "http://schemas.microsoft.com/office/spreadsheetml/2015/revision2"
XR3_NS = "http://schemas.microsoft.com/office/spreadsheetml/2016/revision3"
XR6_NS = "http://schemas.microsoft.com/office/spreadsheetml/2016/revision6"
XR10_NS = "http://schemas.microsoft.com/office/spreadsheetml/2016/revision10"
X15_NS = "http://schemas.microsoft.com/office/spreadsheetml/2010/11/main"
XLWCV_NS = "http://schemas.microsoft.com/office/spreadsheetml/2024/workbookCompatibilityVersion"

for prefix, uri in {
    "": SHEET_NS,
    "r": DOC_REL_NS,
    "mc": MC_NS,
    "x14ac": X14AC_NS,
    "xr": XR_NS,
    "xr2": XR2_NS,
    "xr3": XR3_NS,
    "xr6": XR6_NS,
    "xr10": XR10_NS,
    "x15": X15_NS,
    "xlwcv": XLWCV_NS,
}.items():
    ET.register_namespace(prefix, uri)


def main() -> None:
    args = _parse_args()
    pptx_path = Path(args.pptx)
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    datasource_bytes, datasource_name = _read_datasource(args)
    target = _select_target(
        pptx_path,
        slide_number=args.slide,
        shape_name=args.shape_name,
        target_id=args.target_id,
        title_hint=args.title_hint,
    )
    source = parse_xlsx_table(datasource_bytes, file_name=datasource_name)
    plan = normalize_to_target(
        target,
        source,
        confidence=1.0,
        match_reason=f"Teste openpyxl/Fargate usando {datasource_name}.",
    )
    output_bytes = patch_pptx_surgically(
        pptx_path.read_bytes(),
        plan,
        writer=args.writer,
        update_chart_cache=not args.skip_chart_cache,
    )
    output_path.write_bytes(output_bytes)

    print("PPTX gerado:", output_path)
    print("Target:", plan.target_id, "| shape:", plan.target.shape_name)
    print("Chart XML:", plan.target.chart_xml)
    print("Workbook embutido:", plan.target.workbook_embedded)
    print("Aba preservada:", plan.target.sheet_name or "ativa")
    print("Datasource:", datasource_name)
    print("Writer:", args.writer)
    print("Cache visual atualizado:", "nao" if args.skip_chart_cache else "sim")
    print("Matriz:", len(_workbook_matrix(plan)), "linha(s) x", len(_workbook_matrix(plan)[0]), "coluna(s)")


def patch_pptx_surgically(
    pptx_bytes: bytes,
    plan: TransformPlan,
    writer: str = "xml",
    *,
    update_chart_cache: bool = True,
) -> bytes:
    replacements: dict[str, bytes] = {}
    matrix = _workbook_matrix(plan)
    with ZipFile(BytesIO(pptx_bytes)) as zf:
        if not plan.target.workbook_embedded:
            raise ValueError("O target selecionado nao possui workbook Excel embutido.")
        if plan.target.workbook_embedded not in zf.namelist():
            raise ValueError(f"Workbook embutido nao encontrado: {plan.target.workbook_embedded}")

        workbook_bytes = zf.read(plan.target.workbook_embedded)
        if writer == "openpyxl":
            replacements[plan.target.workbook_embedded] = update_embedded_workbook_openpyxl(
                workbook_bytes,
                plan.target.sheet_name,
                matrix,
            )
        elif writer == "sheet-values":
            replacements[plan.target.workbook_embedded] = update_embedded_workbook_sheet_values_only(
                workbook_bytes,
                plan.target.sheet_name,
                matrix,
            )
        elif writer == "raw-values":
            replacements[plan.target.workbook_embedded] = update_embedded_workbook_raw_values_only(
                workbook_bytes,
                plan.target.sheet_name,
                matrix,
            )
        else:
            replacements[plan.target.workbook_embedded] = update_embedded_workbook_xml(
                workbook_bytes,
                plan.target.sheet_name,
                matrix,
            )
        if update_chart_cache:
            replacements.update(chart_replacements(zf, plan.target, plan))

    return replace_zip_parts_preserving_structure(pptx_bytes, replacements)


def update_embedded_workbook_raw_values_only(
    workbook_bytes: bytes,
    sheet_name: str,
    matrix: list[list[Any]],
) -> bytes:
    row_count = len(matrix)
    col_count = max((len(row) for row in matrix), default=0)
    if row_count == 0 or col_count == 0:
        raise ValueError("Matriz vazia; nada para gravar no workbook embutido.")

    replacements: dict[str, bytes] = {}
    with ZipFile(BytesIO(workbook_bytes)) as workbook_zip:
        sheet_path = _worksheet_path_for_sheet(workbook_zip, sheet_name)
        sheet_xml = workbook_zip.read(sheet_path)
        replacements[sheet_path] = _raw_patch_sheet_values_xml(sheet_xml, matrix)

    return replace_zip_parts_preserving_structure(workbook_bytes, replacements)


def update_embedded_workbook_sheet_values_only(
    workbook_bytes: bytes,
    sheet_name: str,
    matrix: list[list[Any]],
) -> bytes:
    row_count = len(matrix)
    col_count = max((len(row) for row in matrix), default=0)
    if row_count == 0 or col_count == 0:
        raise ValueError("Matriz vazia; nada para gravar no workbook embutido.")

    replacements: dict[str, bytes] = {}
    with ZipFile(BytesIO(workbook_bytes)) as workbook_zip:
        sheet_path = _worksheet_path_for_sheet(workbook_zip, sheet_name)
        sheet_xml = workbook_zip.read(sheet_path)
        replacements[sheet_path] = _updated_sheet_values_only_xml(sheet_xml, matrix)

    return replace_zip_parts_preserving_structure(workbook_bytes, replacements)


def update_embedded_workbook_xml(
    workbook_bytes: bytes,
    sheet_name: str,
    matrix: list[list[Any]],
) -> bytes:
    row_count = len(matrix)
    col_count = max((len(row) for row in matrix), default=0)
    if row_count == 0 or col_count == 0:
        raise ValueError("Matriz vazia; nada para gravar no workbook embutido.")

    replacements: dict[str, bytes] = {}
    with ZipFile(BytesIO(workbook_bytes)) as workbook_zip:
        sheet_path = _worksheet_path_for_sheet(workbook_zip, sheet_name)
        shared_strings_path = "xl/sharedStrings.xml"
        use_shared_strings = shared_strings_path in workbook_zip.namelist()
        sheet_xml = workbook_zip.read(sheet_path)
        replacements[sheet_path] = _updated_sheet_xml(sheet_xml, matrix, use_shared_strings=use_shared_strings)
        if use_shared_strings:
            replacements[shared_strings_path] = _shared_strings_xml(_matrix_strings(matrix))
        table_path = _first_table_path(workbook_zip, sheet_path)
        if table_path:
            replacements[table_path] = _updated_table_xml(workbook_zip.read(table_path), matrix)

    return replace_zip_parts_preserving_structure(workbook_bytes, replacements)


def replace_zip_parts_preserving_structure(zip_bytes: bytes, replacements: dict[str, bytes]) -> bytes:
    if not replacements:
        return zip_bytes

    with ZipFile(BytesIO(zip_bytes)) as zin:
        infos = zin.infolist()
        names = {info.filename for info in infos}
        missing = sorted(set(replacements) - names)
        if missing:
            raise ValueError(f"Partes ZIP nao encontradas: {', '.join(missing)}")

        raw_end_by_name = _raw_local_record_ends(zip_bytes, infos, getattr(zin, "start_dir", len(zip_bytes)))
        output = BytesIO()
        central_records: list[bytes] = []
        for info in infos:
            local_offset = output.tell()
            replacement = replacements.get(info.filename)
            if replacement is None:
                start = info.header_offset
                end = raw_end_by_name[info.filename]
                output.write(zip_bytes[start:end])
                crc = info.CRC
                compress_size = info.compress_size
                file_size = info.file_size
            else:
                compressed, crc, compress_size, file_size = _compress_zip_payload(replacement, info.compress_type, info.flag_bits)
                output.write(_zip_local_header(info, crc, compress_size, file_size))
                output.write(_zip_filename_bytes(info))
                output.write(info.extra)
                output.write(compressed)
            central_records.append(_zip_central_header(info, crc, compress_size, file_size, local_offset))

        central_offset = output.tell()
        for record in central_records:
            output.write(record)
        central_size = output.tell() - central_offset
        output.write(
            struct.pack(
                "<IHHHHIIH",
                0x06054B50,
                0,
                0,
                len(infos),
                len(infos),
                central_size,
                central_offset,
                len(zin.comment),
            )
        )
        output.write(zin.comment)
    return output.getvalue()


def _raw_local_record_ends(zip_bytes: bytes, infos: list[Any], central_offset: int) -> dict[str, int]:
    sorted_infos = sorted(infos, key=lambda item: item.header_offset)
    output: dict[str, int] = {}
    for index, info in enumerate(sorted_infos):
        output[info.filename] = sorted_infos[index + 1].header_offset if index + 1 < len(sorted_infos) else central_offset
    return output


def _compress_zip_payload(data: bytes, compress_type: int, flag_bits: int) -> tuple[bytes, int, int, int]:
    crc = binascii.crc32(data) & 0xFFFFFFFF
    if compress_type == 0:
        compressed = data
    elif compress_type == ZIP_DEFLATED:
        level = 1 if flag_bits & 0x0006 == 0x0006 else 6
        compressor = zlib.compressobj(level=level, wbits=-15)
        compressed = compressor.compress(data) + compressor.flush()
    else:
        raise ValueError(f"Metodo de compressao ZIP nao suportado para patch cirurgico: {compress_type}")
    return compressed, crc, len(compressed), len(data)


def _zip_datetime(info: Any) -> tuple[int, int]:
    year, month, day, hour, minute, second = info.date_time
    dos_time = (hour << 11) | (minute << 5) | (second // 2)
    dos_date = ((year - 1980) << 9) | (month << 5) | day
    return dos_time, dos_date


def _zip_filename_bytes(info: Any) -> bytes:
    encoding = "utf-8" if info.flag_bits & 0x800 else "cp437"
    return info.filename.encode(encoding)


def _zip_local_header(info: Any, crc: int, compress_size: int, file_size: int) -> bytes:
    filename = _zip_filename_bytes(info)
    dos_time, dos_date = _zip_datetime(info)
    return struct.pack(
        "<IHHHHHIIIHH",
        0x04034B50,
        info.extract_version,
        info.flag_bits,
        info.compress_type,
        dos_time,
        dos_date,
        crc,
        compress_size,
        file_size,
        len(filename),
        len(info.extra),
    )


def _zip_central_header(info: Any, crc: int, compress_size: int, file_size: int, local_offset: int) -> bytes:
    filename = _zip_filename_bytes(info)
    comment = info.comment or b""
    dos_time, dos_date = _zip_datetime(info)
    version_made_by = (info.create_system << 8) | info.create_version
    return struct.pack(
        "<IHHHHHHIIIHHHHHII",
        0x02014B50,
        version_made_by,
        info.extract_version,
        info.flag_bits,
        info.compress_type,
        dos_time,
        dos_date,
        crc,
        compress_size,
        file_size,
        len(filename),
        len(info.extra),
        len(comment),
        getattr(info, "volume", 0),
        info.internal_attr,
        info.external_attr,
        local_offset,
    ) + filename + info.extra + comment


def _raw_patch_sheet_values_xml(sheet_xml: bytes, matrix: list[list[Any]]) -> bytes:
    patched = sheet_xml
    row_count = len(matrix)
    col_count = max((len(row) for row in matrix), default=0)
    for row_index in range(2, row_count + 1):
        row_values = matrix[row_index - 1] if row_index - 1 < len(matrix) else []
        for col_index in range(2, col_count + 1):
            value = row_values[col_index - 1] if col_index - 1 < len(row_values) else None
            cell_ref = f"{_excel_col(col_index)}{row_index}"
            patched = _replace_cell_value_bytes(patched, cell_ref, _xml_number_text(_number_or_original(value)))
    return patched


def _replace_cell_value_bytes(sheet_xml: bytes, cell_ref: str, value: str) -> bytes:
    cell_ref_bytes = re.escape(cell_ref.encode("ascii"))
    pattern = re.compile(
        rb'(<c\b(?=[^>]*\br="' + cell_ref_bytes + rb'")[^>]*>)(.*?)(</c>)',
        re.DOTALL,
    )
    match = pattern.search(sheet_xml)
    if not match:
        return sheet_xml

    prefix, body, suffix = match.groups()
    value_bytes = value.encode("ascii")
    if re.search(rb"<v>.*?</v>", body, flags=re.DOTALL):
        new_body = re.sub(rb"(<v>)(.*?)(</v>)", rb"\g<1>" + value_bytes + rb"\g<3>", body, count=1, flags=re.DOTALL)
    else:
        new_body = body + b"<v>" + value_bytes + b"</v>"
    return sheet_xml[: match.start()] + prefix + new_body + suffix + sheet_xml[match.end() :]


def _updated_sheet_values_only_xml(sheet_xml: bytes, matrix: list[list[Any]]) -> bytes:
    root = ET.fromstring(sheet_xml)
    sheet_data = root.find(f"{{{SHEET_NS}}}sheetData")
    if sheet_data is None:
        raise ValueError("Aba do workbook embutido nao tem sheetData para alterar.")

    cells_by_ref = {
        cell.attrib.get("r"): cell
        for row in sheet_data.findall(f"{{{SHEET_NS}}}row")
        for cell in row.findall(f"{{{SHEET_NS}}}c")
        if cell.attrib.get("r")
    }
    row_count = len(matrix)
    col_count = max((len(row) for row in matrix), default=0)
    for row_index in range(2, row_count + 1):
        row_values = matrix[row_index - 1] if row_index - 1 < len(matrix) else []
        for col_index in range(2, col_count + 1):
            cell = cells_by_ref.get(f"{_excel_col(col_index)}{row_index}")
            if cell is None:
                continue
            value = row_values[col_index - 1] if col_index - 1 < len(row_values) else None
            cell.attrib.pop("t", None)
            for child in list(cell):
                cell.remove(child)
            v = ET.SubElement(cell, f"{{{SHEET_NS}}}v")
            v.text = _xml_number_text(_number_or_original(value))
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _worksheet_path_for_sheet(workbook_zip: ZipFile, sheet_name: str) -> str:
    workbook_root = ET.fromstring(workbook_zip.read("xl/workbook.xml"))
    rels_root = ET.fromstring(workbook_zip.read("xl/_rels/workbook.xml.rels"))
    rels = {
        rel.attrib.get("Id"): rel.attrib.get("Target", "")
        for rel in rels_root.findall(f"{{{REL_NS}}}Relationship")
    }
    sheets = workbook_root.find(f"{{{SHEET_NS}}}sheets")
    if sheets is None:
        raise ValueError("Workbook embutido nao tem lista de abas.")
    selected_sheet = None
    for sheet in sheets.findall(f"{{{SHEET_NS}}}sheet"):
        if sheet_name and sheet.attrib.get("name") == sheet_name:
            selected_sheet = sheet
            break
        if selected_sheet is None:
            selected_sheet = sheet
    if selected_sheet is None:
        raise ValueError("Workbook embutido nao tem abas.")
    rel_id = selected_sheet.attrib.get(f"{{{DOC_REL_NS}}}id")
    target = rels.get(rel_id or "")
    if not target:
        raise ValueError(f"Nao encontrei relationship da aba '{selected_sheet.attrib.get('name')}'.")
    return _join_xlsx_path("xl/workbook.xml", target)


def _first_table_path(workbook_zip: ZipFile, sheet_path: str) -> str:
    rels_path = _rels_path(sheet_path)
    if rels_path not in workbook_zip.namelist():
        return ""
    rels_root = ET.fromstring(workbook_zip.read(rels_path))
    for rel in rels_root.findall(f"{{{REL_NS}}}Relationship"):
        if str(rel.attrib.get("Type") or "").endswith("/table"):
            return _join_xlsx_path(sheet_path, str(rel.attrib.get("Target") or ""))
    return ""


def _updated_sheet_xml(sheet_xml: bytes, matrix: list[list[Any]], *, use_shared_strings: bool) -> bytes:
    root = ET.fromstring(sheet_xml)
    row_count = len(matrix)
    col_count = max((len(row) for row in matrix), default=0)
    dimension = root.find(f"{{{SHEET_NS}}}dimension")
    if dimension is not None:
        dimension.attrib["ref"] = f"A1:{_excel_col(col_count)}{row_count}"

    sheet_data = root.find(f"{{{SHEET_NS}}}sheetData")
    if sheet_data is None:
        sheet_data = ET.Element(f"{{{SHEET_NS}}}sheetData")
        root.append(sheet_data)
    for child in list(sheet_data):
        sheet_data.remove(child)

    string_index = {value: index for index, value in enumerate(_matrix_strings(matrix))} if use_shared_strings else {}
    for row_index, row_values in enumerate(matrix, start=1):
        row = ET.Element(
            f"{{{SHEET_NS}}}row",
            {
                "r": str(row_index),
                "spans": f"1:{col_count}",
                f"{{{X14AC_NS}}}dyDescent": "0.25",
            },
        )
        for col_index in range(1, col_count + 1):
            value = row_values[col_index - 1] if col_index - 1 < len(row_values) else ""
            cell_ref = f"{_excel_col(col_index)}{row_index}"
            cell = ET.Element(f"{{{SHEET_NS}}}c", {"r": cell_ref})
            if row_index == 1 or col_index == 1:
                text = "" if value is None else str(value)
                if use_shared_strings:
                    cell.attrib["t"] = "s"
                    v = ET.SubElement(cell, f"{{{SHEET_NS}}}v")
                    v.text = str(string_index[text])
                else:
                    cell.attrib["t"] = "inlineStr"
                    inline = ET.SubElement(cell, f"{{{SHEET_NS}}}is")
                    text_node = ET.SubElement(inline, f"{{{SHEET_NS}}}t")
                    if text != text.strip() or text == "":
                        text_node.attrib["{http://www.w3.org/XML/1998/namespace}space"] = "preserve"
                    text_node.text = text
            else:
                number = _number_or_original(value)
                v = ET.SubElement(cell, f"{{{SHEET_NS}}}v")
                v.text = _xml_number_text(number)
            row.append(cell)
        sheet_data.append(row)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _updated_table_xml(table_xml: bytes, matrix: list[list[Any]]) -> bytes:
    root = ET.fromstring(table_xml)
    row_count = len(matrix)
    col_count = max((len(row) for row in matrix), default=0)
    root.attrib["ref"] = f"A1:{_excel_col(col_count)}{row_count}"
    auto_filter = root.find(f"{{{SHEET_NS}}}autoFilter")
    if auto_filter is not None:
        auto_filter.attrib["ref"] = root.attrib["ref"]
    table_columns = root.find(f"{{{SHEET_NS}}}tableColumns")
    if table_columns is None:
        table_columns = ET.SubElement(root, f"{{{SHEET_NS}}}tableColumns")
    old_columns = list(table_columns.findall(f"{{{SHEET_NS}}}tableColumn"))
    for child in old_columns:
        table_columns.remove(child)
    table_columns.attrib["count"] = str(col_count)
    header = matrix[0] if matrix else []
    for index in range(1, col_count + 1):
        template = copy.deepcopy(old_columns[index - 1]) if index - 1 < len(old_columns) else ET.Element(f"{{{SHEET_NS}}}tableColumn")
        template.attrib["id"] = str(index)
        template.attrib["name"] = str(header[index - 1] if index - 1 < len(header) else f"Column{index}") or " "
        table_columns.append(template)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _shared_strings_xml(strings: list[str]) -> bytes:
    root = ET.Element(f"{{{SHEET_NS}}}sst", {"count": str(len(strings)), "uniqueCount": str(len(strings))})
    for value in strings:
        si = ET.SubElement(root, f"{{{SHEET_NS}}}si")
        text = ET.SubElement(si, f"{{{SHEET_NS}}}t")
        if value != value.strip() or value == "":
            text.attrib["{http://www.w3.org/XML/1998/namespace}space"] = "preserve"
        text.text = value
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _matrix_strings(matrix: list[list[Any]]) -> list[str]:
    output: list[str] = []
    seen: set[str] = set()
    for row_index, row in enumerate(matrix, start=1):
        for col_index, value in enumerate(row, start=1):
            if row_index != 1 and col_index != 1:
                continue
            text = "" if value is None else str(value)
            if text not in seen:
                seen.add(text)
                output.append(text)
    return output


def update_embedded_workbook_openpyxl(
    workbook_bytes: bytes,
    sheet_name: str,
    matrix: list[list[Any]],
) -> bytes:
    workbook = openpyxl.load_workbook(BytesIO(workbook_bytes))
    worksheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook.worksheets[0]
    row_count = len(matrix)
    col_count = max((len(row) for row in matrix), default=0)
    if row_count == 0 or col_count == 0:
        raise ValueError("Matriz vazia; nada para gravar no workbook embutido.")

    for row_index, row in enumerate(matrix, start=1):
        for col_index in range(1, col_count + 1):
            value = row[col_index - 1] if col_index - 1 < len(row) else None
            cell = worksheet.cell(row=row_index, column=col_index)
            if col_index == 1 or row_index == 1:
                cell.number_format = "@"
                cell.value = "" if value is None else str(value)
            else:
                cell.value = _number_or_original(value)

    _resize_first_table_if_present(worksheet, row_count, col_count)
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _resize_first_table_if_present(worksheet: Any, row_count: int, col_count: int) -> None:
    if not worksheet.tables:
        return
    table = next(iter(worksheet.tables.values()))
    if not isinstance(table, Table):
        return
    table.ref = f"A1:{get_column_letter(col_count)}{row_count}"
    if table.autoFilter is not None:
        table.autoFilter.ref = table.ref


def _workbook_matrix(plan: TransformPlan) -> list[list[Any]]:
    if plan.orientation_ppt == "series_rows_categories_columns":
        matrix = [[" ", *plan.categories]]
        for index, series_name in enumerate(plan.series):
            values = plan.values[index] if index < len(plan.values) else []
            matrix.append([series_name, *values])
        return matrix

    matrix = [[" ", *plan.series]]
    for index, category in enumerate(plan.categories):
        values = plan.values[index] if index < len(plan.values) else []
        matrix.append([category, *values])
    return matrix


def _number_or_original(value: Any) -> Any:
    if value is None or isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    if not text:
        return ""
    normalized = text.replace("%", "").strip()
    if "," in normalized and "." in normalized:
        normalized = normalized.replace(".", "").replace(",", ".")
    else:
        normalized = normalized.replace(",", ".")
    try:
        number = float(normalized)
    except ValueError:
        return value
    if "%" in text:
        return number / 100 if abs(number) > 1 else number
    return number


def _xml_number_text(value: Any) -> str:
    if value is None or value == "":
        return "0"
    if isinstance(value, (int, float)):
        return f"{float(value):.15g}"
    parsed = _number_or_original(value)
    if isinstance(parsed, (int, float)):
        return f"{float(parsed):.15g}"
    return "0"


def _excel_col(index: int) -> str:
    letters = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def _join_xlsx_path(base_part: str, target: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    base_dir = str(Path(base_part).parent).replace("\\", "/")
    parts: list[str] = []
    for part in f"{base_dir}/{target}".split("/"):
        if part in {"", "."}:
            continue
        if part == "..":
            if parts:
                parts.pop()
            continue
        parts.append(part)
    return "/".join(parts)


def _rels_path(part_name: str) -> str:
    path = Path(part_name)
    return str(path.parent / "_rels" / f"{path.name}.rels").replace("\\", "/")


def _select_target(
    pptx_path: Path,
    slide_number: int,
    shape_name: str,
    target_id: str,
    title_hint: str,
) -> PptTarget:
    targets = [
        target
        for target in discover_ppt_targets(pptx_path, numeric_only=False, include_text_shapes=False)
        if target.slide_number == slide_number and target.object_type == "chart"
    ]
    if not targets:
        raise ValueError(f"Nenhum grafico encontrado no slide {slide_number}.")
    if target_id:
        matched = [target for target in targets if target.target_id == target_id]
        if matched:
            return matched[0]
    if shape_name:
        matched = [target for target in targets if target.shape_name == shape_name]
        if matched:
            return matched[0]
    if title_hint:
        hint = title_hint.casefold()
        ranked = sorted(
            targets,
            key=lambda target: (
                hint in target.nearby_text.casefold(),
                any(label.casefold() in {"cristal", "diamante", "diamante +"} for label in target.expected_series),
            ),
            reverse=True,
        )
        return ranked[0]
    if len(targets) == 1:
        return targets[0]
    raise ValueError(
        "Mais de um grafico encontrado. Informe --shape-name ou --target-id. "
        + ", ".join(f"{target.target_id}/{target.shape_name}" for target in targets)
    )


def _read_datasource(args: argparse.Namespace) -> tuple[bytes, str]:
    if args.xlsx:
        path = Path(args.xlsx)
        return path.read_bytes(), path.name
    if not args.datasource_zip or not args.datasource_member:
        raise ValueError("Informe --xlsx ou a dupla --datasource-zip/--datasource-member.")
    with ZipFile(args.datasource_zip) as zf:
        return zf.read(args.datasource_member), args.datasource_member


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Teste Linux/Fargate: edita workbook Excel embutido com openpyxl e atualiza cache XML "
            "do grafico sem usar COM nem python-pptx chart.replace_data()."
        )
    )
    parser.add_argument("--pptx", required=True, help="PPTX template de entrada.")
    parser.add_argument("--xlsx", default="", help="XLSX datasource direto.")
    parser.add_argument("--datasource-zip", default="", help="ZIP contendo o XLSX datasource.")
    parser.add_argument("--datasource-member", default="", help="Caminho do XLSX dentro do ZIP.")
    parser.add_argument("--output", required=True, help="PPTX de saida.")
    parser.add_argument("--slide", type=int, default=4, help="Numero do slide a testar.")
    parser.add_argument("--shape-name", default=DEFAULT_SHAPE_NAME, help="Nome do objeto grafico no PPT.")
    parser.add_argument("--target-id", default="", help="ID interno alvo, se ja existir.")
    parser.add_argument("--title-hint", default="Plano de Crescimento", help="Texto auxiliar para escolher grafico.")
    parser.add_argument(
        "--writer",
        choices=["xml", "openpyxl", "sheet-values", "raw-values"],
        default="xml",
        help=(
            "xml edita o .xlsx embutido por XML; sheet-values altera so valores numericos existentes via XML parser; "
            "raw-values troca apenas o texto de <v> nas celulas numericas; openpyxl e o teste comparativo anterior."
        ),
    )
    parser.add_argument(
        "--skip-chart-cache",
        action="store_true",
        help="Edita somente o workbook embutido e deixa ppt/charts/chartX.xml intacto para isolar a causa do erro.",
    )
    return parser.parse_args()


if __name__ == "__main__":
    main()
